#!/usr/bin/env python3
"""
build_intake.py  --  RedIQ-replacement underwriting intake workbook writer.

Reads a raw T12 + rent roll (+ optional HelloData CSV) and produces a single
formatted .xlsx with an EDITABLE chart-of-accounts categorization whose OS
Summary rolls up LIVE via SUMIFS.  Change a code in 'T12 Categorized' col A and
the standardized OS Summary updates automatically -- then copy/paste the dump
tabs straight into the TMG acquisition model.

Tabs:  Dashboard | T12 Categorized | OS Summary | Rent Roll (One-Line) |
       Unit Mix | HelloData (opt) | Reconciliation | Trends | Codes

CLI:
    python build_intake.py --t12 T12.xlsx --rr RentRoll.xlsx \
        [--hd hello.csv] [--name "Property"] [--out out.xlsx]
"""
import argparse
import datetime as _dt
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

import intake_lib as il
import account_map as am

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
FONT = "Arial"
BLUE = "0000FF"      # editable inputs (the code column)
BLACK = "000000"     # same-sheet formulas / data
GREEN = "008000"     # cross-sheet links (SUMIFS / VLOOKUP)
GREY = "808080"
WHITE = "FFFFFF"
NAVY = "1F3864"
HDR_FILL = "1F3864"
SUB_FILL = "D9E1F2"
EDIT_FILL = "FFF2CC"   # soft yellow → "edit here"
SECT_FILL = "8EAADB"
ZEBRA = "F2F5FB"

MONEY = '#,##0;(#,##0);"-"'
MONEY2 = '#,##0.00;(#,##0.00);"-"'
INT = '#,##0;(#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'
DATEF = "mm/dd/yyyy"
MONF = "mmm yyyy"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPLINE = Border(top=Side(style="thin", color="404040"))
DBLTOP = Border(top=Side(style="double", color="404040"))


def _f(bold=False, color=BLACK, size=10, italic=False):
    return Font(name=FONT, bold=bold, color=color, size=size, italic=italic)


def _fill(hexc):
    return PatternFill("solid", fgColor=hexc)


def _set(ws, r, c, val, *, font=None, fill=None, fmt=None, align=None,
         wrap=False, border=None):
    cell = ws.cell(r, c, val)
    cell.font = font or _f()
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align or wrap:
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        cell.border = border
    return cell


def _title_block(ws, title, subtitle=""):
    _set(ws, 1, 1, title, font=_f(bold=True, color=NAVY, size=15))
    if subtitle:
        _set(ws, 2, 1, subtitle, font=_f(color=GREY, size=10, italic=True))


# ===========================================================================
# TAB: T12 Categorized  (full T12, editable code column, Total per row)
# ===========================================================================
def _t12_patterns(st):
    """A few auto-detected trend observations for the T12 Categorized notes block."""
    n = st.n_months
    notes = []
    if n < 3:
        return notes
    a0 = st.code_total("Rentinc", 0) + st.code_total("ltl", 0)
    a1 = st.code_total("Rentinc", n - 1) + st.code_total("ltl", n - 1)
    v0, v1 = st.code_total("vac", 0), st.code_total("vac", n - 1)
    if a0 and a1:
        occ0, occ1 = 1 + v0 / a0, 1 + v1 / a1
        if occ1 - occ0 > 0.10:
            notes.append(f"Lease-up: economic occupancy ramped from {occ0*100:.0f}% "
                         f"({st.month_labels[0]}) to {occ1*100:.0f}% ({st.month_labels[-1]}) — "
                         f"early-period revenue reflects fill, not run-rate; trust T3 over T12.")
    ltl0, ltl1 = st.code_total("ltl", 0), st.code_total("ltl", n - 1)
    if abs(ltl0) > 1 and abs(ltl1 - ltl0) > abs(ltl0) * 0.5:
        notes.append(f"Loss-to-Lease moved from ${ltl0:,.0f} ({st.month_labels[0]}) to "
                     f"${ltl1:,.0f} ({st.month_labels[-1]}) as contract rents reset toward market.")
    return notes


def write_t12_categorized(ws, st, code_list_rows: int):
    """One clean chart of accounts across the full union of months. Every line appears
    once (merged across statements); each month's value comes from the freshest statement
    that owns it, so a SUMIFS by code reproduces the resolved series with no double-count.
    Column A (code) stays editable; section bands separate income / expenses / non-op."""
    n = st.n_months
    months = st.months
    MS = 4                                       # month columns start at D
    total_col = MS + n
    last_col = total_col
    multi = len(st.files) > 1

    _set(ws, 1, 1, "Code", font=_f(bold=True, color=WHITE), align="center")
    ws.cell(1, 1).fill = _fill("C9A227")
    _set(ws, 1, 2, "Category", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center")
    _set(ws, 1, 3, "Line Item (raw GL)", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))
    for k in range(n):
        d = _dt.datetime(months[k][1][0], months[k][1][1], 1)
        _set(ws, 1, MS + k, d, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL),
             fmt=MONF, align="center")
    _set(ws, 1, total_col, ("T12 Total" if n > 12 else "Total"),
         font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center", wrap=True)
    # optional sourcing note when multiple statements are stitched
    if multi:
        src = " · ".join(f"{f.label} ({f.t12.month_labels[0]}\u2013{f.t12.month_labels[-1]})"
                         for f in st.files)
        _set(ws, 2, 3, f"Stitched from: {src}. Each month is sourced from the freshest "
                       f"statement covering it; blank cells = that line not itemized in the "
                       f"statement owning that month.", font=_f(color=GREY, italic=True, size=8))

    OUTLIER_FILL = "FCE4D6"                      # light amber for one-off spike cells
    notable, mlabels = [], st.month_labels
    r = 3 if multi else 2
    for item in il.unified_lines(st):
        if item["type"] == "header":
            _set(ws, r, 3, item["name"], font=_f(bold=True, color=NAVY))
            for c in range(1, last_col + 1):
                ws.cell(r, c).fill = _fill(SUB_FILL)
            r += 1
            continue
        _set(ws, r, 1, item["code"], font=_f(bold=True, color=BLUE), align="center")
        _set(ws, r, 2, f'=IFERROR(VLOOKUP($A{r},Codes!$A:$B,2,FALSE()),"")', font=_f(color=GREEN))
        _set(ws, r, 3, item["name"], font=_f(color=BLACK))
        vals = item["values"]
        nz = sorted(abs(v) for v in vals if v is not None and abs(v) > 1e-9)
        med = nz[len(nz) // 2] if nz else 0.0
        for k, v in enumerate(vals):
            if v is None:
                continue
            is_out = med > 0 and len(nz) >= 4 and abs(v) > 3.5 * med and abs(v) > 2000
            _set(ws, r, MS + k, v, font=_f(color=BLACK), fmt=MONEY,
                 fill=(_fill(OUTLIER_FILL) if is_out else None))
            if is_out:
                notable.append((abs(v) / med, item["name"], mlabels[k], v))
        _set(ws, r, total_col,
             f"=SUM({get_column_letter(MS + max(0, n - 12))}{r}:{get_column_letter(MS + n - 1)}{r})",
             font=_f(bold=True, color=BLACK), fmt=MONEY)
        r += 1
    last_row = r - 1

    # ---- outliers & notable patterns ----
    r = last_row + 2
    _set(ws, r, 1, "Outliers & Notable Patterns  (amber month cells = > 3.5× that line's median)",
         font=_f(bold=True, color=NAVY))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=min(total_col, 12))
    r += 1
    for note in _t12_patterns(st):
        _set(ws, r, 1, "•  " + note, font=_f(color="843C0C"), wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=min(total_col, 12))
        ws.row_dimensions[r].height = 28
        r += 1
    for ratio, name, mlab, v in sorted(notable, reverse=True)[:8]:
        _set(ws, r, 1, f"•  {name}: {mlab} = ${v:,.0f}  ({ratio:.1f}× the line's median) "
                       f"— one-off / reclass worth checking.", font=_f(color="843C0C"), wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=min(total_col, 12))
        r += 1

    dv = DataValidation(type="list", formula1="CodeList", allow_blank=True, showDropDown=False)
    dv.error = "Pick a standardized code from the list (or clear to exclude)."
    dv.errorTitle = "Invalid code"
    dv.prompt = "Edit this code to re-map the line. OS Summary updates automatically."
    dv.promptTitle = "Editable code"
    ws.add_data_validation(dv)
    dv.add(f"A2:A{last_row}")

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 42
    for k in range(n):
        ws.column_dimensions[get_column_letter(MS + k)].width = 12
    ws.column_dimensions[get_column_letter(total_col)].width = 13
    ws.freeze_panes = "D2"
    ws.sheet_view.showGridLines = False
    return last_row, total_col, MS, n


# ===========================================================================
# TAB: OS Summary / Operating History  (RedIQ Overview template, live SUMIFS)
# ===========================================================================
def write_os_grid(ws, st, prop_name, show_idx, *, title, subtitle):
    """RedIQ-style standardized statement. `show_idx` is the list of union-month
    indices to display (e.g. last 12 for the model paste, or all for the full
    history). Each shown column's SUMIFS points at the matching 'T12 Categorized'
    month column so the rollup stays live and editable."""
    n = len(show_idx)
    months = st.months
    TC = "'T12 Categorized'"
    MS = 4                                    # T12 Categorized month columns start at D

    def os_col(k):
        return 6 + k                          # display column (F=6)

    def tc_col(k):
        return get_column_letter(MS + show_idx[k])   # matching source column

    _set(ws, 1, 2, prop_name, font=_f(bold=True, color=NAVY, size=13))
    _set(ws, 2, 2, title, font=_f(color=GREY, italic=True))
    _set(ws, 3, 2, subtitle, font=_f(color=GREY, italic=True, size=9))

    for cc in (3, 4, 5):
        _set(ws, 4, cc, "Annual", font=_f(bold=True, color=GREY, size=9), align="center")
    for k in range(n):
        _set(ws, 4, os_col(k), "Monthly", font=_f(bold=True, color=WHITE, size=9),
             fill=_fill(HDR_FILL), align="center")
        d = _dt.datetime(months[show_idx[k]][1][0], months[show_idx[k]][1][1], 1)
        _set(ws, 5, os_col(k), d, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL),
             fmt=MONF, align="center")
    _set(ws, 6, 1, "Code", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center")
    _set(ws, 6, 2, "Effective Gross Revenue", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))

    def sumifs(row, anchor):
        for k in range(n):
            col = tc_col(k)
            ws.cell(row, os_col(k), f"=SUMIFS({TC}!{col}:{col},{TC}!$A:$A,$A{anchor})")
            c = ws.cell(row, os_col(k)); c.font = _f(color=GREEN); c.number_format = MONEY

    def subtotal(row, fn, *, bold=True, border=TOPLINE, fill=None):
        for k in range(n):
            L = get_column_letter
            ws.cell(row, os_col(k), fn(L(os_col(k))))
            c = ws.cell(row, os_col(k)); c.font = _f(bold=bold, color=BLACK)
            c.number_format = MONEY; c.border = border
            if fill:
                c.fill = fill

    r = 7
    for code, label in am.REVENUE_CODES:
        _set(ws, r, 1, code, font=_f(bold=True, color=BLACK), align="center")
        _set(ws, r, 2, label, font=_f(color=BLACK)); sumifs(r, r); r += 1
    _set(ws, 20, 2, "Effective Gross Revenue", font=_f(bold=True, color=NAVY), fill=_fill(SUB_FILL))
    ws.cell(20, 1).fill = _fill(SUB_FILL)
    subtotal(20, lambda c: f"=SUM({c}7:{c}19)", fill=_fill(SUB_FILL))

    _set(ws, 22, 2, "Operating Expenses", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))
    ws.cell(22, 1).fill = _fill(HDR_FILL)
    r = 23
    for code, label in am.EXPENSE_CODES:
        _set(ws, r, 1, code, font=_f(bold=True, color=BLACK), align="center")
        _set(ws, r, 2, label, font=_f(color=BLACK)); sumifs(r, r); r += 1
    _set(ws, 42, 2, "Operating Expenses", font=_f(bold=True, color=NAVY), fill=_fill(SUB_FILL))
    ws.cell(42, 1).fill = _fill(SUB_FILL)
    subtotal(42, lambda c: f"=SUM({c}23:{c}41)", fill=_fill(SUB_FILL))

    _set(ws, 44, 2, "Net Operating Income", font=_f(bold=True, color=WHITE, size=11), fill=_fill(NAVY))
    ws.cell(44, 1).fill = _fill(NAVY)
    subtotal(44, lambda c: f"={c}20-{c}42", border=DBLTOP, fill=_fill(NAVY))
    for k in range(n):
        ws.cell(44, os_col(k)).font = _f(bold=True, color=WHITE, size=11)

    _set(ws, 46, 2, "Non-Operating Revenue", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))
    ws.cell(46, 1).fill = _fill(HDR_FILL)
    r = 47
    for code, label in am.NONOP_REV_CODES:
        _set(ws, r, 1, code, font=_f(bold=True, color=BLACK), align="center")
        _set(ws, r, 2, label, font=_f(color=BLACK)); sumifs(r, r); r += 1
    _set(ws, 48, 2, "Non-Operating Revenue", font=_f(bold=True, color=NAVY), fill=_fill(SUB_FILL))
    ws.cell(48, 1).fill = _fill(SUB_FILL)
    subtotal(48, lambda c: f"=SUM({c}47:{c}47)", fill=_fill(SUB_FILL))

    _set(ws, 50, 2, "Non-Operating Expenses", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))
    ws.cell(50, 1).fill = _fill(HDR_FILL)
    r = 51
    for code, label in am.NONOP_EXP_CODES:
        _set(ws, r, 1, code, font=_f(bold=True, color=BLACK), align="center")
        _set(ws, r, 2, label, font=_f(color=BLACK)); sumifs(r, r); r += 1
    _set(ws, 73, 2, "Non-Operating Expenses", font=_f(bold=True, color=NAVY), fill=_fill(SUB_FILL))
    ws.cell(73, 1).fill = _fill(SUB_FILL)
    subtotal(73, lambda c: f"=SUM({c}51:{c}72)", fill=_fill(SUB_FILL))

    _set(ws, 75, 2, "Net Income", font=_f(bold=True, color=WHITE, size=11), fill=_fill(NAVY))
    ws.cell(75, 1).fill = _fill(NAVY)
    subtotal(75, lambda c: f"={c}44+{c}48-{c}73", border=DBLTOP, fill=_fill(NAVY))
    for k in range(n):
        ws.cell(75, os_col(k)).font = _f(bold=True, color=WHITE, size=11)

    _set(ws, 77, 2, f"Generated {_dt.date.today():%m/%d/%Y} \u2014 live SUMIFS over 'T12 Categorized' code column",
         font=_f(color=GREY, italic=True, size=9))

    for col, w in (("A", 11), ("B", 32), ("C", 9), ("D", 9), ("E", 9)):
        ws.column_dimensions[col].width = w
    for k in range(n):
        ws.column_dimensions[get_column_letter(6 + k)].width = 12
    ws.freeze_panes = "F7"
    ws.sheet_view.showGridLines = False


# ===========================================================================
# TAB: Rent Roll (One-Line)
# ===========================================================================
RR_HEADERS = ["Unit No.", "Floor Plan", "Net sf", "Bed", "Bath", "Lease Type",
              "Occupancy Status", "Market Rent", "Contractual Rent",
              "Net Effective Rent", "Lease Start Date", "Lease Expiration",
              "Move In Date"]


def write_rent_roll(ws, rr: il.RentRoll, hd):
    for c, h in enumerate(RR_HEADERS, 1):
        _set(ws, 1, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL),
             align="center", wrap=True)
    # charge-code columns, GROUPED so it is explicit which charges roll into contract rent.
    # one blank spacer column between the core fields (A-M) and the charge block.
    ncore = len(RR_HEADERS)                       # 13 -> A..M
    sched_tot, act_tot = {}, {}
    for u in rr.units:
        for cc, amt in u.charges.items():
            sched_tot[cc] = sched_tot.get(cc, 0.0) + (amt or 0.0)
        for cc, amt in u.actual_charges.items():
            act_tot[cc] = act_tot.get(cc, 0.0) + (amt or 0.0)
    by_mag = lambda d: [cc for cc, t in sorted(d.items(), key=lambda kv: -abs(kv[1])) if abs(t) > 0.005]
    sched_codes = by_mag(sched_tot)
    # Partition the SCHEDULED charges into contract-rent vs other-income.
    contract_cc, other_cc = [], []
    for cc in sched_codes:
        _code, is_cr, _rec = am.categorize_charge(cc)
        (contract_cc if is_cr else other_cc).append(cc)
    # RUBS / utility recoveries are billed in arrears off metered usage, so they carry $0
    # scheduled and only appear in ACTUAL charges. Surface them (and only them — not the
    # one-time actual noise like late/termination/referral fees) from actuals, keyed by a
    # RUBS code (RWS/RT/RF) with ~0 scheduled.
    RUBS = {"RWS", "RT", "RF"}
    rubs_cc = []
    for cc in by_mag(act_tot):
        code, _is_cr, _rec = am.categorize_charge(cc)
        if code in RUBS and abs(sched_tot.get(cc, 0.0)) < 1.0 and act_tot.get(cc, 0.0) > 0.005:
            rubs_cc.append(cc)

    # groups: (banner, banner_fill, subhdr_fill, charge-list, source-dict-attr)
    groups = []
    if contract_cc:
        groups.append(("↓ IN Contractual Rent  (scheduled $)", "548235", "C6E0B4", contract_cc, "charges"))
    if other_cc:
        groups.append(("↓ Other recurring — Other Income, NOT in contract rent  (scheduled $)",
                       SECT_FILL, "D9E1F2", other_cc, "charges"))
    if rubs_cc:
        groups.append(("↓ RUBS / utility recoveries — ACTUAL $  ($0 scheduled; billed in arrears)",
                       "C9A227", "FFE699", rubs_cc, "actual_charges"))

    all_cols = [(cc, attr, sub) for (_b, _bf, sub, lst, attr) in groups for cc in lst]
    has_block = bool(all_cols)
    spacer = ncore + 1                            # N (blank)
    cstart = ncore + 2                            # O -> first charge column
    if has_block:
        col = cstart
        for banner, bfill, _sub, lst, _attr in groups:
            _set(ws, 1, col, banner, font=_f(bold=True, color=WHITE), fill=_fill(bfill))
            if len(lst) > 1:
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(lst) - 1)
            col += len(lst)
        for j, (cc, _attr, sub) in enumerate(all_cols):
            _set(ws, 2, cstart + j, cc, font=_f(bold=True, color=NAVY), fill=_fill(sub),
                 align="center", wrap=True)

    data_start = 3 if has_block else 2
    if has_block:
        for c in range(1, ncore + 1):
            ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)

    r = data_start
    for u in rr.units:
        bed, bath, _src = il.infer_bed_bath(u.floorplan, hd)
        zeb = _fill(ZEBRA) if (r - data_start) % 2 == 0 else None
        vals = [u.unit, u.floorplan, u.sqft, bed, bath, u.lease_type, u.occupancy,
                u.market_rent, u.contract_rent, u.net_effective,
                u.lease_start, u.lease_end, u.move_in]
        for c, v in enumerate(vals, 1):
            fmt = None
            al = None
            if c in (3,):
                fmt = INT; al = "right"
            elif c in (4, 5):
                fmt = "0"; al = "center"
            elif c in (8, 9, 10):
                fmt = MONEY; al = "right"
            elif c in (11, 12, 13):
                fmt = DATEF; al = "center"
            elif c in (1, 2, 6, 7):
                al = "left" if c in (1, 2) else "center"
            _set(ws, r, c, v, font=_f(), fmt=fmt, align=al, fill=zeb)
        for j, (cc, attr, _sub) in enumerate(all_cols):
            amt = getattr(u, attr).get(cc)
            _set(ws, r, cstart + j, (amt if amt not in (None, 0) else None),
                 font=_f(), fmt=MONEY, align="right", fill=zeb)
        r += 1
    last = r - 1
    # totals row
    _set(ws, last + 1, 1, "TOTAL / AVG", font=_f(bold=True, color=NAVY))
    _set(ws, last + 1, 3, f"=SUM(C{data_start}:C{last})", font=_f(bold=True, color=NAVY), fmt=INT, align="right")
    _set(ws, last + 1, 8, f"=SUM(H{data_start}:H{last})", font=_f(bold=True, color=NAVY), fmt=MONEY, align="right")
    _set(ws, last + 1, 9, f"=SUM(I{data_start}:I{last})", font=_f(bold=True, color=NAVY), fmt=MONEY, align="right")
    _set(ws, last + 1, 10, f"=SUM(J{data_start}:J{last})", font=_f(bold=True, color=NAVY), fmt=MONEY, align="right")
    for j in range(len(all_cols)):
        L = get_column_letter(cstart + j)
        _set(ws, last + 1, cstart + j, f"=SUM({L}{data_start}:{L}{last})",
             font=_f(bold=True, color=NAVY), fmt=MONEY, align="right")
    for c in range(1, cstart + len(all_cols)):
        ws.cell(last + 1, c).border = DBLTOP

    # contract-rent reconciliation note under the totals
    if has_block:
        note = ("Contractual Rent (col I) = the green charges above — base Rent + Amenity Rent "
                "(scheduled). Other Income & RUBS recoveries are NOT in contract rent. RUBS is "
                "shown from ACTUAL charges because it has no scheduled value (billed in arrears).")
        _set(ws, last + 3, 1, note, font=_f(italic=True, color=GREY, size=9))

    widths = [11, 11, 9, 6, 6, 12, 16, 13, 14, 14, 14, 14, 14]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    if has_block:
        ws.column_dimensions[get_column_letter(spacer)].width = 2
        for j in range(len(all_cols)):
            ws.column_dimensions[get_column_letter(cstart + j)].width = 13
    ws.freeze_panes = "B" + str(data_start)
    ws.sheet_view.showGridLines = False
    return last


# ===========================================================================
# Unit-mix block (embedded in the Dashboard, not a standalone tab)
# ===========================================================================
def write_unit_mix_block(ws, mix, has_hd, r0, c0):
    """Render the unit-mix table at origin (r0, c0); columns are relative to c0.
    New-lease (last 5) and HelloData T90 / T365 executed are market-rent indicators;
    HD90 YoY compares the trailing-90-day executed rent to the same window a year ago.
    Totals are Python-computed mix-weighted values. Returns the last row used."""
    def col(j):
        return c0 + j

    def CL(j):
        return get_column_letter(c0 + j)

    hdr = ["Floor Plan", "Bed", "Bath", "Units", "Occ", "Vac", "Avg SF",
           "Avg Market Rent\n(per RR — not a\nmkt signal)", "Avg Contract\nRent", "New\nLeases", "Renew\nLeases",
           "Avg New-Lease\n(last 5)", "New #1", "New #2", "New #3", "New #4", "New #5",
           "HD T90\nAsking", "HD T90\nEffective", "HD T365\nAsking", "HD T365\nEffective",
           "HD90 YoY\nAsking", "HD90 YoY\nEffective", "Bed/Bath\nSrc", "HD Plan\n(website)"]
    # j: 0 plan,1 bed,2 bath,3 units,4 occ,5 vac,6 avgSF,7 avgMkt,8 avgCon,9 new,10 ren,
    # 11 avgNew5,12-16 new1-5,17 t90ask,18 t90eff,19 t365ask,20 t365eff,21 yoyAsk,22 yoyEff,23 src,24 HDname
    _set(ws, r0, c0, "UNIT MIX  \u2014  mix-weighted market-rent indicators",
         font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
    for j in range(1, len(hdr)):
        ws.cell(r0, col(j)).fill = _fill(SECT_FILL)
    hr = r0 + 1
    for j, h in enumerate(hdr):
        _set(ws, hr, col(j), h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL),
             align="center", wrap=True)
    r = hr + 1
    for m in mix:
        zeb = _fill(ZEBRA) if (r - hr) % 2 == 0 else None

        def put(j, v, **kw):
            _set(ws, r, col(j), v, fill=zeb, **kw)
        put(0, m.plan, font=_f(bold=True))
        put(1, m.bed, fmt="0", align="center")
        put(2, m.bath, fmt="0", align="center")
        put(3, m.units, fmt="0", align="center")
        put(4, m.occ, fmt="0", align="center")
        put(5, m.vac, fmt="0", align="center")
        put(6, round(m.avg_sqft), fmt=INT, align="right")
        put(7, round(m.avg_market), fmt=MONEY, align="right")
        put(8, round(m.avg_contract), fmt=MONEY, align="right")
        put(9, m.new_count, fmt="0", align="center")
        put(10, m.renewal_count, fmt="0", align="center")
        put(11, round(m.avg_new_last5) if m.avg_new_last5 else None,
            font=_f(bold=True, color="843C0C"), fmt=MONEY, align="right")
        for k in range(5):
            put(12 + k, m.new_last5_rents[k] if k < len(m.new_last5_rents) else None,
                fmt=MONEY, align="right")
        if has_hd:
            put(17, round(m.t90_ask) if m.t90_ask else None, fmt=MONEY, align="right")
            put(18, round(m.t90_eff) if m.t90_eff else None,
                font=_f(bold=True, color=GREEN), fmt=MONEY, align="right")
            put(19, round(m.t365_ask) if m.t365_ask else None, fmt=MONEY, align="right")
            put(20, round(m.t365_eff) if m.t365_eff else None, fmt=MONEY, align="right")
            put(21, m.yoy_ask if m.yoy_ask is not None else None, fmt=PCT, align="right")
            put(22, m.yoy_eff if m.yoy_eff is not None else None, fmt=PCT, align="right")
        put(23, m.bedbath_source, font=_f(color=GREY, size=9), align="center")
        put(24, m.hd_names or None, font=_f(color=NAVY, size=9), align="left")
        r += 1

    # totals (mix-weighted, computed in Python so blanks/None are handled cleanly)
    def wavg(get, wt=lambda m: m.units):
        num = sum(wt(m) * get(m) for m in mix if get(m))
        den = sum(wt(m) for m in mix if get(m))
        return (num / den) if den else 0

    def wyoy(get):
        num = sum(m.units * get(m) for m in mix if get(m) is not None)
        den = sum(m.units for m in mix if get(m) is not None)
        return (num / den) if den else None

    def tot(j, v, **kw):
        color = kw.pop("color", NAVY)
        _set(ws, r, col(j), v, font=_f(bold=True, color=color), **kw)
    _set(ws, r, c0, "TOTAL / AVG", font=_f(bold=True, color=NAVY))
    tot(3, sum(m.units for m in mix), fmt="0", align="center")
    tot(4, sum(m.occ for m in mix), fmt="0", align="center")
    tot(5, sum(m.vac for m in mix), fmt="0", align="center")
    tot(6, round(wavg(lambda m: m.avg_sqft)), fmt=INT, align="right")
    tot(7, round(wavg(lambda m: m.avg_market)), fmt=MONEY, align="right")
    tot(8, round(wavg(lambda m: m.avg_contract)), fmt=MONEY, align="right")
    tot(9, sum(m.new_count for m in mix), fmt="0", align="center")
    tot(10, sum(m.renewal_count for m in mix), fmt="0", align="center")
    _an = wavg(lambda m: m.avg_new_last5, lambda m: m.new_count)
    tot(11, round(_an) if _an else None, color="843C0C", fmt=MONEY, align="right")
    if has_hd:
        for j, fld in ((17, "t90_ask"), (18, "t90_eff"), (19, "t365_ask"), (20, "t365_eff")):
            v = wavg(lambda m, f=fld: getattr(m, f))
            tot(j, round(v) if v else None, color=(GREEN if j == 18 else NAVY), fmt=MONEY, align="right")
        tot(21, wyoy(lambda m: m.yoy_ask), fmt=PCT, align="right")
        tot(22, wyoy(lambda m: m.yoy_eff), fmt=PCT, align="right")
    for j in range(len(hdr)):
        ws.cell(r, col(j)).border = DBLTOP

    w = [12, 5, 5, 6, 5, 5, 8, 11, 12, 7, 7, 12, 9, 9, 9, 9, 9, 10, 10, 10, 10, 10, 10, 9]
    for j, wd in enumerate(w):
        ws.column_dimensions[CL(j)].width = wd
    return r


# ===========================================================================
# TAB: Lease Trend  (market-rent trajectory + seasonality + new-lease detail)
# ===========================================================================
def write_lease_trend(ws, lt, st, tr, rr: il.RentRoll, has_hd=True):
    _title_block(ws, "Lease Trend & Operating History (monthly)",
                 "True market rent from HelloData executed leases (mix-weighted by floor plan) + "
                 "new-lease signings; loss-to-lease backed into from T12 AGPR; concessions compared "
                 "T12 vs HelloData. Months flow left-to-right; renewals excluded from market reads.")
    units = max(1, len(rr.units))
    axis = sorted(set(lt.months) | set(st.monthkeys))

    finmap, finidx = {}, {ym: i for i, ym in enumerate(st.monthkeys)}
    for ym, i in finidx.items():
        gpr = st.code_total("Rentinc", i); ltl = st.code_total("ltl", i)
        vac = st.code_total("vac", i); conc = st.code_total("conc", i)
        finmap[ym] = {"i": i, "gpr": gpr, "ltl": ltl, "vac": vac, "conc": conc, "agpr": gpr + ltl}

    r = 4

    def month_header(r):
        _set(ws, r, 1, "Month →", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))
        for j, ym in enumerate(axis):
            _set(ws, r, 2 + j, _dt.date(ym[0], ym[1], 1).strftime("%b %Y"),
                 font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center")
        return r + 1

    def section(r, label):
        _set(ws, r, 1, label, font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
        for j in range(len(axis)):
            ws.cell(r, 2 + j).fill = _fill(SECT_FILL)
        return r + 1

    def mrow(r, label, fn, fmt, bold=False):
        _set(ws, r, 1, label, font=_f(bold=bold, color=(NAVY if bold else BLACK)))
        for j, ym in enumerate(axis):
            v = fn(ym)
            if v is not None and v != "":
                _set(ws, r, 2 + j, v, fmt=fmt, align="right")
        return r + 1

    r = month_header(r)
    r = section(r, "MARKET RENT — HelloData executed (mix-weighted) + new-lease signings")
    r = mrow(r, "HD Market Rent / unit (asking)", lambda ym: round(lt.hd_ask.get(ym, 0)) or None, MONEY, bold=True)
    r = mrow(r, "HD Effective Rent / unit", lambda ym: round(lt.hd_eff.get(ym, 0)) or None, MONEY)
    r = mrow(r, "HD Concession % (1 − eff/ask)", lambda ym: lt.hd_conc.get(ym) or None, PCT)
    r = mrow(r, "HD Executed Leases (#)", lambda ym: lt.hd_n.get(ym) or None, "0")
    r = mrow(r, "New-Lease Contract / unit (rent roll)", lambda ym: round(lt.new_rent.get(ym, 0)) or None, MONEY, bold=True)
    r = mrow(r, "New Leases (#)", lambda ym: lt.new_n.get(ym) or None, "0")

    r += 1
    r = section(r, "OCCUPANCY & RENT POSITION — from financial statements")
    r = mrow(r, "Physical Occupancy % (T12: 1 − vac/AGPR)",
             lambda ym: (lambda f: (1 + f["vac"] / f["agpr"]) if f and f["agpr"] else None)(finmap.get(ym)), PCT)
    r = mrow(r, "Contract Rent / unit — AGPR (T12)",
             lambda ym: (lambda f: round(f["agpr"] / units) if f else None)(finmap.get(ym)), MONEY, bold=True)
    r = mrow(r, "Market Rent / unit — GPR basis (T12, ref only)",
             lambda ym: (lambda f: round(f["gpr"] / units) if f else None)(finmap.get(ym)), MONEY)

    def ltl_u(ym):
        a = lt.hd_ask.get(ym, 0); f = finmap.get(ym)
        return round(a - f["agpr"] / units) if (a > 0 and f) else None

    def ltl_pct(ym):
        a = lt.hd_ask.get(ym, 0); f = finmap.get(ym)
        return ((a - f["agpr"] / units) / a) if (a > 0 and f) else None

    r = mrow(r, "Loss-to-Lease / unit (HD mkt − AGPR)", ltl_u, MONEY, bold=True)
    r = mrow(r, "Loss-to-Lease % (vs HD market)", ltl_pct, PCT)

    r += 1
    r = section(r, "CONCESSIONS — T12 (portfolio) vs HelloData (new-lease)")
    r = mrow(r, "T12 Concessions ($/mo)", lambda ym: (lambda f: round(f["conc"]) if (f and f["conc"]) else None)(finmap.get(ym)), MONEY)
    r = mrow(r, "T12 Concessions % of AGPR",
             lambda ym: (lambda f: (-f["conc"] / f["agpr"]) if (f and f["agpr"]) else None)(finmap.get(ym)), PCT, bold=True)
    r = mrow(r, "HD Concession % (new-lease)", lambda ym: lt.hd_conc.get(ym) or None, PCT, bold=True)

    r += 1
    r = section(r, "OPERATING TREND — from stitched T12")
    r = mrow(r, "Effective Gross Revenue", lambda ym: (lambda f: round(tr.egr_by_month[f["i"]]) if f else None)(finmap.get(ym)), MONEY)
    r = mrow(r, "Operating Expenses", lambda ym: (lambda f: round(tr.opex_by_month[f["i"]]) if f else None)(finmap.get(ym)), MONEY)
    r = mrow(r, "Net Operating Income", lambda ym: (lambda f: round(tr.noi_by_month[f["i"]]) if f else None)(finmap.get(ym)), MONEY, bold=True)

    r += 1
    r = section(r, "TRAILING ANNUALIZATIONS")
    for c, h in enumerate(["T12", "T6", "T3"]):
        _set(ws, r, 2 + c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center")
    r += 1
    for lbl, key in [("EGR (annualized)", "_EGR"), ("Operating Expenses", "_OPEX"), ("NOI", "_NOI")]:
        _set(ws, r, 1, lbl, font=_f(bold=(key == "_NOI")))
        for c, p in enumerate(["T12", "T6", "T3"]):
            v = tr.periods.get(p, {}).get(key)
            if v is not None:
                _set(ws, r, 2 + c, round(v), fmt=MONEY, align="right")
        r += 1

    r += 1
    notes = list(lt.notes)
    ov = [ym for ym in axis if ym in finmap]
    t12c = [(-finmap[ym]["conc"] / finmap[ym]["agpr"]) for ym in ov if finmap[ym]["agpr"]]
    hdc = [lt.hd_conc[ym] for ym in ov if lt.hd_conc.get(ym, 0) > 0]
    avg_t12c = (sum(t12c) / len(t12c) * 100) if t12c else 0
    avg_hdc = (sum(hdc) / len(hdc) * 100) if hdc else 0
    new_share = sum(1 for u in rr.units if il.classify_lease(u) == "new") / units * 100
    notes.append(
        f"Concessions by source: the T12 books concessions across the WHOLE portfolio "
        f"(~{avg_t12c:.1f}% of AGPR here), while HelloData's is per NEW lease (~{avg_hdc:.1f}% here). "
        f"The portfolio figure ≈ new-lease concession × the share of units on a recently-signed lease — "
        f"here ~{new_share:.0f}% of in-place leases are new (lease-up), so the two run close (little "
        f"dilution). On a stabilized deal with ~40–50% annual turnover, expect the T12 figure nearer half "
        f"the new-lease rate. "
        + ("NOTE: HelloData effective tracks asking in the most recent months (specials text not captured), "
           "so the T12 concession line is the more reliable recent concession signal." if avg_hdc < 1 else ""))
    if notes:
        _set(ws, r, 1, "Observations", font=_f(bold=True, color=NAVY))
        r += 1
        for note in notes:
            _set(ws, r, 1, "•  " + note, font=_f(color="843C0C"), wrap=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=min(2 + len(axis), 13))
            ws.row_dimensions[r].height = 42
            r += 1
    r += 2

    # T90 by base plan
    if has_hd and lt.plan_t90:
        _set(ws, r, 1, "Trailing-90-Day Executed by Floor Plan (HelloData)",
             font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
        for c in range(2, 8):
            ws.cell(r, c).fill = _fill(SECT_FILL)
        r += 1
        for c, h in enumerate(["Base Plan", "T90 Asking", "T90 Effective", "Concession %", "Exec (n)"], 1):
            _set(ws, r, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center", wrap=True)
        r += 1
        for bp in sorted(lt.plan_t90):
            d = lt.plan_t90[bp]
            _set(ws, r, 1, bp, font=_f(bold=True), align="center")
            _set(ws, r, 2, round(d["ask"]) or None, fmt=MONEY, align="right")
            _set(ws, r, 3, round(d["eff"]) or None, fmt=MONEY, align="right")
            _set(ws, r, 4, f'=IF(OR(B{r}="",B{r}=0),"-",1-C{r}/B{r})', fmt=PCT, align="right")
            _set(ws, r, 5, d["n"] or None, fmt="0", align="center")
            r += 1
        r += 1

    # last-5 new leases per plan (with dates / units / PSF)
    _set(ws, r, 1, "Last 5 New Leases by Floor Plan (rent roll)",
         font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
    for c in range(2, 8):
        ws.cell(r, c).fill = _fill(SECT_FILL)
    r += 1
    for c, h in enumerate(["Floor Plan", "Unit", "Lease Start", "Contract Rent", "Rent PSF (mo.)"], 1):
        _set(ws, r, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center", wrap=True)
    r += 1
    sqft_by_unit = {u.unit: u.sqft for u in rr.units}
    for plan in sorted(lt.recent_new):
        for u in lt.recent_new[plan]:
            sf = sqft_by_unit.get(u.unit, 0) or u.sqft
            psf = round(u.contract_rent / sf, 2) if sf else None
            _set(ws, r, 1, plan, font=_f(bold=True), align="center")
            _set(ws, r, 2, u.unit, align="center")
            _set(ws, r, 3, u.lease_start, fmt=DATEF, align="center")
            _set(ws, r, 4, round(u.contract_rent), fmt=MONEY, align="right")
            _set(ws, r, 5, psf, fmt=MONEY2, align="right")
            r += 1

    ws.column_dimensions["A"].width = 38           # metric labels / plan names
    for c in range(2, 30):                          # monthly columns (and sub-table cols)
        ws.column_dimensions[get_column_letter(c)].width = 9.5
    ws.freeze_panes = "B5"
    ws.sheet_view.showGridLines = False


# ===========================================================================
# TAB: HelloData pass-through  (matches model 'HD Dump' A..U)
# ===========================================================================
def write_hellodata(ws, hd: il.HelloData):
    out_headers = ["Property Name", "Floorplan", "Unit", "Floor", "Bedrooms",
                   "Bathrooms", "Partial Bathrooms", "SF", "Available On",
                   "Deposit", "Term", "AMI Level", "Last Asking Rent",
                   "Asking Rent PSF", "Last Effective Rent", "Effective Rent PSF",
                   "On Market Date", "Off Market Date", "Days on Market",
                   "Days Vacant", "Floorplan Mapped"]
    src_keys = ["Property Name", "Floorplan", "Unit", "Floor", "Bedrooms",
                "Bathrooms", "Partial Bathrooms", "SF", "Available On",
                "Deposit", "Term", "Est AMI Level", "Last Asking Rent",
                "Asking Rent PSF", "Last Effective Rent", "Effective Rent PSF",
                "On Market Date", "Off Market Date", "Days on Market", "Days Vacant"]
    for c, h in enumerate(out_headers, 1):
        _set(ws, 1, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL),
             align="center", wrap=True)
    r = 2
    date_keys = {"Available On", "On Market Date", "Off Market Date"}
    for row in hd.rows:
        for c, key in enumerate(src_keys, 1):
            v = row.get(key, "")
            if key in date_keys:
                d = il._to_date(v)
                if d is not None:
                    _set(ws, r, c, d, font=_f(), fmt=DATEF)
                    continue
                v = ""
            else:
                try:
                    v = float(v) if v not in ("", None) and key not in (
                        "Property Name", "Floorplan", "Unit",
                        "AMI Level", "Est AMI Level") else v
                except (ValueError, TypeError):
                    pass
            _set(ws, r, c, v, font=_f())
        _set(ws, r, 21, il._norm_plan(row.get("Floorplan", "")), font=_f(color=GREEN))
        r += 1
    for c in range(1, 22):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = "C2"
    ws.sheet_view.showGridLines = False


# ===========================================================================
# TAB: Reconciliation
# ===========================================================================
def write_reconciliation(ws, rec: il.Reconciliation, rr: il.RentRoll):
    _title_block(ws, "Reconciliation — Rent Roll ↔ T12",
                 f"Latest T12 month: {rec.latest_month}.  'In Contract Rent?' flags which charges roll into contract rent.")
    r = 4
    _set(ws, r, 1, "Control Tie-Outs (rent roll vs latest T12 month)", font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
    for c in range(2, 7):
        ws.cell(r, c).fill = _fill(SECT_FILL)
    r += 1
    for c, h in enumerate(["Measure", "Rent Roll", "T12 (latest mo. / ann.)", "Variance", "Var %", "Note"], 1):
        _set(ws, r, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), wrap=True, align="center")
    r += 1
    for ln in rec.lines:
        _set(ws, r, 1, ln.label, font=_f(bold=True))
        _set(ws, r, 2, ln.rr_value, fmt=MONEY, align="right")
        _set(ws, r, 3, ln.t12_value, fmt=MONEY, align="right")
        _set(ws, r, 4, f"=B{r}-C{r}", fmt=MONEY, align="right")
        _set(ws, r, 5, f'=IF(C{r}=0,"-",D{r}/C{r})', fmt=PCT, align="right")
        _set(ws, r, 6, ln.note, font=_f(color=GREY, size=9), wrap=True)
        r += 1

    nett = getattr(rec, "hd_fee_netting", None)
    if nett:
        r += 1
        _set(ws, r, 1, "HelloData Market Rent: Fee Netting (asking AND effective)",
             font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
        for c in range(2, 7):
            ws.cell(r, c).fill = _fill(SECT_FILL)
        r += 1
        for c, h in enumerate(["Measure", "Amount", "", "", "", "Note"], 1):
            _set(ws, r, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL),
                 align="center", wrap=True)
        r += 1
        fee = nett["applied"]
        net_ask = (nett["hd_raw"] - fee) if nett["hd_raw"] else 0.0
        eff_g = nett.get("hd_raw_eff", 0.0)
        net_eff = (eff_g - fee) if eff_g else 0.0
        net_rows = [
            ("HelloData T90 asking — gross (mix-wtd)", nett["hd_raw"],
             "What HelloData scrapes: the website 'Total Monthly', which can bundle mandatory fees."),
            ("HelloData T90 effective — gross (mix-wtd)", eff_g,
             "HD effective (net of concessions) — also carries the bundled fee, so it is netted too."),
            ("Rent-roll new-lease base rent (T90)", nett["base"],
             "Recently signed base/contract rent from the rent roll — no bundled fees."),
            ("Implied gap (gross HD asking − base)", nett["gap"],
             "If HD sits a roughly constant amount above base across units, that delta is a bundled fee."),
            ("Fee netted (from BOTH asking & effective)", fee, nett["source"]),
            ("HelloData T90 asking — net of fee", net_ask,
             "Net asking shown on the Lease Trend / Unit Mix tabs."),
            ("HelloData T90 effective — net of fee", net_eff,
             "Net effective shown on the Lease Trend / Unit Mix tabs. The same fee is removed "
             "from T365 and the YoY reads as well."),
        ]
        for label, val, note in net_rows:
            bold_red = (label.startswith("Fee netted") and fee > 0)
            _set(ws, r, 1, label, font=_f(bold=True))
            _set(ws, r, 2, val, fmt=MONEY, align="right",
                 font=_f(bold=True, color="C00000") if bold_red else _f())
            _set(ws, r, 6, note, font=_f(color=GREY, size=9), wrap=True)
            ws.row_dimensions[r].height = 28
            r += 1
        comp_txt = ", ".join(f"{cc} ${amt:,.2f}" for cc, amt in nett["candidates"]) or "none detected"
        _set(ws, r, 1, "Candidate flat fees on rent roll", font=_f(bold=True, color=NAVY))
        _set(ws, r, 2, f"{comp_txt}  (disclosure only — NOT auto-netted; "
                       f"confirm against the property website before netting)",
             font=_f(color=BLACK, size=9), wrap=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    _set(ws, r, 1, "Charge-Code Map (rent roll scheduled, monthly)", font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
    for c in range(2, 7):
        ws.cell(r, c).fill = _fill(SECT_FILL)
    r += 1
    for c, h in enumerate(["Charge Code", "Mapped Code", "Yardi Type", "Monthly $", "In Contract Rent?"], 1):
        _set(ws, r, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center", wrap=True)
    r += 1
    for cc, code, ytype, amt, isc in rec.charge_map:
        _set(ws, r, 1, cc, font=_f())
        _set(ws, r, 2, code, font=_f(bold=True, color=BLUE), align="center")
        _set(ws, r, 3, ytype or "", font=_f(color=GREY, size=9), align="center")
        _set(ws, r, 4, amt, fmt=MONEY, align="right")
        _set(ws, r, 5, "Yes" if isc else "No",
             font=_f(bold=True, color=("008000" if isc else "C00000")), align="center")
        r += 1

    if getattr(rec, "charge_t12", None):
        r += 1
        _set(ws, r, 1, "Charge → T12 Placement (empirical contract-rent test)",
             font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
        for c in range(2, 7):
            ws.cell(r, c).fill = _fill(SECT_FILL)
        r += 1
        _set(ws, r, 1, "Where each rent-roll charge actually LANDS on the T12 — matched by $ "
             "(and name) to a real T12 line — decides what's in contract rent, not the charge "
             "name. RUBS recoveries are booked as contra-expenses, so both sides are searched.",
             font=_f(color=GREY, size=9), wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 30
        r += 1
        for c, h in enumerate(["Charge", "RR $/mo", "T12 placement", "Conf.",
                               "In contract rent?", "Basis"], 1):
            _set(ws, r, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL),
                 align="center", wrap=True)
        r += 1
        _BUCKET = {"contract": "Contract rent", "other_income": "Other income",
                   "rubs_recovery": "Utility/RUBS recovery",
                   "rental_contra": "Concession / rental contra", "unmatched": "—"}
        for ct in rec.charge_t12:
            in_contract = ct["bucket"] == "contract"
            disagree = ct["agrees"] is False
            _set(ws, r, 1, str(ct["cc"])[:30], font=_f())
            _set(ws, r, 2, ct["amt"], fmt=MONEY, align="right")
            _set(ws, r, 3, _BUCKET.get(ct["bucket"], ct["bucket"]),
                 font=_f(color=(NAVY if not disagree else "C00000")), align="center")
            _set(ws, r, 4, ct["conf"], font=_f(color=GREY, size=9), align="center")
            _set(ws, r, 5, "Yes" if in_contract else "No",
                 font=_f(bold=True, color=("008000" if in_contract else "C00000")), align="center")
            _set(ws, r, 6, ct["note"], font=_f(color=(GREY if not disagree else "843C0C"),
                 size=9), wrap=True)
            ws.row_dimensions[r].height = 30
            r += 1

    if getattr(rec, "correlations", None):
        r += 1
        _set(ws, r, 1, "Correlated Cross-Checks (rent-roll counts/charges ↔ T12 income)",
             font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
        for c in range(2, 7):
            ws.cell(r, c).fill = _fill(SECT_FILL)
        r += 1
        for label, detail in rec.correlations:
            _set(ws, r, 1, label, font=_f(bold=True, color=NAVY))
            _set(ws, r, 2, detail, font=_f(color=BLACK), wrap=True)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 30
            r += 1

    if getattr(rec, "noi_tie", None):
        r += 1
        _set(ws, r, 1, "NOI Tie-Out to Operator Statement (each source statement, annual $)",
             font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
        for c in range(2, 7):
            ws.cell(r, c).fill = _fill(SECT_FILL)
        r += 1
        for c, h in enumerate(["Statement", "Reported NOI", "Standardized NOI",
                               "Δ", "Note"], 1):
            _set(ws, r, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL),
                 align="center", wrap=True)
        r += 1
        for t in rec.noi_tie:
            _set(ws, r, 1, t["label"], font=_f(bold=True))
            if t["rep_noi"] is None:
                _set(ws, r, 2, "no NOI line", font=_f(color=GREY, size=9), align="right")
                _set(ws, r, 3, round(t["comp_noi"]), fmt=MONEY, align="right")
                _set(ws, r, 5, "Source statement has no 'Net Operating Income' subtotal; "
                     "standardized NOI shown for reference.", font=_f(color=GREY, size=9), wrap=True)
            else:
                gap = t["comp_noi"] - t["rep_noi"]
                ties = abs(gap) <= max(50.0, 0.001 * abs(t["rep_noi"]))
                _set(ws, r, 2, round(t["rep_noi"]), fmt=MONEY, align="right")
                _set(ws, r, 3, round(t["comp_noi"]), fmt=MONEY, align="right")
                _set(ws, r, 4, round(gap), fmt=MONEY, align="right",
                     font=_f(bold=True, color=("008000" if ties else "C00000")))
                _set(ws, r, 5, "Ties to the operator's reported NOI (RUBS gross-up nets to zero)."
                     if ties else "Differs — a line crossed the NOI boundary; review the Code column.",
                     font=_f(color=GREY, size=9), wrap=True)
            ws.row_dimensions[r].height = 28
            r += 1

    if rec.flags:
        r += 1
        _set(ws, r, 1, "⚑ Flags for Underwriting", font=_f(bold=True, color=WHITE), fill=_fill("C00000"))
        for c in range(2, 7):
            ws.cell(r, c).fill = _fill("C00000")
        r += 1
        for fl in rec.flags:
            _set(ws, r, 1, "•  " + fl, font=_f(color="843C0C"), wrap=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 30
            r += 1

    widths = [26, 14, 14, 14, 16, 40]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.sheet_view.showGridLines = False


# ===========================================================================
# TAB: Codes  (reference + dropdown source)
# ===========================================================================
def write_codes(ws):
    _title_block(ws, "Standardized Code Legend", "Source list for the 'T12 Categorized' dropdown. Col A drives VLOOKUP & SUMIFS.")
    _set(ws, 4, 1, "Code", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center")
    _set(ws, 4, 2, "Category", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))
    _set(ws, 4, 3, "Section", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center")
    r = 5
    groups = [("EFFECTIVE GROSS REVENUE", am.REVENUE_CODES, "rev"),
              ("OPERATING EXPENSES", am.EXPENSE_CODES, "opex"),
              ("NON-OPERATING REVENUE", am.NONOP_REV_CODES, "nonop-rev"),
              ("NON-OPERATING EXPENSES", am.NONOP_EXP_CODES, "nonop-exp")]
    for gname, codes, sect in groups:
        for code, label in codes:
            _set(ws, r, 1, code, font=_f(bold=True, color=BLUE), align="center")
            _set(ws, r, 2, label, font=_f())
            _set(ws, r, 3, sect, font=_f(color=GREY, size=9), align="center")
            r += 1
    last = r - 1
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    return last


# ===========================================================================
# TAB: Dashboard
# ===========================================================================
def write_dashboard(ws, prop_name, st, rr, mix, rec, tr, lt, has_hd,
                    rr_asof="n/a", stmt_asof="n/a", rr_tab="Rent Roll (One-Line)"):
    _set(ws, 1, 1, f"{prop_name} — Underwriting Intake", font=_f(bold=True, color=NAVY, size=16))
    nfiles = len(st.files)
    span = f"{st.month_labels[0]} – {st.month_labels[-1]} ({st.n_months} mo" \
           + (f", {nfiles} statements stitched" if nfiles > 1 else "") + ")"
    _set(ws, 2, 1, f"Generated {_dt.date.today():%B %d, %Y}  ·  RR-T12 Processor intake  ·  Operating period {span}",
         font=_f(color=GREY, italic=True))
    _set(ws, 3, 1, f"Data vintage  ·  Rent roll as-of {rr_asof}  ·  Latest financial statement thru {stmt_asof}",
         font=_f(color=GREY, italic=True))

    occ = sum(1 for u in rr.units if u.occupancy == "Occupied")
    vac = sum(1 for u in rr.units if u.occupancy == "Vacant")
    nonrev = sum(1 for u in rr.units if u.occupancy == "Non-Revenue")
    nunits = len(rr.units)
    tsf = rr.totals.get("sqft", 0)
    mkt = sum(u.market_rent for u in rr.units)
    con = sum(u.contract_rent for u in rr.units)
    newc = sum(1 for u in rr.units if il.classify_lease(u) == "new")
    renc = sum(1 for u in rr.units if il.classify_lease(u) == "renewal")
    t12p = tr.periods.get("T12", {})

    # T90 market-rent indicators (trailing 90 days), mix-weighted by unit count
    def _mixwt(attr):
        num = den = 0.0
        for m in mix:
            v = getattr(m, attr, 0) or 0
            if v > 0 and m.units:
                num += v * m.units; den += m.units
        return num / den if den else 0
    hd_t90_ask = _mixwt("t90_ask")
    hd_t90_eff = _mixwt("t90_eff")
    new_t90 = il.new_lease_t90(rr)        # avg new-lease contract rent, trailing 90d

    # ---- snapshot (left, cols A-B) ----
    r = 4
    _set(ws, r, 1, "SNAPSHOT", font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
    ws.cell(r, 2).fill = _fill(SECT_FILL)
    r += 1
    rows = [
        ("Units", nunits, "0"),
        ("Occupied / Vacant / Non-Rev", f"{occ} / {vac} / {nonrev}", None),
        ("Occupancy %", (occ / nunits) if nunits else 0, "0.0%"),
        ("Total Rentable SF", tsf, INT),
        ("New / Renewal leases", f"{newc} / {renc}", None),
        ("— Market Rent Indicators (true signal) —", "", None),
    ]
    if has_hd:
        rows.append(("HelloData asking, T90 (mix-wtd)", round(hd_t90_ask) if hd_t90_ask else "n/a", MONEY if hd_t90_ask else None))
        rows.append(("HelloData effective, T90 (mix-wtd)", round(hd_t90_eff) if hd_t90_eff else "n/a", MONEY if hd_t90_eff else None))
    rows.append(("New-lease contract, T90 (avg)", round(new_t90) if new_t90 else "n/a", MONEY if new_t90 else None))
    rows += [
        ("— T12 Annualized (most recent 12 mo.) —", "", None),
        ("Effective Gross Revenue (T12)", t12p.get("_EGR"), MONEY),
        ("Operating Expenses (T12)", t12p.get("_OPEX"), MONEY),
        ("Net Operating Income (T12)", t12p.get("_NOI"), MONEY),
        ("Operating Margin (T12)", (t12p.get("_NOI", 0) / t12p.get("_EGR", 1)) if t12p.get("_EGR") else 0, "0.0%"),
    ]
    for label, val, fmt in rows:
        sect = label.startswith("—")
        _set(ws, r, 1, label, font=_f(bold=sect, color=(NAVY if sect else BLACK)))
        if val != "":
            _set(ws, r, 2, val, font=_f(bold=True), fmt=fmt, align="right")
        r += 1
    snap_last = r - 1

    # ---- unit mix (right, cols D+) ----
    um_last = write_unit_mix_block(ws, mix, has_hd, r0=4, c0=4)
    _set(ws, um_last + 1, 4,
         "New-lease = lease start \u2264 move-in (renewals excluded). HD T90/T365 = executed "
         "(off-market) asking/effective; HD90 YoY compares the trailing 90 days to the same window a year ago.",
         font=_f(color=GREY, italic=True, size=8))

    # ---- flags (full width, below both) ----
    r = max(snap_last, um_last + 1) + 2
    _set(ws, r, 1, "⚑ UNDERWRITING FLAGS", font=_f(bold=True, color=WHITE), fill=_fill("C00000"))
    for c in range(2, 13):
        ws.cell(r, c).fill = _fill("C00000")
    r += 1
    flaglist = list(getattr(st, "overlap_flags", [])) + list(getattr(st, "granularity_flags", [])) \
        + list(rec.flags) + list(lt.notes) + list(tr.notes)
    if not flaglist:
        _set(ws, r, 1, "No automatic flags raised — review categorization and reconciliation tabs.", font=_f(color=GREY, italic=True))
        r += 1
    else:
        for fl in flaglist:
            _set(ws, r, 1, "•  " + fl, font=_f(color="843C0C"), wrap=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
            ws.row_dimensions[r].height = 28
            r += 1

    # ---- how to use / paste targets ----
    r += 1
    _set(ws, r, 1, "HOW TO USE → paste into the TMG acquisition model", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))
    for c in range(2, 13):
        ws.cell(r, c).fill = _fill(HDR_FILL)
    r += 1
    steps = [
        "1.  Review 'T12 Categorized' col A (amber). Every GL line appears once with all months "
        "flowing across; edit a code via the dropdown to re-map. Category (col B) and OS Summary roll "
        "up automatically." + ("  Overlapping months are sourced from the freshest statement so nothing "
                               "double-counts." if nfiles > 1 else ""),
        "2.  'OS Summary' is the standardized chart of accounts (live SUMIFS, most recent 12 months). "
        "Copy A1:Q77 → model 'OS Summary Dump' A1 (Paste Special → Values).",
        "3.  'T12 Categorized' → copy Code, Category, Line Item + the most recent 12 month columns "
        "(NOT the Total column) → model 'T12 Dump' A2.",
        f"4.  '{rr_tab}' (newest rent roll) → copy the core columns A2:M(last) → model 'RR Dump' A2.  "
        "(The per-unit charge-code columns to the right of the gap are reference detail, not a model paste target.)",
        ("5.  'HelloData' → copy A2:U(last) → model 'HD Dump' A2." if has_hd else "5.  (No HelloData provided — skip HD Dump.)"),
        "6.  The unit mix (with new-lease + HD T90/T365/YoY market-rent indicators) is on this Dashboard; "
        "check 'Lease Trend' (seasonality / market-rent trajectory) and 'Reconciliation' (contract-rent ↔ AGPR "
        "tie-out) before underwriting. Paste targets only — do NOT populate the model itself from here.",
    ]
    for s in steps:
        _set(ws, r, 1, s, font=_f(), wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        ws.row_dimensions[r].height = 32
        r += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 2
    ws.sheet_view.showGridLines = False


# ===========================================================================
# DRIVER
# ===========================================================================
def build(t12_paths, rr_paths, hd_path=None, prop_name=None, out_path=None,
          charge_codes_path=None, hd_fee_offset=None):
    if isinstance(t12_paths, (str, bytes)):
        t12_paths = [t12_paths]
    if isinstance(rr_paths, (str, bytes)):
        rr_paths = [rr_paths]
    st = il.stitch_statements(list(t12_paths))
    charge_lookup = il.parse_charge_codes(charge_codes_path)
    hd = il.parse_hellodata(hd_path) if hd_path else None

    # data vintage — prefer a precise date from the filename, else the in-file label
    def _file_date(path):
        m = re.search(r"(20\d{2})[._\-](\d{2})[._\-](\d{2})", os.path.basename(path))
        return f"{m.group(2)}/{m.group(3)}/{m.group(1)}" if m else None

    # One or more rent rolls. The NEWEST (by as-of date) is the primary — it drives the
    # dashboard, unit mix, reconciliation and lease-type analysis. Older rolls each get their
    # own one-line tab and feed the lease-trend new-lease history (signings since turned over).
    parsed_rolls = [(p, il.parse_rent_roll(p, charge_lookup)) for p in rr_paths]
    def _roll_key(item):
        p, r = item
        return r.as_of_date or il._to_date(_file_date(p)) or _dt.date.min
    parsed_rolls.sort(key=_roll_key, reverse=True)
    rr_path, rr = parsed_rolls[0]
    extra_paths = [p for p, _ in parsed_rolls[1:]]
    extra_rolls = [r for _, r in parsed_rolls[1:]]
    rr_asof = _file_date(rr_path) or rr.as_of or "n/a"
    # latest statement = the freshest stitched file's window end (and its file date if present)
    latest_t12 = max(t12_paths, key=lambda p: (_file_date(p) or "", p))
    stmt_asof = (st.month_labels[-1] if st.month_labels else "n/a")
    if not prop_name:
        prop_name = (st.title or os.path.splitext(os.path.basename(rr_path))[0]).strip()
    # HelloData scrapes the website "Total Monthly", which can bundle mandatory flat fees
    # (pest, amenity, valet trash, utility admin, …) on top of base rent — inflating HD vs the
    # rent roll's base/contract rent. We DO NOT silently infer or auto-apply a bundle: keying
    # on rent-roll charge codes is unreliable (it grabs per-unit "Varies" charges the website
    # excludes, and misses fees that aren't itemized per unit). Instead HD is shown GROSS by
    # default, and a fee is netted only when --hd-fee-offset gives a website-confirmed amount.
    # The rent-roll candidate bundle and the gross-HD-vs-base gap are SURFACED on the
    # Reconciliation tab as evidence, but they never change the numbers on their own.
    def _mw(rows, attr):
        num = den = 0.0
        for m in rows:
            v = getattr(m, attr, 0) or 0
            if v > 0 and m.units:
                num += v * m.units; den += m.units
        return num / den if den else 0.0
    bundle_total, bundle_comps = il.mandatory_fee_bundle(rr)
    hd_fee = hd_fee_offset if hd_fee_offset is not None else 0.0
    hd_raw = hd_raw_eff = base = calib_gap = None
    if hd:
        _gross_mix = il.build_unit_mix(rr, hd, hd_fee=0.0)               # gross HD (no fee)
        hd_raw = _mw(_gross_mix, "t90_ask")                             # gross HD T90 asking
        hd_raw_eff = _mw(_gross_mix, "t90_eff")                         # gross HD T90 effective
        base = il.new_lease_t90(rr)                                      # rent-roll new-lease base
        calib_gap = (hd_raw - base) if (hd_raw and base) else None
    mix = il.build_unit_mix(rr, hd, hd_fee=hd_fee)   # joins HD by unit #; RR as-of for new/renewal
    rec = il.reconcile(st, rr)
    rec.noi_tie = il.reconcile_noi(st)
    for t in rec.noi_tie:
        if t["rep_noi"] is None:
            continue
        gap = t["comp_noi"] - t["rep_noi"]
        if abs(gap) > max(50.0, 0.001 * abs(t["rep_noi"])):
            rec.flags.append(
                f"Standardized NOI for {t['label']} (${t['comp_noi']:,.0f}) differs from the "
                f"operator's reported Net Operating Income (${t['rep_noi']:,.0f}) by ${gap:,.0f} "
                f"-- a line crossed the NOI boundary in categorization; review the Code column.")
    # Ambiguous categorizations: lines the categorizer could only place weakly (CAM/
    # amenity fees, generic 'other/misc', non-utility 'reimbursement', adjustments).
    # Surface MATERIAL ones (>= $25k/yr) so the skill ASKS the user to confirm rather
    # than silently trusting the guess. (See SKILL.md "Confirm ambiguous categorizations".)
    _seen_amb = set()
    for item in il.unified_lines(st):
        if item.get("type") != "line":
            continue
        v = [x for x in item["values"] if isinstance(x, (int, float))]
        annual = sum(v[-12:])                      # most-recent-12-month run-rate
        key = (item["name"], item.get("code"))
        # (a) ambiguously-named lines (CAM/amenity, generic other/misc, non-utility reimbursement)
        why = am.ambiguity_reason(item["name"], item.get("code"))
        if why and abs(annual) >= 25000 and key not in _seen_amb:
            _seen_amb.add(key)
            rec.flags.append(
                f"CONFIRM categorization: '{item['name']}' (~${annual:,.0f}/yr) is coded "
                f"'{item.get('code')}' — {why}. Verify this code before underwriting.")
        # (b) a NET-NEGATIVE expense line is often INCOME (a resident recovery / contra booked
        #     against expense — tenant rebills, utility passthroughs). Surface materially- and
        #     consistently-negative expense lines so the right ones get reclassified to revenue.
        if item.get("side") == "expense" and key not in _seen_amb:
            nz = [x for x in v if abs(x) > 1e-9]
            if nz and annual <= -2500 and sum(1 for x in nz if x < 0) / len(nz) >= 0.6:
                _seen_amb.add(key)
                rec.flags.append(
                    f"CONFIRM categorization: '{item['name']}' is a NET-NEGATIVE expense "
                    f"(~${annual:,.0f}/yr) coded '{item.get('code')}' — a negative expense is "
                    f"often income (a resident recovery/contra); confirm whether it belongs in revenue.")
    rec.charge_t12 = il.match_charges_to_t12(st, rr)
    if hd:
        rec.hd_fee_netting = {
            "hd_raw": hd_raw or 0.0,
            "hd_raw_eff": hd_raw_eff or 0.0,
            "base": base or 0.0,
            "gap": calib_gap if calib_gap is not None else 0.0,
            "applied": hd_fee,
            "candidates": bundle_comps,
            "candidate_total": bundle_total,
            "offset_given": hd_fee_offset is not None,
            "source": ("Explicit override (--hd-fee-offset) — confirmed against the property "
                       "website 'Total Monthly'."
                       if hd_fee_offset is not None else
                       "Not netted — HD shown gross. Set --hd-fee-offset to net a "
                       "website-confirmed bundle."),
        }
        # Flag (don't auto-apply): when HD gross sits materially above the new-lease base and
        # no override was given, the website may be bundling mandatory fees. Surface it as a
        # CONFIRM action — the true bundle lives on the property website, not in this data.
        if (hd_fee_offset is None and calib_gap is not None
                and calib_gap > max(10.0, 0.005 * (base or 0.0))):
            comp_txt = ", ".join(f"{cc} ${amt:,.2f}" for cc, amt in bundle_comps) or "none itemized"
            rec.flags.append(
                f"HelloData asking sits ${calib_gap:,.0f}/mo above the new-lease base rent "
                f"(gross HD T90 ${hd_raw:,.0f} vs base ${base:,.0f}). This is usually ordinary "
                f"market premium (current asking above recently-signed rents), but on properties "
                f"whose website advertises ALL-IN pricing, HelloData can pick up bundled fees. HD "
                f"is shown GROSS (not netted). To check: compare HD's asking for a currently-listed "
                f"unit to that unit's base rent vs 'Total Monthly' on the property website — if HD "
                f"ties to the all-in total, re-run with --hd-fee-offset <$/mo>; if it ties to base "
                f"(as at Aura), do NOT net. Candidate flat fees on the rent roll: {comp_txt}.")
    tr = il.build_trends(st, rr)
    lt = il.build_lease_trend(rr, hd, hd_fee=hd_fee, extra_rolls=extra_rolls)
    if extra_rolls:
        older = ", ".join(sorted((_file_date(p) or "older") for p in extra_paths))
        lt.notes.insert(0, f"New-Lease Contract / unit history draws on {len(extra_rolls)+1} rent "
                           f"rolls (current {rr_asof} + older {older}); older rolls supply signings "
                           f"that have since turned over. The dashboard and unit mix use the current "
                           f"rent roll only.")
    if hd:
        comp_txt = ", ".join(f"{cc} ${amt:,.2f}" for cc, amt in bundle_comps) or "—"
        if hd_fee > 0:
            note = (f"HelloData reflects the website 'Total Monthly', which can bundle mandatory "
                    f"fees. HD market rent here is shown NET of ${hd_fee:,.2f}/mo (explicit override). "
                    f"See 'Reconciliation → HelloData Market Rent: Fee Netting' for the full derivation.")
        else:
            note = (f"HelloData market rent is shown GROSS — no fee netted. Candidate flat fees on "
                    f"the rent roll: {comp_txt}. If the property website bundles fees into its asking "
                    f"price, set --hd-fee-offset; see 'Reconciliation → HelloData Market Rent: Fee Netting'.")
        lt.notes.insert(0, note)

    n = st.n_months

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_t12 = wb.create_sheet("T12 Categorized")
    ws_os = wb.create_sheet("OS Summary")
    # One-line tab per rent roll. With a single roll, keep the canonical (undated) name; with
    # several, DATE EVERY tab for symmetry. The newest stays the model-paste target.
    def _tab_date(path, r):
        d = _file_date(path)
        if d:
            mm, dd, yy = d.split("/")
            return f"{mm}-{dd}-{yy[2:]}"
        return r.as_of_date.strftime("%m-%d-%y") if r.as_of_date else "older"
    if extra_rolls:
        primary_tab = f"Rent Roll (One-Line) {_tab_date(rr_path, rr)}"[:31]
    else:
        primary_tab = "Rent Roll (One-Line)"
    ws_rr = wb.create_sheet(primary_tab)
    ws_rr_hist = []
    for p, r in zip(extra_paths, extra_rolls):
        name = f"Rent Roll (One-Line) {_tab_date(p, r)}"[:31]
        ws_rr_hist.append((wb.create_sheet(name), r))
    ws_lt = wb.create_sheet("Lease Trend")
    ws_hd = wb.create_sheet("HelloData") if hd else None
    ws_rec = wb.create_sheet("Reconciliation")
    ws_codes = wb.create_sheet("Codes")

    codes_last = write_codes(ws_codes)
    wb.defined_names["CodeList"] = DefinedName("CodeList", attr_text=f"Codes!$A$5:$A${codes_last}")

    write_t12_categorized(ws_t12, st, codes_last)
    last12 = list(range(max(0, n - 12), n))
    sub = "Adjusted Amounts ($ totals) — most recent 12 months · model paste target (A1:Q77)" if n > 12 \
        else "Adjusted Amounts ($ totals) — model paste target (A1:Q77)"
    write_os_grid(ws_os, st, prop_name, last12,
                  title="Operating Statement — Standardized (live rollup)", subtitle=sub)
    write_rent_roll(ws_rr, rr, hd)
    for ws_h, r in ws_rr_hist:
        write_rent_roll(ws_h, r, hd)
    write_lease_trend(ws_lt, lt, st, tr, rr, has_hd=bool(hd))
    if ws_hd is not None:
        write_hellodata(ws_hd, hd)
    write_reconciliation(ws_rec, rec, rr)
    write_dashboard(ws_dash, prop_name, st, rr, mix, rec, tr, lt, has_hd=bool(hd),
                    rr_asof=rr_asof, stmt_asof=stmt_asof, rr_tab=primary_tab)

    if not out_path:
        safe = "".join(ch if ch.isalnum() else "_" for ch in prop_name).strip("_")
        out_path = f"/mnt/user-data/outputs/{safe}__Underwriting_Intake.xlsx"
    wb.save(out_path)

    # ---- COMPLETENESS GATE: never present a half-built workbook silently. ----
    # If a core input layout wasn't recognized, the build still "succeeds" but tabs come
    # out empty. Detect that loudly so the caller refuses to deliver (see SKILL.md
    # "Completeness gate"). The marker line is machine-readable; the banner is on stderr.
    problems = []
    if not getattr(st, "monthkeys", None):
        problems.append("T12 produced NO month columns — the statement layout wasn't recognized.")
    elif not any(r.rtype == "line" for f in st.files for r in f.t12.rows):
        problems.append("T12 produced NO categorized line items.")
    if not rr.units:
        problems.append("Rent roll parsed 0 units — the rent-roll layout wasn't recognized; "
                        "Dashboard, unit mix, occupancy and reconciliation are EMPTY.")
    elif not mix:
        problems.append("Unit mix is empty (rent roll + HelloData produced no floor-plan rows).")
    # NOI TIE-OUT: standardized NOI must reconcile to the operator's OWN reported NOI —
    # re-bucketing lines (incl. the RUBS gross-up) is NOI-neutral, so a material gap means a
    # categorization/parse BUG (a parent/statistic row summed as a leaf, a sign error, or a
    # line crossing the NOI boundary), NOT a methodology choice. Surface it loudly.
    noi_gaps = []
    for t in getattr(rec, "noi_tie", []) or []:
        if t.get("rep_noi") is None:
            continue
        gap = t["comp_noi"] - t["rep_noi"]
        if abs(gap) > max(500.0, 0.005 * abs(t["rep_noi"])):
            noi_gaps.append(f"{t['label']}: standardized ${t['comp_noi']:,.0f} vs operator "
                            f"reported ${t['rep_noi']:,.0f} (Δ ${gap:,.0f})")
    if problems:
        print("\n".join(["", "=" * 70,
              "⚠  BUILD INCOMPLETE — DO NOT DELIVER THIS WORKBOOK.",
              "Missing / unrecognized:"] + [f"  - {p}" for p in problems]
              + ["Tell the user exactly what's missing and what's needed (usually an",
                 "unrecognized T12/rent-roll layout to fix in the parser), then stop.",
                 "=" * 70, ""]), file=sys.stderr)
        print("BUILD_INCOMPLETE: " + " | ".join(problems))
    elif noi_gaps:
        print("\n".join(["", "=" * 70,
              "⚠  NOI DOES NOT TIE to the operator's reported NOI — INVESTIGATE before delivering.",
              "Re-bucketing is NOI-neutral, so a gap = a categorization/parse bug (parent or",
              "statistic row summed as a leaf, sign error, or a line crossing the NOI boundary)."]
              + [f"  - {g}" for g in noi_gaps] + ["=" * 70, ""]), file=sys.stderr)
        print("BUILD_OK_BUT_NOI_UNTIED: " + " | ".join(noi_gaps))
    else:
        print(f"BUILD_OK: {len(rr.units)} units, {len(mix)} floor plans, "
              f"{len(st.monthkeys)} months. NOI ties to operator.")
    return out_path


def main():
    ap = argparse.ArgumentParser(description="Build a RedIQ-replacement underwriting intake workbook.")
    ap.add_argument("--t12", required=True, nargs="+",
                    help="One or more T12 / monthly operating statements (stitched into one continuous series).")
    ap.add_argument("--rr", required=True, nargs="+",
                    help="One or more rent rolls. The newest (by as-of date) drives the "
                         "dashboard/unit mix; each gets its own one-line tab and older rolls "
                         "extend the lease-trend new-lease history.")
    ap.add_argument("--hd", default=None, help="HelloData unit-details CSV (optional).")
    ap.add_argument("--charge-codes", default=None, dest="charge_codes",
                    help="Optional charge-code lookup (Account/Name[/Type]) for rent rolls "
                         "that bill by numeric code rather than name.")
    ap.add_argument("--hd-fee-offset", dest="hd_fee_offset", type=float, default=None,
                    help="$/mo of website-bundled mandatory fees to net from HelloData asking "
                         "AND effective. Default: none (HD shown gross; the build flags when HD "
                         "sits materially above the new-lease base so you can confirm the "
                         "website's 'Total Monthly' and set this).")
    ap.add_argument("--name", default=None)
    ap.add_argument("--out", default=None)
    a = ap.parse_args()
    out = build(a.t12, a.rr, a.hd, a.name, a.out, charge_codes_path=a.charge_codes,
                hd_fee_offset=a.hd_fee_offset)
    print(out)


if __name__ == "__main__":
    main()
