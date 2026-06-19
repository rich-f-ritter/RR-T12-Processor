#!/usr/bin/env python3
"""
build_intake.py  --  RedIQ-replacement underwriting intake workbook writer.

Reads a raw T12 + rent roll (+ optional HelloData CSV) and produces a single
formatted .xlsx with an EDITABLE chart-of-accounts categorization whose OS
Summary rolls up LIVE via SUMIFS.  Change a code in 'T12 Categorized' col A and
the standardized OS Summary updates automatically -- then copy/paste the dump
tabs straight into the TMG acquisition model.

Tabs:  Dashboard | T12 Categorized | OS Summary | Rent Roll (One-Line) |
       Unit Mix | HelloData (opt) | Reconciliation | Trends | Codes

CLI:
    python build_intake.py --t12 T12.xlsx --rr RentRoll.xlsx \
        [--hd hello.csv] [--name "Property"] [--out out.xlsx]
"""
import argparse
import datetime as _dt
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

import intake_lib as il
import account_map as am

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
FONT = "Arial"
BLUE = "0000FF"      # editable inputs (the code column)
BLACK = "000000"     # same-sheet formulas / data
GREEN = "008000"     # cross-sheet links (SUMIFS / VLOOKUP)
GREY = "808080"
WHITE = "FFFFFF"
NAVY = "1F3864"
HDR_FILL = "1F3864"
SUB_FILL = "D9E1F2"
EDIT_FILL = "FFF2CC"   # soft yellow → "edit here"
SECT_FILL = "8EAADB"
ZEBRA = "F2F5FB"

MONEY = '#,##0;(#,##0);"-"'
MONEY2 = '#,##0.00;(#,##0.00);"-"'
INT = '#,##0;(#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'
DATEF = "mm/dd/yyyy"
MONF = "mmm yyyy"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPLINE = Border(top=Side(style="thin", color="404040"))
DBLTOP = Border(top=Side(style="double", color="404040"))


def _f(bold=False, color=BLACK, size=10, italic=False):
    return Font(name=FONT, bold=bold, color=color, size=size, italic=italic)


def _fill(hexc):
    return PatternFill("solid", fgColor=hexc)


def _set(ws, r, c, val, *, font=None, fill=None, fmt=None, align=None,
         wrap=False, border=None):
    cell = ws.cell(r, c, val)
    cell.font = font or _f()
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align or wrap:
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        cell.border = border
    return cell


def _title_block(ws, title, subtitle=""):
    _set(ws, 1, 1, title, font=_f(bold=True, color=NAVY, size=15))
    if subtitle:
        _set(ws, 2, 1, subtitle, font=_f(color=GREY, size=10, italic=True))


# ===========================================================================
# TAB: T12 Categorized  (full T12, editable code column, Total per row)
# ===========================================================================
def write_t12_categorized(ws, st, code_list_rows: int):
    """One clean chart of accounts across the full union of months. Every line appears
    once (merged across statements); each month's value comes from the freshest statement
    that owns it, so a SUMIFS by code reproduces the resolved series with no double-count.
    Column A (code) stays editable; section bands separate income / expenses / non-op."""
    n = st.n_months
    months = st.months
    MS = 4                                       # month columns start at D
    total_col = MS + n
    last_col = total_col
    multi = len(st.files) > 1

    _set(ws, 1, 1, "Code", font=_f(bold=True, color=WHITE), align="center")
    ws.cell(1, 1).fill = _fill("C9A227")
    _set(ws, 1, 2, "Category", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center")
    _set(ws, 1, 3, "Line Item (raw GL)", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))
    for k in range(n):
        d = _dt.datetime(months[k][1][0], months[k][1][1], 1)
        _set(ws, 1, MS + k, d, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL),
             fmt=MONF, align="center")
    _set(ws, 1, total_col, "Total", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center")
    # optional sourcing note when multiple statements are stitched
    if multi:
        src = " · ".join(f"{f.label} ({f.t12.month_labels[0]}\u2013{f.t12.month_labels[-1]})"
                         for f in st.files)
        _set(ws, 2, 3, f"Stitched from: {src}. Each month is sourced from the freshest "
                       f"statement covering it; blank cells = that line not itemized in the "
                       f"statement owning that month.", font=_f(color=GREY, italic=True, size=8))

    r = 3 if multi else 2
    for item in il.unified_lines(st):
        if item["type"] == "header":
            _set(ws, r, 3, item["name"], font=_f(bold=True, color=NAVY))
            for c in range(1, last_col + 1):
                ws.cell(r, c).fill = _fill(SUB_FILL)
            r += 1
            continue
        _set(ws, r, 1, item["code"], font=_f(bold=True, color=BLUE), align="center")
        _set(ws, r, 2, f'=IFERROR(VLOOKUP($A{r},Codes!$A:$B,2,FALSE()),"")', font=_f(color=GREEN))
        _set(ws, r, 3, item["name"], font=_f(color=BLACK))
        for k, v in enumerate(item["values"]):
            if v is not None:
                _set(ws, r, MS + k, v, font=_f(color=BLACK), fmt=MONEY)
        _set(ws, r, total_col,
             f"=SUM({get_column_letter(MS)}{r}:{get_column_letter(MS + n - 1)}{r})",
             font=_f(bold=True, color=BLACK), fmt=MONEY)
        r += 1
    last_row = r - 1

    dv = DataValidation(type="list", formula1="CodeList", allow_blank=True, showDropDown=False)
    dv.error = "Pick a standardized code from the list (or clear to exclude)."
    dv.errorTitle = "Invalid code"
    dv.prompt = "Edit this code to re-map the line. OS Summary updates automatically."
    dv.promptTitle = "Editable code"
    ws.add_data_validation(dv)
    dv.add(f"A2:A{last_row}")

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 42
    for k in range(n):
        ws.column_dimensions[get_column_letter(MS + k)].width = 12
    ws.column_dimensions[get_column_letter(total_col)].width = 13
    ws.freeze_panes = "D2"
    ws.sheet_view.showGridLines = False
    return last_row, total_col, MS, n


# ===========================================================================
# TAB: OS Summary / Operating History  (RedIQ Overview template, live SUMIFS)
# ===========================================================================
def write_os_grid(ws, st, prop_name, show_idx, *, title, subtitle):
    """RedIQ-style standardized statement. `show_idx` is the list of union-month
    indices to display (e.g. last 12 for the model paste, or all for the full
    history). Each shown column's SUMIFS points at the matching 'T12 Categorized'
    month column so the rollup stays live and editable."""
    n = len(show_idx)
    months = st.months
    TC = "'T12 Categorized'"
    MS = 4                                    # T12 Categorized month columns start at D

    def os_col(k):
        return 6 + k                          # display column (F=6)

    def tc_col(k):
        return get_column_letter(MS + show_idx[k])   # matching source column

    _set(ws, 1, 2, prop_name, font=_f(bold=True, color=NAVY, size=13))
    _set(ws, 2, 2, title, font=_f(color=GREY, italic=True))
    _set(ws, 3, 2, subtitle, font=_f(color=GREY, italic=True, size=9))

    for cc in (3, 4, 5):
        _set(ws, 4, cc, "Annual", font=_f(bold=True, color=GREY, size=9), align="center")
    for k in range(n):
        _set(ws, 4, os_col(k), "Monthly", font=_f(bold=True, color=WHITE, size=9),
             fill=_fill(HDR_FILL), align="center")
        d = _dt.datetime(months[show_idx[k]][1][0], months[show_idx[k]][1][1], 1)
        _set(ws, 5, os_col(k), d, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL),
             fmt=MONF, align="center")
    _set(ws, 6, 1, "Code", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center")
    _set(ws, 6, 2, "Effective Gross Revenue", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))

    def sumifs(row, anchor):
        for k in range(n):
            col = tc_col(k)
            ws.cell(row, os_col(k), f"=SUMIFS({TC}!{col}:{col},{TC}!$A:$A,$A{anchor})")
            c = ws.cell(row, os_col(k)); c.font = _f(color=GREEN); c.number_format = MONEY

    def subtotal(row, fn, *, bold=True, border=TOPLINE, fill=None):
        for k in range(n):
            L = get_column_letter
            ws.cell(row, os_col(k), fn(L(os_col(k))))
            c = ws.cell(row, os_col(k)); c.font = _f(bold=bold, color=BLACK)
            c.number_format = MONEY; c.border = border
            if fill:
                c.fill = fill

    r = 7
    for code, label in am.REVENUE_CODES:
        _set(ws, r, 1, code, font=_f(bold=True, color=BLACK), align="center")
        _set(ws, r, 2, label, font=_f(color=BLACK)); sumifs(r, r); r += 1
    _set(ws, 20, 2, "Effective Gross Revenue", font=_f(bold=True, color=NAVY), fill=_fill(SUB_FILL))
    ws.cell(20, 1).fill = _fill(SUB_FILL)
    subtotal(20, lambda c: f"=SUM({c}7:{c}19)", fill=_fill(SUB_FILL))

    _set(ws, 22, 2, "Operating Expenses", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))
    ws.cell(22, 1).fill = _fill(HDR_FILL)
    r = 23
    for code, label in am.EXPENSE_CODES:
        _set(ws, r, 1, code, font=_f(bold=True, color=BLACK), align="center")
        _set(ws, r, 2, label, font=_f(color=BLACK)); sumifs(r, r); r += 1
    _set(ws, 42, 2, "Operating Expenses", font=_f(bold=True, color=NAVY), fill=_fill(SUB_FILL))
    ws.cell(42, 1).fill = _fill(SUB_FILL)
    subtotal(42, lambda c: f"=SUM({c}23:{c}41)", fill=_fill(SUB_FILL))

    _set(ws, 44, 2, "Net Operating Income", font=_f(bold=True, color=WHITE, size=11), fill=_fill(NAVY))
    ws.cell(44, 1).fill = _fill(NAVY)
    subtotal(44, lambda c: f"={c}20-{c}42", border=DBLTOP, fill=_fill(NAVY))
    for k in range(n):
        ws.cell(44, os_col(k)).font = _f(bold=True, color=WHITE, size=11)

    _set(ws, 46, 2, "Non-Operating Revenue", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))
    ws.cell(46, 1).fill = _fill(HDR_FILL)
    r = 47
    for code, label in am.NONOP_REV_CODES:
        _set(ws, r, 1, code, font=_f(bold=True, color=BLACK), align="center")
        _set(ws, r, 2, label, font=_f(color=BLACK)); sumifs(r, r); r += 1
    _set(ws, 48, 2, "Non-Operating Revenue", font=_f(bold=True, color=NAVY), fill=_fill(SUB_FILL))
    ws.cell(48, 1).fill = _fill(SUB_FILL)
    subtotal(48, lambda c: f"=SUM({c}47:{c}47)", fill=_fill(SUB_FILL))

    _set(ws, 50, 2, "Non-Operating Expenses", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))
    ws.cell(50, 1).fill = _fill(HDR_FILL)
    r = 51
    for code, label in am.NONOP_EXP_CODES:
        _set(ws, r, 1, code, font=_f(bold=True, color=BLACK), align="center")
        _set(ws, r, 2, label, font=_f(color=BLACK)); sumifs(r, r); r += 1
    _set(ws, 73, 2, "Non-Operating Expenses", font=_f(bold=True, color=NAVY), fill=_fill(SUB_FILL))
    ws.cell(73, 1).fill = _fill(SUB_FILL)
    subtotal(73, lambda c: f"=SUM({c}51:{c}72)", fill=_fill(SUB_FILL))

    _set(ws, 75, 2, "Net Income", font=_f(bold=True, color=WHITE, size=11), fill=_fill(NAVY))
    ws.cell(75, 1).fill = _fill(NAVY)
    subtotal(75, lambda c: f"={c}44+{c}48-{c}73", border=DBLTOP, fill=_fill(NAVY))
    for k in range(n):
        ws.cell(75, os_col(k)).font = _f(bold=True, color=WHITE, size=11)

    _set(ws, 77, 2, f"Generated {_dt.date.today():%m/%d/%Y} \u2014 live SUMIFS over 'T12 Categorized' code column",
         font=_f(color=GREY, italic=True, size=9))

    for col, w in (("A", 11), ("B", 32), ("C", 9), ("D", 9), ("E", 9)):
        ws.column_dimensions[col].width = w
    for k in range(n):
        ws.column_dimensions[get_column_letter(6 + k)].width = 12
    ws.freeze_panes = "F7"
    ws.sheet_view.showGridLines = False


# ===========================================================================
# TAB: Rent Roll (One-Line)
# ===========================================================================
RR_HEADERS = ["Unit No.", "Floor Plan", "Net sf", "Bed", "Bath", "Lease Type",
              "Occupancy Status", "Market Rent", "Contractual Rent",
              "Net Effective Rent", "Lease Start Date", "Lease Expiration",
              "Move In Date"]


def write_rent_roll(ws, rr: il.RentRoll, hd):
    for c, h in enumerate(RR_HEADERS, 1):
        _set(ws, 1, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL),
             align="center", wrap=True)
    # charge-code columns: union of charges with a non-zero total, ordered by magnitude.
    # one blank spacer column between the core fields (A-M) and the charge block.
    ncore = len(RR_HEADERS)                       # 13 -> A..M
    cc_tot = {}
    for u in rr.units:
        for cc, amt in u.charges.items():
            cc_tot[cc] = cc_tot.get(cc, 0.0) + (amt or 0.0)
    charge_codes = [cc for cc, t in sorted(cc_tot.items(), key=lambda kv: -abs(kv[1])) if abs(t) > 0.005]
    spacer = ncore + 1                            # N (blank)
    cstart = ncore + 2                            # O -> first charge column
    if charge_codes:
        # group label over the charge block (spacer column N is left empty)
        _set(ws, 1, cstart, "Scheduled Charges by Code (monthly $)", font=_f(bold=True, color=WHITE),
             fill=_fill(SECT_FILL))
        for j, cc in enumerate(charge_codes):
            _set(ws, 2, cstart + j, cc, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL),
                 align="center", wrap=True)

    data_start = 3 if charge_codes else 2
    # if no charge block, headers sit on row 1 and data on row 2 (original layout)
    hdr_row = 1
    if charge_codes:
        # re-place core headers on row 2 as well so the charge sub-header row aligns;
        # keep row-1 core headers but merge each down two rows for a clean look
        for c in range(1, ncore + 1):
            ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)
        ws.merge_cells(start_row=1, start_column=cstart, end_row=1, end_column=cstart + len(charge_codes) - 1)

    r = data_start
    for u in rr.units:
        bed, bath, _src = il.infer_bed_bath(u.floorplan, hd)
        zeb = _fill(ZEBRA) if (r - data_start) % 2 == 0 else None
        vals = [u.unit, u.floorplan, u.sqft, bed, bath, u.lease_type, u.occupancy,
                u.market_rent, u.contract_rent, u.net_effective,
                u.lease_start, u.lease_end, u.move_in]
        for c, v in enumerate(vals, 1):
            fmt = None
            al = None
            if c in (3,):
                fmt = INT; al = "right"
            elif c in (4, 5):
                fmt = "0"; al = "center"
            elif c in (8, 9, 10):
                fmt = MONEY; al = "right"
            elif c in (11, 12, 13):
                fmt = DATEF; al = "center"
            elif c in (1, 2, 6, 7):
                al = "left" if c in (1, 2) else "center"
            _set(ws, r, c, v, font=_f(), fmt=fmt, align=al, fill=zeb)
        for j, cc in enumerate(charge_codes):
            amt = u.charges.get(cc)
            _set(ws, r, cstart + j, (amt if amt not in (None, 0) else None),
                 font=_f(), fmt=MONEY, align="right", fill=zeb)
        r += 1
    last = r - 1
    # totals row
    _set(ws, last + 1, 1, "TOTAL / AVG", font=_f(bold=True, color=NAVY))
    _set(ws, last + 1, 3, f"=SUM(C{data_start}:C{last})", font=_f(bold=True, color=NAVY), fmt=INT, align="right")
    _set(ws, last + 1, 8, f"=SUM(H{data_start}:H{last})", font=_f(bold=True, color=NAVY), fmt=MONEY, align="right")
    _set(ws, last + 1, 9, f"=SUM(I{data_start}:I{last})", font=_f(bold=True, color=NAVY), fmt=MONEY, align="right")
    _set(ws, last + 1, 10, f"=SUM(J{data_start}:J{last})", font=_f(bold=True, color=NAVY), fmt=MONEY, align="right")
    for j, cc in enumerate(charge_codes):
        L = get_column_letter(cstart + j)
        _set(ws, last + 1, cstart + j, f"=SUM({L}{data_start}:{L}{last})",
             font=_f(bold=True, color=NAVY), fmt=MONEY, align="right")
    for c in range(1, cstart + len(charge_codes)):
        ws.cell(last + 1, c).border = DBLTOP

    widths = [11, 11, 9, 6, 6, 12, 16, 13, 14, 14, 14, 14, 14]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    if charge_codes:
        ws.column_dimensions[get_column_letter(spacer)].width = 2
        for j in range(len(charge_codes)):
            ws.column_dimensions[get_column_letter(cstart + j)].width = 13
    ws.freeze_panes = "B" + str(data_start)
    ws.sheet_view.showGridLines = False
    return last


# ===========================================================================
# Unit-mix block (embedded in the Dashboard, not a standalone tab)
# ===========================================================================
def write_unit_mix_block(ws, mix, has_hd, r0, c0):
    """Render the unit-mix table at origin (r0, c0); columns are relative to c0.
    New-lease (last 5) and HelloData T90 / T365 executed are market-rent indicators;
    HD90 YoY compares the trailing-90-day executed rent to the same window a year ago.
    Totals are Python-computed mix-weighted values. Returns the last row used."""
    def col(j):
        return c0 + j

    def CL(j):
        return get_column_letter(c0 + j)

    hdr = ["Floor Plan", "Bed", "Bath", "Units", "Occ", "Vac", "Avg SF",
           "Avg Market\nRent", "Avg Contract\nRent", "New\nLeases", "Renew\nLeases",
           "Avg New-Lease\n(last 5)", "New #1", "New #2", "New #3", "New #4", "New #5",
           "HD T90\nAsking", "HD T90\nEffective", "HD T365\nAsking", "HD T365\nEffective",
           "HD90 YoY\nAsking", "HD90 YoY\nEffective", "Bed/Bath\nSrc"]
    # j: 0 plan,1 bed,2 bath,3 units,4 occ,5 vac,6 avgSF,7 avgMkt,8 avgCon,9 new,10 ren,
    # 11 avgNew5,12-16 new1-5,17 t90ask,18 t90eff,19 t365ask,20 t365eff,21 yoyAsk,22 yoyEff,23 src
    _set(ws, r0, c0, "UNIT MIX  \u2014  mix-weighted market-rent indicators",
         font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
    for j in range(1, len(hdr)):
        ws.cell(r0, col(j)).fill = _fill(SECT_FILL)
    hr = r0 + 1
    for j, h in enumerate(hdr):
        _set(ws, hr, col(j), h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL),
             align="center", wrap=True)
    r = hr + 1
    for m in mix:
        zeb = _fill(ZEBRA) if (r - hr) % 2 == 0 else None

        def put(j, v, **kw):
            _set(ws, r, col(j), v, fill=zeb, **kw)
        put(0, m.plan, font=_f(bold=True))
        put(1, m.bed, fmt="0", align="center")
        put(2, m.bath, fmt="0", align="center")
        put(3, m.units, fmt="0", align="center")
        put(4, m.occ, fmt="0", align="center")
        put(5, m.vac, fmt="0", align="center")
        put(6, round(m.avg_sqft), fmt=INT, align="right")
        put(7, round(m.avg_market), fmt=MONEY, align="right")
        put(8, round(m.avg_contract), fmt=MONEY, align="right")
        put(9, m.new_count, fmt="0", align="center")
        put(10, m.renewal_count, fmt="0", align="center")
        put(11, round(m.avg_new_last5) if m.avg_new_last5 else None,
            font=_f(bold=True, color="843C0C"), fmt=MONEY, align="right")
        for k in range(5):
            put(12 + k, m.new_last5_rents[k] if k < len(m.new_last5_rents) else None,
                fmt=MONEY, align="right")
        if has_hd:
            put(17, round(m.t90_ask) if m.t90_ask else None, fmt=MONEY, align="right")
            put(18, round(m.t90_eff) if m.t90_eff else None,
                font=_f(bold=True, color=GREEN), fmt=MONEY, align="right")
            put(19, round(m.t365_ask) if m.t365_ask else None, fmt=MONEY, align="right")
            put(20, round(m.t365_eff) if m.t365_eff else None, fmt=MONEY, align="right")
            put(21, m.yoy_ask if m.yoy_ask is not None else None, fmt=PCT, align="right")
            put(22, m.yoy_eff if m.yoy_eff is not None else None, fmt=PCT, align="right")
        put(23, m.bedbath_source, font=_f(color=GREY, size=9), align="center")
        r += 1

    # totals (mix-weighted, computed in Python so blanks/None are handled cleanly)
    def wavg(get, wt=lambda m: m.units):
        num = sum(wt(m) * get(m) for m in mix if get(m))
        den = sum(wt(m) for m in mix if get(m))
        return (num / den) if den else 0

    def wyoy(get):
        num = sum(m.units * get(m) for m in mix if get(m) is not None)
        den = sum(m.units for m in mix if get(m) is not None)
        return (num / den) if den else None

    def tot(j, v, **kw):
        color = kw.pop("color", NAVY)
        _set(ws, r, col(j), v, font=_f(bold=True, color=color), **kw)
    _set(ws, r, c0, "TOTAL / AVG", font=_f(bold=True, color=NAVY))
    tot(3, sum(m.units for m in mix), fmt="0", align="center")
    tot(4, sum(m.occ for m in mix), fmt="0", align="center")
    tot(5, sum(m.vac for m in mix), fmt="0", align="center")
    tot(6, round(wavg(lambda m: m.avg_sqft)), fmt=INT, align="right")
    tot(7, round(wavg(lambda m: m.avg_market)), fmt=MONEY, align="right")
    tot(8, round(wavg(lambda m: m.avg_contract)), fmt=MONEY, align="right")
    tot(9, sum(m.new_count for m in mix), fmt="0", align="center")
    tot(10, sum(m.renewal_count for m in mix), fmt="0", align="center")
    _an = wavg(lambda m: m.avg_new_last5, lambda m: m.new_count)
    tot(11, round(_an) if _an else None, color="843C0C", fmt=MONEY, align="right")
    if has_hd:
        for j, fld in ((17, "t90_ask"), (18, "t90_eff"), (19, "t365_ask"), (20, "t365_eff")):
            v = wavg(lambda m, f=fld: getattr(m, f))
            tot(j, round(v) if v else None, color=(GREEN if j == 18 else NAVY), fmt=MONEY, align="right")
        tot(21, wyoy(lambda m: m.yoy_ask), fmt=PCT, align="right")
        tot(22, wyoy(lambda m: m.yoy_eff), fmt=PCT, align="right")
    for j in range(len(hdr)):
        ws.cell(r, col(j)).border = DBLTOP

    w = [12, 5, 5, 6, 5, 5, 8, 11, 12, 7, 7, 12, 9, 9, 9, 9, 9, 10, 10, 10, 10, 10, 10, 9]
    for j, wd in enumerate(w):
        ws.column_dimensions[CL(j)].width = wd
    return r


# ===========================================================================
# TAB: Lease Trend  (market-rent trajectory + seasonality + new-lease detail)
# ===========================================================================
def write_lease_trend(ws, lt, rr: il.RentRoll, has_hd=True):
    _title_block(ws, "Lease Trend & Seasonality",
                 "True-market-rent movement from executed leases — HelloData (mix-weighted by floor "
                 "plan) and new-lease signings. Renewals are excluded (never market-tested).")
    r = 4
    # seasonality notes
    if lt.notes:
        _set(ws, r, 1, "Observations", font=_f(bold=True, color=NAVY))
        r += 1
        for note in lt.notes:
            _set(ws, r, 1, "\u2022  " + note, font=_f(color="843C0C"), wrap=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            ws.row_dimensions[r].height = 30
            r += 1
        r += 1

    # quarterly market-rent matrix
    _set(ws, r, 1, "Market Rent by Quarter (executed)", font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
    for c in range(2, 8):
        ws.cell(r, c).fill = _fill(SECT_FILL)
    r += 1
    cols = ["Quarter", "HD Asking\n(mix-wtd)", "HD Effective\n(mix-wtd)", "Concession\n%",
            "HD Exec\n(n)", "New-Lease Rent\n(rent roll)", "New Leases\n(n)"]
    for c, h in enumerate(cols, 1):
        _set(ws, r, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center", wrap=True)
    r += 1
    for i, q in enumerate(lt.quarters):
        zeb = _fill(ZEBRA) if i % 2 == 0 else None
        _set(ws, r, 1, q, font=_f(bold=True), align="center", fill=zeb)
        _set(ws, r, 2, round(lt.hd_ask_by_q[i]) or None, fmt=MONEY, align="right", fill=zeb)
        _set(ws, r, 3, round(lt.hd_eff_by_q[i]) or None, fmt=MONEY, align="right", fill=zeb)
        _set(ws, r, 4, f'=IF(OR(B{r}="",B{r}=0),"-",1-C{r}/B{r})', fmt=PCT, align="right", fill=zeb)
        _set(ws, r, 5, lt.hd_n_by_q[i] or None, fmt="0", align="center", fill=zeb)
        _set(ws, r, 6, round(lt.new_rent_by_q[i]) or None, fmt=MONEY, align="right", fill=zeb)
        _set(ws, r, 7, lt.new_n_by_q[i] or None, fmt="0", align="center", fill=zeb)
        r += 1
    r += 1

    # T90 by base plan
    if has_hd and lt.plan_t90:
        _set(ws, r, 1, "Trailing-90-Day Executed by Floor Plan (HelloData)",
             font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
        for c in range(2, 8):
            ws.cell(r, c).fill = _fill(SECT_FILL)
        r += 1
        for c, h in enumerate(["Base Plan", "T90 Asking", "T90 Effective", "Concession %", "Exec (n)"], 1):
            _set(ws, r, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center", wrap=True)
        r += 1
        for bp in sorted(lt.plan_t90):
            d = lt.plan_t90[bp]
            _set(ws, r, 1, bp, font=_f(bold=True), align="center")
            _set(ws, r, 2, round(d["ask"]) or None, fmt=MONEY, align="right")
            _set(ws, r, 3, round(d["eff"]) or None, fmt=MONEY, align="right")
            _set(ws, r, 4, f'=IF(OR(B{r}="",B{r}=0),"-",1-C{r}/B{r})', fmt=PCT, align="right")
            _set(ws, r, 5, d["n"] or None, fmt="0", align="center")
            r += 1
        r += 1

    # last-5 new leases per plan (with dates / units / PSF)
    _set(ws, r, 1, "Last 5 New Leases by Floor Plan (rent roll)",
         font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
    for c in range(2, 8):
        ws.cell(r, c).fill = _fill(SECT_FILL)
    r += 1
    for c, h in enumerate(["Floor Plan", "Unit", "Lease Start", "Contract Rent", "Rent PSF (mo.)"], 1):
        _set(ws, r, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center", wrap=True)
    r += 1
    sqft_by_unit = {u.unit: u.sqft for u in rr.units}
    for plan in sorted(lt.recent_new):
        for u in lt.recent_new[plan]:
            sf = sqft_by_unit.get(u.unit, 0) or u.sqft
            psf = round(u.contract_rent / sf, 2) if sf else None
            _set(ws, r, 1, plan, font=_f(bold=True), align="center")
            _set(ws, r, 2, u.unit, align="center")
            _set(ws, r, 3, u.lease_start, fmt=DATEF, align="center")
            _set(ws, r, 4, round(u.contract_rent), fmt=MONEY, align="right")
            _set(ws, r, 5, psf, fmt=MONEY2, align="right")
            r += 1

    widths = [13, 11, 13, 14, 13, 12, 12]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.sheet_view.showGridLines = False


# ===========================================================================
# TAB: HelloData pass-through  (matches model 'HD Dump' A..U)
# ===========================================================================
def write_hellodata(ws, hd: il.HelloData):
    out_headers = ["Property Name", "Floorplan", "Unit", "Floor", "Bedrooms",
                   "Bathrooms", "Partial Bathrooms", "SF", "Available On",
                   "Deposit", "Term", "AMI Level", "Last Asking Rent",
                   "Asking Rent PSF", "Last Effective Rent", "Effective Rent PSF",
                   "On Market Date", "Off Market Date", "Days on Market",
                   "Days Vacant", "Floorplan Mapped"]
    src_keys = ["Property Name", "Floorplan", "Unit", "Floor", "Bedrooms",
                "Bathrooms", "Partial Bathrooms", "SF", "Available On",
                "Deposit", "Term", "Est AMI Level", "Last Asking Rent",
                "Asking Rent PSF", "Last Effective Rent", "Effective Rent PSF",
                "On Market Date", "Off Market Date", "Days on Market", "Days Vacant"]
    for c, h in enumerate(out_headers, 1):
        _set(ws, 1, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL),
             align="center", wrap=True)
    r = 2
    for row in hd.rows:
        for c, key in enumerate(src_keys, 1):
            v = row.get(key, "")
            try:
                v = float(v) if v not in ("", None) and key not in (
                    "Property Name", "Floorplan", "Unit", "Available On",
                    "On Market Date", "Off Market Date", "AMI Level", "Est AMI Level") else v
            except (ValueError, TypeError):
                pass
            _set(ws, r, c, v, font=_f())
        _set(ws, r, 21, il._norm_plan(row.get("Floorplan", "")), font=_f(color=GREEN))
        r += 1
    for c in range(1, 22):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = "C2"
    ws.sheet_view.showGridLines = False


# ===========================================================================
# TAB: Reconciliation
# ===========================================================================
def write_reconciliation(ws, rec: il.Reconciliation, rr: il.RentRoll):
    _title_block(ws, "Reconciliation — Rent Roll ↔ T12",
                 f"Latest T12 month: {rec.latest_month}.  'In Contract Rent?' flags which charges roll into contract rent.")
    r = 4
    _set(ws, r, 1, "Control Tie-Outs (rent roll vs latest T12 month)", font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
    for c in range(2, 7):
        ws.cell(r, c).fill = _fill(SECT_FILL)
    r += 1
    for c, h in enumerate(["Measure", "Rent Roll", "T12 (latest mo. / ann.)", "Variance", "Var %", "Note"], 1):
        _set(ws, r, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), wrap=True, align="center")
    r += 1
    for ln in rec.lines:
        _set(ws, r, 1, ln.label, font=_f(bold=True))
        _set(ws, r, 2, ln.rr_value, fmt=MONEY, align="right")
        _set(ws, r, 3, ln.t12_value, fmt=MONEY, align="right")
        _set(ws, r, 4, f"=B{r}-C{r}", fmt=MONEY, align="right")
        _set(ws, r, 5, f'=IF(C{r}=0,"-",D{r}/C{r})', fmt=PCT, align="right")
        _set(ws, r, 6, ln.note, font=_f(color=GREY, size=9), wrap=True)
        r += 1

    r += 1
    _set(ws, r, 1, "Charge-Code Map (rent roll scheduled, monthly)", font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
    for c in range(2, 7):
        ws.cell(r, c).fill = _fill(SECT_FILL)
    r += 1
    for c, h in enumerate(["Charge Code", "Mapped Code", "Yardi Type", "Monthly $", "In Contract Rent?"], 1):
        _set(ws, r, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center", wrap=True)
    r += 1
    for cc, code, ytype, amt, isc in rec.charge_map:
        _set(ws, r, 1, cc, font=_f())
        _set(ws, r, 2, code, font=_f(bold=True, color=BLUE), align="center")
        _set(ws, r, 3, ytype or "", font=_f(color=GREY, size=9), align="center")
        _set(ws, r, 4, amt, fmt=MONEY, align="right")
        _set(ws, r, 5, "Yes" if isc else "No",
             font=_f(bold=True, color=("008000" if isc else "C00000")), align="center")
        r += 1

    if rec.flags:
        r += 1
        _set(ws, r, 1, "⚑ Flags for Underwriting", font=_f(bold=True, color=WHITE), fill=_fill("C00000"))
        for c in range(2, 7):
            ws.cell(r, c).fill = _fill("C00000")
        r += 1
        for fl in rec.flags:
            _set(ws, r, 1, "•  " + fl, font=_f(color="843C0C"), wrap=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 30
            r += 1

    widths = [26, 14, 14, 14, 16, 40]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.sheet_view.showGridLines = False


# ===========================================================================
# TAB: Trends
# ===========================================================================
def write_trends(ws, tr: il.Trends):
    _title_block(ws, "Trends", "Monthly trajectory + trailing annualizations (T12 / T6 / T3).")
    # monthly table
    r = 4
    _set(ws, r, 1, "Monthly Performance", font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
    for c in range(2, 8):
        ws.cell(r, c).fill = _fill(SECT_FILL)
    r += 1
    cols = ["Month", "EGR", "Operating Exp", "NOI", "Rental Income", "Vacancy", "Loss-to-Lease"]
    for c, h in enumerate(cols, 1):
        _set(ws, r, c, h, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center", wrap=True)
    r += 1
    start = r
    for i, mlabel in enumerate(tr.months):
        _set(ws, r, 1, mlabel, font=_f(bold=True), align="center")
        _set(ws, r, 2, tr.egr_by_month[i], fmt=MONEY, align="right")
        _set(ws, r, 3, tr.opex_by_month[i], fmt=MONEY, align="right")
        _set(ws, r, 4, tr.noi_by_month[i], fmt=MONEY, align="right")
        _set(ws, r, 5, tr.rentinc_by_month[i], fmt=MONEY, align="right")
        _set(ws, r, 6, tr.vac_by_month[i], fmt=MONEY, align="right")
        _set(ws, r, 7, tr.ltl_by_month[i], fmt=MONEY, align="right")
        r += 1
    last = r - 1
    _set(ws, r, 1, "Sum / T12", font=_f(bold=True, color=NAVY))
    for c in range(2, 8):
        L = get_column_letter(c)
        _set(ws, r, c, f"=SUM({L}{start}:{L}{last})", font=_f(bold=True, color=NAVY), fmt=MONEY, align="right")
        ws.cell(r, c).border = DBLTOP
    ws.cell(r, 1).border = DBLTOP
    r += 2

    # annualizations
    _set(ws, r, 1, "Trailing Annualizations", font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
    for c in range(2, 5):
        ws.cell(r, c).fill = _fill(SECT_FILL)
    r += 1
    _set(ws, r, 1, "Metric", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))
    for c, p in enumerate(["T12", "T6", "T3"], 2):
        _set(ws, r, c, p, font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center")
    r += 1
    rowdefs = [("Effective Gross Revenue", "_EGR"), ("Operating Expenses", "_OPEX"),
               ("Net Operating Income", "_NOI")]
    for label, key in rowdefs:
        _set(ws, r, 1, label, font=_f(bold=True))
        for c, p in enumerate(["T12", "T6", "T3"], 2):
            v = tr.periods.get(p, {}).get(key)
            _set(ws, r, c, v if v is not None else "n/a", fmt=MONEY, align="right")
        r += 1

    # notes
    if tr.notes:
        r += 1
        _set(ws, r, 1, "Observations", font=_f(bold=True, color=NAVY))
        r += 1
        for note in tr.notes:
            _set(ws, r, 1, "•  " + note, font=_f(color=BLACK), wrap=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            ws.row_dimensions[r].height = 30
            r += 1

    widths = [20, 14, 14, 14, 14, 14, 14]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.sheet_view.showGridLines = False


# ===========================================================================
# TAB: Codes  (reference + dropdown source)
# ===========================================================================
def write_codes(ws):
    _title_block(ws, "Standardized Code Legend", "Source list for the 'T12 Categorized' dropdown. Col A drives VLOOKUP & SUMIFS.")
    _set(ws, 4, 1, "Code", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center")
    _set(ws, 4, 2, "Category", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))
    _set(ws, 4, 3, "Section", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL), align="center")
    r = 5
    groups = [("EFFECTIVE GROSS REVENUE", am.REVENUE_CODES, "rev"),
              ("OPERATING EXPENSES", am.EXPENSE_CODES, "opex"),
              ("NON-OPERATING REVENUE", am.NONOP_REV_CODES, "nonop-rev"),
              ("NON-OPERATING EXPENSES", am.NONOP_EXP_CODES, "nonop-exp")]
    for gname, codes, sect in groups:
        for code, label in codes:
            _set(ws, r, 1, code, font=_f(bold=True, color=BLUE), align="center")
            _set(ws, r, 2, label, font=_f())
            _set(ws, r, 3, sect, font=_f(color=GREY, size=9), align="center")
            r += 1
    last = r - 1
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    return last


# ===========================================================================
# TAB: Dashboard
# ===========================================================================
def write_dashboard(ws, prop_name, st, rr, mix, rec, tr, lt, has_hd):
    _set(ws, 1, 1, f"{prop_name} — Underwriting Intake", font=_f(bold=True, color=NAVY, size=16))
    nfiles = len(st.files)
    span = f"{st.month_labels[0]} – {st.month_labels[-1]} ({st.n_months} mo" \
           + (f", {nfiles} statements stitched" if nfiles > 1 else "") + ")"
    _set(ws, 2, 1, f"Generated {_dt.date.today():%B %d, %Y}  ·  RedIQ-replacement intake  ·  Operating period {span}",
         font=_f(color=GREY, italic=True))

    occ = sum(1 for u in rr.units if u.occupancy == "Occupied")
    vac = sum(1 for u in rr.units if u.occupancy == "Vacant")
    nonrev = sum(1 for u in rr.units if u.occupancy == "Non-Revenue")
    nunits = len(rr.units)
    tsf = rr.totals.get("sqft", 0)
    mkt = sum(u.market_rent for u in rr.units)
    con = sum(u.contract_rent for u in rr.units)
    newc = sum(1 for u in rr.units if il.classify_lease(u) == "new")
    renc = sum(1 for u in rr.units if il.classify_lease(u) == "renewal")
    t12p = tr.periods.get("T12", {})
    last_hd = next((lt.hd_ask_by_q[i] for i in range(len(lt.quarters) - 1, -1, -1)
                    if lt.hd_ask_by_q[i] > 0), 0)
    last_new = next((lt.new_rent_by_q[i] for i in range(len(lt.quarters) - 1, -1, -1)
                     if lt.new_rent_by_q[i] > 0), 0)

    # ---- snapshot (left, cols A-B) ----
    r = 4
    _set(ws, r, 1, "SNAPSHOT", font=_f(bold=True, color=WHITE), fill=_fill(SECT_FILL))
    ws.cell(r, 2).fill = _fill(SECT_FILL)
    r += 1
    rows = [
        ("Units", nunits, "0"),
        ("Occupied / Vacant / Non-Rev", f"{occ} / {vac} / {nonrev}", None),
        ("Occupancy %", (occ / nunits) if nunits else 0, "0.0%"),
        ("Total Rentable SF", tsf, INT),
        ("New / Renewal leases", f"{newc} / {renc}", None),
        ("Gross Market Rent (mo.)", mkt, MONEY),
        ("Contract Rent incl. amenity (mo.)", con, MONEY),
        ("Loss-to-Lease (mo.)", con - mkt, MONEY),
        ("— Market Rent Indicators —", "", None),
    ]
    if has_hd:
        rows.append(("HelloData asking, latest qtr (mix-wtd)", round(last_hd) if last_hd else "n/a", MONEY if last_hd else None))
    rows.append(("New-lease contract, latest qtr", round(last_new) if last_new else "n/a", MONEY if last_new else None))
    rows += [
        ("— T12 Annualized (most recent 12 mo.) —", "", None),
        ("Effective Gross Revenue (T12)", t12p.get("_EGR"), MONEY),
        ("Operating Expenses (T12)", t12p.get("_OPEX"), MONEY),
        ("Net Operating Income (T12)", t12p.get("_NOI"), MONEY),
        ("Operating Margin (T12)", (t12p.get("_NOI", 0) / t12p.get("_EGR", 1)) if t12p.get("_EGR") else 0, "0.0%"),
    ]
    for label, val, fmt in rows:
        sect = label.startswith("—")
        _set(ws, r, 1, label, font=_f(bold=sect, color=(NAVY if sect else BLACK)))
        if val != "":
            _set(ws, r, 2, val, font=_f(bold=True), fmt=fmt, align="right")
        r += 1
    snap_last = r - 1

    # ---- unit mix (right, cols D+) ----
    um_last = write_unit_mix_block(ws, mix, has_hd, r0=4, c0=4)
    _set(ws, um_last + 1, 4,
         "New-lease = lease start \u2264 move-in (renewals excluded). HD T90/T365 = executed "
         "(off-market) asking/effective; HD90 YoY compares the trailing 90 days to the same window a year ago.",
         font=_f(color=GREY, italic=True, size=8))

    # ---- flags (full width, below both) ----
    r = max(snap_last, um_last + 1) + 2
    _set(ws, r, 1, "⚑ UNDERWRITING FLAGS", font=_f(bold=True, color=WHITE), fill=_fill("C00000"))
    for c in range(2, 13):
        ws.cell(r, c).fill = _fill("C00000")
    r += 1
    flaglist = list(getattr(st, "overlap_flags", [])) + list(getattr(st, "granularity_flags", [])) \
        + list(rec.flags) + list(lt.notes) + list(tr.notes)
    if not flaglist:
        _set(ws, r, 1, "No automatic flags raised — review categorization and reconciliation tabs.", font=_f(color=GREY, italic=True))
        r += 1
    else:
        for fl in flaglist:
            _set(ws, r, 1, "•  " + fl, font=_f(color="843C0C"), wrap=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
            ws.row_dimensions[r].height = 28
            r += 1

    # ---- how to use / paste targets ----
    r += 1
    _set(ws, r, 1, "HOW TO USE → paste into the TMG acquisition model", font=_f(bold=True, color=WHITE), fill=_fill(HDR_FILL))
    for c in range(2, 13):
        ws.cell(r, c).fill = _fill(HDR_FILL)
    r += 1
    steps = [
        "1.  Review 'T12 Categorized' col A (amber). Every GL line appears once with all months "
        "flowing across; edit a code via the dropdown to re-map. Category (col B) and OS Summary roll "
        "up automatically." + ("  Overlapping months are sourced from the freshest statement so nothing "
                               "double-counts." if nfiles > 1 else ""),
        "2.  'OS Summary' is the standardized chart of accounts (live SUMIFS, most recent 12 months). "
        "Copy A1:Q77 → model 'OS Summary Dump' A1 (Paste Special → Values).",
        "3.  'T12 Categorized' → copy Code, Category, Line Item + the most recent 12 month columns "
        "(NOT the Total column) → model 'T12 Dump' A2.",
        "4.  'Rent Roll (One-Line)' → copy the core columns A2:M(last) → model 'RR Dump' A2.  "
        "(The per-unit charge-code columns to the right of the gap are reference detail, not a model paste target.)",
        ("5.  'HelloData' → copy A2:U(last) → model 'HD Dump' A2." if has_hd else "5.  (No HelloData provided — skip HD Dump.)"),
        "6.  The unit mix (with new-lease + HD T90/T365/YoY market-rent indicators) is on this Dashboard; "
        "check 'Lease Trend' (seasonality / market-rent trajectory) and 'Reconciliation' (contract-rent ↔ AGPR "
        "tie-out) before underwriting. Paste targets only — do NOT populate the model itself from here.",
    ]
    for s in steps:
        _set(ws, r, 1, s, font=_f(), wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        ws.row_dimensions[r].height = 32
        r += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 2
    ws.sheet_view.showGridLines = False


# ===========================================================================
# DRIVER
# ===========================================================================
def build(t12_paths, rr_path, hd_path=None, prop_name=None, out_path=None):
    if isinstance(t12_paths, (str, bytes)):
        t12_paths = [t12_paths]
    st = il.stitch_statements(list(t12_paths))
    rr = il.parse_rent_roll(rr_path)
    hd = il.parse_hellodata(hd_path) if hd_path else None
    if not prop_name:
        prop_name = (st.title or os.path.splitext(os.path.basename(rr_path))[0]).strip()
    recent_new = il.recent_new_leases(rr, 5)
    t90 = il.hellodata_t90(hd)
    mix = il.build_unit_mix(rr, hd, recent_new, t90)
    rec = il.reconcile(st, rr)
    tr = il.build_trends(st, rr)
    lt = il.build_lease_trend(rr, hd)

    n = st.n_months

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_t12 = wb.create_sheet("T12 Categorized")
    ws_os = wb.create_sheet("OS Summary")
    ws_rr = wb.create_sheet("Rent Roll (One-Line)")
    ws_lt = wb.create_sheet("Lease Trend")
    ws_hd = wb.create_sheet("HelloData") if hd else None
    ws_rec = wb.create_sheet("Reconciliation")
    ws_tr = wb.create_sheet("Trends")
    ws_codes = wb.create_sheet("Codes")

    codes_last = write_codes(ws_codes)
    wb.defined_names["CodeList"] = DefinedName("CodeList", attr_text=f"Codes!$A$5:$A${codes_last}")

    write_t12_categorized(ws_t12, st, codes_last)
    last12 = list(range(max(0, n - 12), n))
    sub = "Adjusted Amounts ($ totals) — most recent 12 months · model paste target (A1:Q77)" if n > 12 \
        else "Adjusted Amounts ($ totals) — model paste target (A1:Q77)"
    write_os_grid(ws_os, st, prop_name, last12,
                  title="Operating Statement — Standardized (live rollup)", subtitle=sub)
    write_rent_roll(ws_rr, rr, hd)
    write_lease_trend(ws_lt, lt, rr, has_hd=bool(hd))
    if ws_hd is not None:
        write_hellodata(ws_hd, hd)
    write_reconciliation(ws_rec, rec, rr)
    write_trends(ws_tr, tr)
    write_dashboard(ws_dash, prop_name, st, rr, mix, rec, tr, lt, has_hd=bool(hd))

    if not out_path:
        safe = "".join(ch if ch.isalnum() else "_" for ch in prop_name).strip("_")
        out_path = f"/mnt/user-data/outputs/{safe}__Underwriting_Intake.xlsx"
    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser(description="Build a RedIQ-replacement underwriting intake workbook.")
    ap.add_argument("--t12", required=True, nargs="+",
                    help="One or more T12 / monthly operating statements (stitched into one continuous series).")
    ap.add_argument("--rr", required=True, help="Rent roll (use the most recent).")
    ap.add_argument("--hd", default=None, help="HelloData unit-details CSV (optional).")
    ap.add_argument("--name", default=None)
    ap.add_argument("--out", default=None)
    a = ap.parse_args()
    out = build(a.t12, a.rr, a.hd, a.name, a.out)
    print(out)


if __name__ == "__main__":
    main()
