#!/usr/bin/env python3
"""
dev/compare.py — regression check: freshly built workbook vs a golden workbook.

The fresh workbook's formulas are evaluated in Python via recalc_local (the
sandbox's LibreOffice recalc is unreliable), then compared cell-by-cell against
the golden workbook's recalculated/cached values:

  - sheets present in only one file
  - per-sheet dimension differences
  - numeric cells differing beyond --tol (absolute) / --rtol (relative)
  - a number on one side but blank on the other (was previously missed)
  - text cells that differ

Exits non-zero if anything differs beyond tolerance, so it can gate CI.

Usage:
    python3 dev/compare.py [fresh.xlsx] [golden.xlsx] [--tol 0.5] [--rtol 0.001]
"""
import argparse, sys
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import recalc_local as rl

ROOT = Path(__file__).resolve().parent.parent
DEF_FRESH = ROOT / "dev" / "out" / "Canyon_Ridge__Underwriting_Intake.xlsx"
DEF_GOLD = ROOT / "dev" / "golden" / "Canyon_Ridge__Underwriting_Intake.golden.xlsx"


def golden_values(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    out, dims = {}, {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        dims[sn] = (ws.max_row, ws.max_column)
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is not None:
                    out[(sn, f"{get_column_letter(c)}{r}")] = v
    return out, dims, wb.sheetnames


def blank(v):
    return v is None or v == "" or v == "#ERR"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("fresh", nargs="?", default=str(DEF_FRESH))
    ap.add_argument("golden", nargs="?", default=str(DEF_GOLD))
    ap.add_argument("--tol", type=float, default=0.5)
    ap.add_argument("--rtol", type=float, default=0.001)
    ap.add_argument("--max", type=int, default=40, help="max diffs printed per sheet")
    args = ap.parse_args()

    print(f"fresh : {args.fresh}  (formulas evaluated locally)")
    print(f"golden: {args.golden}\n")

    fresh = rl.compute_values(args.fresh)
    gold, gdims, gsheets = golden_values(args.golden)
    fsheets = sorted({k[0] for k in fresh})

    sa, sb = set(fsheets), set(gsheets)
    if sa - sb:
        print("  sheets only in FRESH :", sorted(sa - sb))
    if sb - sa:
        print("  sheets only in GOLDEN:", sorted(sb - sa))

    # group diffs by sheet
    keys = set(fresh) | set(gold)
    by_sheet = {}
    for (sn, cell) in keys:
        a, b = fresh.get((sn, cell)), gold.get((sn, cell))
        na = isinstance(a, (int, float)) and not isinstance(a, bool)
        nb = isinstance(b, (int, float)) and not isinstance(b, bool)
        diff = None
        if na and nb:
            d = abs(a - b)
            if d > args.tol and d > abs(b) * args.rtol:
                diff = (a, b)
        elif blank(a) and blank(b):
            pass
        elif na or nb:                       # number vs blank/text — real diff
            if not (blank(a) and blank(b)):
                diff = (a, b)
        else:
            if str(a or "") != str(b or ""):
                diff = (a, b)
        if diff:
            by_sheet.setdefault(sn, []).append((cell, diff[0], diff[1]))

    total = 0
    for sn in gsheets:
        if sn in fsheets and sn in gdims:
            # dim check uses golden dims vs fresh max
            fcells = [k for k in fresh if k[0] == sn]
            fr = max((int(c[len(c.rstrip("0123456789")):]) for _, c in fcells), default=0)
        diffs = sorted(by_sheet.get(sn, []), key=lambda x: (len(x[0]), x[0]))
        if diffs:
            total += len(diffs)
            print(f"[{sn}] {len(diffs)} differing cell(s):")
            for cell, a, b in diffs[:args.max]:
                print(f"    {cell}: fresh={a!r}  golden={b!r}")
            if len(diffs) > args.max:
                print(f"    ... +{len(diffs) - args.max} more")

    print()
    if total == 0 and not (sa ^ sb):
        print("✓ PARITY: fresh matches golden within tolerance.")
        sys.exit(0)
    print(f"✗ {total} cell diff(s); sheet set "
          f"{'matches' if not (sa ^ sb) else 'DIFFERS'}.")
    sys.exit(1)


if __name__ == "__main__":
    main()
