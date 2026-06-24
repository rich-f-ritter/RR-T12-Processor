#!/usr/bin/env python3
"""
dev/run.py — one-command build + recalc for the rr-t12-processor skill.

Builds the intake workbook from the Canyon Ridge fixtures (or any files you
pass through), then runs the mandatory recalc step exactly as the skill spec
requires. This is the local dev loop; the shipped skill is unchanged.

Recalc modes (--recalc):
    local   (default) deterministic Python evaluation via recalc_local — works
            everywhere, reports formula/error counts, leaves formulas intact.
    office  the shipped pipeline's recalc.py (LibreOffice). This is what runs in
            the Claude app; it is unreliable in this sandbox (the macro hangs).
    none    build only.

Usage:
    python3 dev/run.py
    python3 dev/run.py --recalc office
    python3 dev/run.py --rr <rentroll.xlsx> --t12 <a.xlsx> <b.xlsx> ... --hd <h.csv>
"""
import argparse, os, subprocess, sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SKILL = ROOT / ".claude" / "skills" / "rr-t12-processor"
TESTDATA = ROOT / "dev" / "testdata"
OUT = ROOT / "dev" / "out"
sys.path.insert(0, str(ROOT / "dev"))


def find_office_recalc():
    env = os.environ.get("RECALC_PY")
    if env and Path(env).exists():
        return env
    canonical = "/mnt/skills/public/xlsx/scripts/recalc.py"
    return canonical if Path(canonical).exists() else None


def local_recalc(path):
    import recalc_local as rl
    vals = rl.compute_values(path)
    errs = [f"{s}!{c}" for (s, c), v in vals.items() if v == "#ERR"]
    formulas = sum(1 for _ in _formula_cells(path))
    print(f"  local recalc: {formulas} formulas, {len(errs)} error(s)"
          + (f" -> {errs[:10]}" if errs else ""))
    return len(errs)


def _formula_cells(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=False)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    yield cell


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--t12", nargs="+",
                    default=[str(TESTDATA / "Canyon_Ridge__T12_2026.03.xlsx"),
                             str(TESTDATA / "Canyon_Ridge__T12_2026.05.xlsx")])
    ap.add_argument("--rr", nargs="+",
                    default=[str(TESTDATA / "Canyon_Ridge__Rent_Roll_2026.05.31.xlsx")])
    ap.add_argument("--hd",
                    default=str(TESTDATA / "hellodata_unit_details_2026.06.17.csv"))
    ap.add_argument("--charge-codes", dest="charge_codes", default=None)
    ap.add_argument("--hd-fee-offset", dest="hd_fee_offset", default=None)
    ap.add_argument("--name", default="Canyon Ridge Apartments")
    ap.add_argument("--out", default=str(OUT / "Canyon_Ridge__Underwriting_Intake.xlsx"))
    ap.add_argument("--recalc", choices=["local", "office", "none"], default="local")
    ap.add_argument("--timeout", default="120")
    args = ap.parse_args()

    OUT.mkdir(parents=True, exist_ok=True)

    build = [sys.executable, str(SKILL / "scripts" / "build_intake.py"),
             "--t12", *args.t12, "--rr", *args.rr, "--name", args.name,
             "--out", args.out]
    if args.hd:
        build += ["--hd", args.hd]
    if args.charge_codes:
        build += ["--charge-codes", args.charge_codes]
    if args.hd_fee_offset:
        build += ["--hd-fee-offset", str(args.hd_fee_offset)]

    print("→ build:", " ".join(build))
    r = subprocess.run(build)
    if r.returncode != 0:
        sys.exit(r.returncode)

    if args.recalc == "none":
        print("✓ built (recalc skipped):", args.out)
        return

    if args.recalc == "local":
        n_err = local_recalc(args.out)
        print("✓ built + locally recalced:", args.out)
        sys.exit(1 if n_err else 0)

    recalc = find_office_recalc()
    if not recalc:
        print("✗ recalc.py not found. Set $RECALC_PY or run where "
              "/mnt/skills/public/xlsx/scripts/recalc.py exists.", file=sys.stderr)
        sys.exit(2)
    print("→ office recalc:", recalc, args.out, args.timeout)
    r = subprocess.run([sys.executable, recalc, args.out, args.timeout])
    sys.exit(r.returncode)


if __name__ == "__main__":
    main()
