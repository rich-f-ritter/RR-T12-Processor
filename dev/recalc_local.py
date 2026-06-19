#!/usr/bin/env python3
"""
dev/recalc_local.py — deterministic, LibreOffice-free recalc for the intake workbook.

LibreOffice headless recalc is unreliable in some sandboxes (the macro hangs and
the official recalc.py swallows the timeout, leaving formula cells with no cached
value). This module instead evaluates the *closed* formula vocabulary that
build_intake.py emits — SUM, SUMIFS, VLOOKUP, IFERROR, IF, OR, comparisons and
arithmetic over cell/range references — directly in Python.

It is NOT a general Excel engine; it covers exactly what this skill writes. If
build_intake.py grows a new function, add it to FUNCS below and the parser will
pick it up.

API:
    compute_values(path) -> { (sheet_name, "A1"): value }   # every cell, formulas evaluated
CLI:
    python3 dev/recalc_local.py in.xlsx [out.xlsx]
        evaluates and writes a values-only copy (formulas replaced by results).
        Default out: <in>.recalc.xlsx
"""
from __future__ import annotations
import re, sys
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


class XlError(Exception):
    """An Excel-style error (e.g. #N/A) — caught by IFERROR."""


# ---------------------------------------------------------------------------
# tokenizer
# ---------------------------------------------------------------------------
_TOKEN_RE = re.compile(r"""
    \s*(?:
      (?P<str>"(?:[^"]|"")*")
    | (?P<sheet>'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)\!     # sheet qualifier (followed by !)
    | (?P<num>\d+(?:\.\d+)?(?:[eE][+-]?\d+)?)
    | (?P<cell>\$?[A-Za-z]{1,3}\$?\d+)
    | (?P<colrange>\$?[A-Za-z]{1,3}:\$?[A-Za-z]{1,3})   # full-column range A:A
    | (?P<func>[A-Za-z_][A-Za-z0-9_.]*)\(
    | (?P<op><=|>=|<>|[-+*/=<>():,&])
    )
""", re.VERBOSE)


def _tokenize(s):
    toks, i = [], 0
    while i < len(s):
        m = _TOKEN_RE.match(s, i)
        if not m or m.end() == i:
            if s[i:].strip() == "":
                break
            raise ValueError(f"cannot tokenize at {s[i:]!r} in {s!r}")
        i = m.end()
        kind = m.lastgroup
        val = m.group(kind)
        toks.append((kind, val))
    toks.append(("end", None))
    return toks


# ---------------------------------------------------------------------------
# parser (recursive descent) -> AST of nested tuples/python values
# ---------------------------------------------------------------------------
class _Parser:
    def __init__(self, toks):
        self.toks = toks
        self.p = 0

    def peek(self):
        return self.toks[self.p]

    def next(self):
        t = self.toks[self.p]; self.p += 1; return t

    def expect(self, val):
        k, v = self.next()
        if v != val:
            raise ValueError(f"expected {val!r} got {v!r}")

    def parse(self):
        node = self.expr()
        return node

    def expr(self):
        return self.comparison()

    def comparison(self):
        left = self.concat()
        while self.peek()[0] == "op" and self.peek()[1] in ("=", "<>", "<", ">", "<=", ">="):
            op = self.next()[1]
            right = self.concat()
            left = ("cmp", op, left, right)
        return left

    def concat(self):
        left = self.add()
        while self.peek()[1] == "&":
            self.next(); right = self.add(); left = ("concat", left, right)
        return left

    def add(self):
        left = self.mul()
        while self.peek()[1] in ("+", "-"):
            op = self.next()[1]; right = self.mul(); left = ("bin", op, left, right)
        return left

    def mul(self):
        left = self.unary()
        while self.peek()[1] in ("*", "/"):
            op = self.next()[1]; right = self.unary(); left = ("bin", op, left, right)
        return left

    def unary(self):
        if self.peek()[1] == "-":
            self.next(); return ("neg", self.unary())
        return self.primary()

    def primary(self):
        kind, val = self.peek()
        if kind == "num":
            self.next(); return ("num", float(val))
        if kind == "str":
            self.next(); return ("str", val[1:-1].replace('""', '"'))
        if kind == "func":
            self.next()                      # consumes NAME(
            name = val.upper()
            args = []
            if self.peek()[1] != ")":
                args.append(self.expr())
                while self.peek()[1] == ",":
                    self.next(); args.append(self.expr())
            self.expect(")")
            return ("func", name, args)
        if kind == "sheet":
            self.next()
            sheet = val.strip("'")           # group excludes the trailing !
            return self.ref(sheet)
        if kind in ("cell", "colrange"):
            return self.ref(None)
        if val == "(":
            self.next(); e = self.expr(); self.expect(")"); return e
        raise ValueError(f"unexpected token {val!r}")

    def ref(self, sheet):
        kind, val = self.next()
        if kind == "colrange":
            a, b = val.split(":")
            return ("range", sheet, a.replace("$", ""), None, b.replace("$", ""), None)
        # cell, maybe a range cell:cell
        c1 = val.replace("$", "")
        if self.peek()[1] == ":":
            self.next()
            _, v2 = self.next()
            c2 = v2.replace("$", "")
            ca, ra = _split(c1); cb, rb = _split(c2)
            return ("range", sheet, ca, ra, cb, rb)
        ca, ra = _split(c1)
        return ("cell", sheet, ca, ra)


def _split(cellref):
    m = re.match(r"([A-Za-z]+)(\d+)", cellref)
    return m.group(1), int(m.group(2))


# ---------------------------------------------------------------------------
# evaluator
# ---------------------------------------------------------------------------
class Recalc:
    def __init__(self, path):
        self.wb = openpyxl.load_workbook(path, data_only=False)
        self.cache = {}     # (sheet, col, row) -> value
        self.busy = set()
        self.ast = {}       # memoized parse per formula string

    # cell access ----------------------------------------------------------
    def cell_value(self, sheet, col, row):
        key = (sheet, col, row)
        if key in self.cache:
            return self.cache[key]
        if key in self.busy:
            raise XlError(f"circular {sheet}!{col}{row}")
        ws = self.wb[sheet]
        raw = ws.cell(row=row, column=column_index_from_string(col)).value
        if isinstance(raw, str) and raw.startswith("="):
            self.busy.add(key)
            try:
                node = self.ast.get(raw)
                if node is None:
                    node = _Parser(_tokenize(raw[1:])).parse()
                    self.ast[raw] = node
                val = self.eval(node, sheet)
            finally:
                self.busy.discard(key)
            self.cache[key] = val
            return val
        self.cache[key] = raw
        return raw

    # AST evaluation -------------------------------------------------------
    def eval(self, node, ctx_sheet):
        t = node[0]
        if t == "num":
            return node[1]
        if t == "str":
            return node[1]
        if t == "neg":
            return -_num(self.eval(node[1], ctx_sheet))
        if t == "concat":
            return _txt(self.eval(node[1], ctx_sheet)) + _txt(self.eval(node[2], ctx_sheet))
        if t == "bin":
            a = _num(self.eval(node[2], ctx_sheet)); b = _num(self.eval(node[3], ctx_sheet))
            op = node[1]
            if op == "+": return a + b
            if op == "-": return a - b
            if op == "*": return a * b
            if op == "/":
                if b == 0: raise XlError("#DIV/0!")
                return a / b
        if t == "cmp":
            return _compare(node[1], self.eval(node[2], ctx_sheet), self.eval(node[3], ctx_sheet))
        if t == "cell":
            sheet = node[1] or ctx_sheet
            return self.cell_value(sheet, node[2], node[3])
        if t == "range":
            return ("RANGE", node[1] or ctx_sheet, node[2], node[3], node[4], node[5])
        if t == "func":
            return self.call(node[1], node[2], ctx_sheet)
        raise ValueError(f"bad node {node}")

    def call(self, name, args, ctx_sheet):
        if name == "FALSE": return False
        if name == "TRUE":  return True
        if name == "IF":
            cond = self.eval(args[0], ctx_sheet)
            return self.eval(args[1], ctx_sheet) if _truth(cond) else (
                self.eval(args[2], ctx_sheet) if len(args) > 2 else False)
        if name == "OR":
            return any(_truth(self.eval(a, ctx_sheet)) for a in args)
        if name == "IFERROR":
            try:
                v = self.eval(args[0], ctx_sheet)
                if isinstance(v, str) and v.startswith("#"):
                    raise XlError(v)
                return v
            except XlError:
                return self.eval(args[1], ctx_sheet)
        if name == "SUM":
            tot = 0.0
            for a in args:
                for v in self._cells(a, ctx_sheet):
                    if isinstance(v, (int, float)):
                        tot += v
            return tot
        if name == "SUMIFS":
            sum_cells = list(self._cells(args[0], ctx_sheet, raw=True))
            crit_cells = list(self._cells(args[1], ctx_sheet, raw=True))
            criterion = self.eval(args[2], ctx_sheet)
            tot = 0.0
            for sv, cv in zip(sum_cells, crit_cells):
                if _eq_criterion(cv, criterion) and isinstance(sv, (int, float)):
                    tot += sv
            return tot
        if name == "VLOOKUP":
            key = self.eval(args[0], ctx_sheet)
            rng = self.eval(args[1], ctx_sheet)
            col_idx = int(self.eval(args[2], ctx_sheet))
            grid = self._grid(rng)
            for rowcells in grid:
                if rowcells and _eq_criterion(rowcells[0], key):
                    return rowcells[col_idx - 1]
            raise XlError("#N/A")
        raise ValueError(f"unsupported function {name}")

    # range helpers --------------------------------------------------------
    def _bounds(self, rng):
        _, sheet, c1, r1, c2, r2 = rng
        ws = self.wb[sheet]
        ci1, ci2 = column_index_from_string(c1), column_index_from_string(c2)
        if r1 is None:                      # full-column range
            r1, r2 = 1, ws.max_row
        return sheet, min(ci1, ci2), max(ci1, ci2), min(r1, r2), max(r1, r2)

    def _cells(self, arg_node, ctx_sheet, raw=False):
        v = self.eval(arg_node, ctx_sheet) if not isinstance(arg_node, tuple) or arg_node[0] != "range" \
            else self.eval(arg_node, ctx_sheet)
        if isinstance(v, tuple) and v and v[0] == "RANGE":
            sheet, ci1, ci2, r1, r2 = self._bounds(v)
            for r in range(r1, r2 + 1):
                for c in range(ci1, ci2 + 1):
                    yield self.cell_value(sheet, get_column_letter(c), r)
        else:
            yield v

    def _grid(self, rng):
        sheet, ci1, ci2, r1, r2 = self._bounds(rng)
        out = []
        for r in range(r1, r2 + 1):
            out.append([self.cell_value(sheet, get_column_letter(c), r)
                        for c in range(ci1, ci2 + 1)])
        return out

    # public ---------------------------------------------------------------
    def compute_all(self):
        result = {}
        for sheet in self.wb.sheetnames:
            ws = self.wb[sheet]
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    col = get_column_letter(c)
                    try:
                        v = self.cell_value(sheet, col, r)
                    except XlError:
                        v = "#ERR"
                    result[(sheet, f"{col}{r}")] = v
        return result

    def write_values(self, out_path):
        wb = openpyxl.load_workbook(self.wb_path) if hasattr(self, "wb_path") else self.wb
        for sheet in self.wb.sheetnames:
            ws = self.wb[sheet]
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        col = get_column_letter(c)
                        try:
                            v = self.cell_value(sheet, col, r)
                        except XlError:
                            v = None
                        cell.value = None if (isinstance(v, str) and v == "") else v
        self.wb.save(out_path)


# ---------------------------------------------------------------------------
# value coercion helpers
# ---------------------------------------------------------------------------
def _num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        return float(v)
    raise XlError(f"#VALUE! ({v!r})")


def _txt(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _truth(v):
    if isinstance(v, bool): return v
    if isinstance(v, (int, float)): return v != 0
    if v is None or v == "": return False
    return True


def _compare(op, a, b):
    # blank/None behaves like "" or 0 depending on the other operand
    if a is None: a = "" if isinstance(b, str) else 0
    if b is None: b = "" if isinstance(a, str) else 0
    if isinstance(a, str) or isinstance(b, str):
        a, b = str(a), str(b)
    if op == "=":  return a == b
    if op == "<>": return a != b
    if op == "<":  return a < b
    if op == ">":  return a > b
    if op == "<=": return a <= b
    if op == ">=": return a >= b


def _eq_criterion(cell, criterion):
    if cell is None:
        cell = ""
    if isinstance(criterion, float) and criterion.is_integer():
        criterion = int(criterion)
    return str(cell).strip() == str(criterion).strip()


# ---------------------------------------------------------------------------
def compute_values(path):
    return Recalc(path).compute_all()


def main():
    if len(sys.argv) < 2:
        print("usage: recalc_local.py in.xlsx [out.xlsx]"); sys.exit(1)
    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else str(Path(inp).with_suffix(".recalc.xlsx"))
    rc = Recalc(inp)
    rc.write_values(out)
    print(f"✓ recalced (local): {out}")


if __name__ == "__main__":
    main()
