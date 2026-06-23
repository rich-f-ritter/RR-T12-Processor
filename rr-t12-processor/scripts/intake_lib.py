"""
intake_lib.py
=============
Parsing + analysis layer for the RedIQ-replacement underwriting intake engine.

Pure data work only -- NO Excel writing here (that lives in build_intake.py).
Everything is defensive: real-world T12s / rent rolls vary, so each parser
isolates "structural" assumptions (where the header row is, which column holds
what) into small, override-able helpers and degrades gracefully.

Public entry points:
    parse_t12(path)            -> T12 object (months, rows, sections, totals)
    parse_rent_roll(path)      -> RentRoll object (units, footer summaries)
    parse_hellodata(path)      -> HelloData object (rows, floorplan bed/bath map)
    build_unit_mix(rentroll, hd)        -> list[UnitMixRow]
    reconcile(t12, rentroll)            -> Reconciliation object
    build_trends(t12, rentroll)         -> Trends object
"""
from __future__ import annotations
import re, csv, datetime as _dt
from dataclasses import dataclass, field
from typing import Optional
import openpyxl

import account_map as am


# ===========================================================================
# workbook loader — reads .xlsx natively and legacy binary .xls via xlrd, behind a
# uniform (openpyxl-like) interface so every parser works regardless of file format.
# Extension is not trusted (operators mislabel .xlsx as .XLS and vice versa): we try
# openpyxl first and fall back to xlrd on failure.
# ===========================================================================
class _Cell:
    __slots__ = ("value",)

    def __init__(self, value):
        self.value = value


class _XlsSheet:
    def __init__(self, sheet):
        self._s = sheet
        self.title = sheet.name
        self.max_row = sheet.nrows
        self.max_column = sheet.ncols

    def cell(self, row, column):
        r, c = row - 1, column - 1
        if 0 <= r < self._s.nrows and 0 <= c < self._s.ncols:
            v = self._s.cell_value(r, c)
            return _Cell(None if v == "" else v)
        return _Cell(None)

    def iter_rows(self):
        for r in range(1, self.max_row + 1):
            yield [self.cell(r, c) for c in range(1, self.max_column + 1)]


class _XlsBook:
    def __init__(self, path):
        import xlrd
        self._b = xlrd.open_workbook(path)
        self.worksheets = [_XlsSheet(self._b.sheet_by_index(i)) for i in range(self._b.nsheets)]
        self.sheetnames = list(self._b.sheet_names())

    def __getitem__(self, name):
        return next(w for w in self.worksheets if w.title == name)


def _load_workbook(path, data_only=True):
    """Return a workbook (openpyxl for .xlsx, an xlrd-backed adapter for legacy .xls)."""
    try:
        return openpyxl.load_workbook(path, data_only=data_only)
    except Exception:
        return _XlsBook(path)


# ===========================================================================
# small helpers
# ===========================================================================
def _num(v):
    """Coerce a cell to float, treating blanks/dashes/parentheses sanely."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-", "--", "—"):
        return 0.0
    neg = s.startswith("(") and s.endswith(")")
    s = s.replace("(", "").replace(")", "").replace("$", "").replace(",", "").replace("%", "").strip()
    try:
        f = float(s)
        return -f if neg else f
    except ValueError:
        return 0.0


def _s(v):
    return "" if v is None else str(v).strip()


def _is_glnum(s):
    """True if a string looks like a GL account number — bare ('4001', '6504'),
    segmented ('4020-5100'), or long ('40200500')."""
    return bool(re.fullmatch(r"\d{3,8}(?:[-.]\d{1,6})*", _s(s)))


def _is_rollup(name, gl):
    """True if a GL row is a subtotal/parent rollup rather than a postable leaf.
    Detected by an unambiguous 'Total/Net/Subtotal' name prefix or the Yardi-tree
    rollup account suffix (…-098 / …-099). Deliberately does NOT key on words like
    'potential', which appear in real leaf line names (e.g. 'Gross Potential Rent')."""
    n = _s(name).lower()
    # Some exports repeat the GL number inside the description cell
    # ("49999-999 - Total Income"), which would defeat the ^total anchor. Strip a
    # leading account-number prefix before testing the name prefix.
    n = re.sub(r"^\s*\d{3,}[-_]\d+\s*[-–—:]+\s*", "", n)
    if re.match(r"^(total\b|net\b|subtotal|sub-total)", n):
        return True
    return bool(re.search(r"-0?9[89]$", _s(gl)))


def _is_date(v):
    return isinstance(v, (_dt.datetime, _dt.date))


def _month_label(v):
    """Normalize a month header cell to a 'Mon YYYY' label + sort key."""
    if _is_date(v):
        return v.strftime("%b %Y"), (v.year, v.month)
    s = _s(v)
    # try to parse 'Jun 2025', 'Jun-2025', '2025-06', etc.
    for fmt in ("%b %Y", "%b-%Y", "%B %Y", "%m/%Y", "%Y-%m"):
        try:
            d = _dt.datetime.strptime(s, fmt)
            return d.strftime("%b %Y"), (d.year, d.month)
        except ValueError:
            continue
    return s, (9999, 99)


# ===========================================================================
# T12
# ===========================================================================
@dataclass
class T12Row:
    rtype: str            # 'header' | 'line' | 'subtotal'
    name: str
    section: str = ""
    side: str = ""        # 'rev' | 'exp' (line rows only)
    acct: str = ""
    code: str = ""        # auto-assigned standardized code (line rows only)
    values: list = field(default_factory=list)   # 12 monthly floats
    total: float = 0.0


@dataclass
class T12:
    months: list                  # list of (label, sortkey)
    rows: list                    # list[T12Row]
    sheet_name: str = ""
    title: str = ""

    @property
    def month_labels(self):
        return [m[0] for m in self.months]

    @property
    def n_months(self):
        return len(self.months)

    def lines(self):
        return [r for r in self.rows if r.rtype == "line"]

    def code_total(self, code, month_idx=None):
        """Sum of all line rows with `code` (a single month or full total)."""
        tot = 0.0
        for r in self.lines():
            if r.code != code:
                continue
            if month_idx is None:
                tot += r.total
            elif month_idx < len(r.values):
                tot += r.values[month_idx]
        return tot


# revenue-side native sections (everything else is treated as expense)
_REV_SECTIONS = re.compile(
    r"gross potential|rental adjustment|other income|other revenue|"
    r"income write.?off|gross rent|rental income|total income|total rental",
    re.I,
)


def parse_t12(path: str) -> T12:
    wb = _load_workbook(path)
    # pick the sheet that looks like an income statement (most numeric rows)
    ws = wb[wb.sheetnames[0]]
    best, best_score = ws, -1
    for sn in wb.sheetnames:
        w = wb[sn]
        score = sum(
            1 for r in range(1, min(w.max_row, 60) + 1)
            for c in range(2, min(w.max_column, 16) + 1)
            if isinstance(w.cell(r, c).value, (int, float))
        )
        if score > best_score:
            best, best_score, = w, score
    ws = best

    title = _s(ws.cell(3, 1).value) or _s(ws.cell(2, 1).value)

    # --- locate header row (the one containing month columns) ---
    hdr_row = None
    for r in range(1, min(ws.max_row, 20) + 1):
        rowvals = [ws.cell(r, c).value for c in range(2, min(ws.max_column, 16) + 1)]
        n_dates = sum(1 for v in rowvals if _is_date(v))
        n_monthish = sum(
            1 for v in rowvals
            if _is_date(v) or re.match(r"^[A-Za-z]{3,9}[ \-]\d{4}$", _s(v))
        )
        if n_monthish >= 3:
            hdr_row = r
            break
    if hdr_row is None:
        hdr_row = 7  # sensible Yardi default

    # --- month columns: the contiguous run of month-labelled header cells
    #     (may start at col B, col C, ... — wherever the labels end) ---
    def _ismonth(v):
        return _is_date(v) or bool(re.match(r"^[A-Za-z]{3,9}[ \-]\d{4}$", _s(v)))
    months, month_cols = [], []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hdr_row, c).value
        if _ismonth(v):
            lbl, key = _month_label(v)
            months.append((lbl, key)); month_cols.append(c)
        elif months:
            break  # run ended (next col is Total/blank)
    if not month_cols:
        return T12(months=[], rows=[], sheet_name=ws.title, title=title)

    # --- detect the description column and the GL-account-number column.
    #     Universal across layouts:
    #       * description = the text column immediately LEFT of the month block
    #         (where the line names always sit).
    #       * GL number   = the column with the most DISTINCT "code-like" account
    #         values (3–5 digit codes or dashed codes) — searched on the LEFT
    #         first, else on the RIGHT. Distinctness + the code-like shape skip
    #         constant property-id columns (e.g. '9521', '14108') and the
    #         monetary amount / sign-flip helper columns. ---
    first_month = month_cols[0]
    nmonths = len(month_cols)
    total_cols = {c for c in range(1, ws.max_column + 1)
                  if _s(ws.cell(hdr_row, c).value).lower() in ("total", "ytd", "annual")}
    left = list(range(1, first_month))
    right = [c for c in range(first_month + nmonths, ws.max_column + 1) if c not in total_cols]
    data_rows = [r for r in range(hdr_row + 1, ws.max_row + 1)
                 if any(isinstance(ws.cell(r, c).value, (int, float)) for c in month_cols)]

    def _codelike(s):
        return _is_glnum(s) and ("-" in s or (s.isdigit() and 3 <= len(s) <= 6))

    def _scan(cols):
        gl, txt, tl = {}, {}, {}
        for c in cols:
            gs = set(); t = L = 0
            for r in data_rows:
                s = _s(ws.cell(r, c).value)
                if not s:
                    continue
                if _codelike(s):
                    gs.add(s)
                if re.search(r"[A-Za-z]{2,}", s):
                    t += 1; L += len(s)
            gl[c], txt[c], tl[c] = len(gs), t, L
        return gl, txt, tl

    lgl, ltxt, ltl = _scan(left)
    desc_col = max(left, key=lambda c: (ltxt[c], ltl[c])) if left else 1
    gl_col = None
    if left and max(lgl.values(), default=0) >= 2:
        gl_col = max(left, key=lambda c: lgl[c])
    elif right:
        rgl, _, _ = _scan(right)
        if max(rgl.values(), default=0) >= 2:
            gl_col = max(right, key=lambda c: rgl[c])
    if gl_col == desc_col:
        gl_col = None

    rows, cur_section = [], ""
    for r in range(hdr_row + 1, ws.max_row + 1):
        gl_raw = _s(ws.cell(r, gl_col).value) if gl_col else ""
        gl = gl_raw if _is_glnum(gl_raw) else ""
        name = _s(ws.cell(r, desc_col).value)
        if not gl and not name and gl_raw:
            name = gl_raw                 # section/subtotal label sitting in the gl column
        vals = [_num(ws.cell(r, c).value) for c in month_cols]
        has_vals = any(abs(v) > 1e-9 for v in vals)
        if not gl and not name and not has_vals:
            continue

        if not gl:
            # no account number on the row -> section header (no values) or a
            # subtotal/total line (has values). Header text drives `section`.
            if has_vals:
                rows.append(T12Row("subtotal", name or "", cur_section,
                                   values=vals, total=sum(vals)))
            elif name:
                cur_section = name
                rows.append(T12Row("header", name, cur_section))
            continue

        # row carries a GL account number
        if not name:
            name = gl
        if not has_vals:
            # parent/rollup node with no values (hierarchical tree export) -> a
            # section header that provides categorization context for its children
            cur_section = name
            rows.append(T12Row("header", name, cur_section))
            continue
        if _is_rollup(name, gl):
            # a subtotal that happens to carry an account number (Yardi tree: the
            # '-098/-099' rollup accounts, or 'Total …' / 'Net …' lines). Exclude
            # from postable totals so children aren't double-counted.
            rows.append(T12Row("subtotal", name, cur_section, values=vals, total=sum(vals)))
            continue

        # real postable (leaf) line
        digits = re.sub(r"\D", "", gl)
        if _REV_SECTIONS.search(cur_section) or digits[:1] == "4":
            side = "rev"
        else:
            side = "exp"
        code = am.categorize_t12_line(name, side, cur_section, gl)
        # Reimbursement / rebill lines often sit on the EXPENSE side as a contra-expense
        # (a reimbursement that reduces a utility cost is booked negative; the rebilling
        # service fee the property pays is booked positive). When such a line is
        # reclassified to a REVENUE code (RWS/RT/RF), its sign must flip so it reads as
        # income — a negative contra-expense becomes positive RUBS revenue, and a
        # positive service-fee cost becomes a negative contra to that revenue. This
        # matches how RedIQ presents these lines.
        if side == "exp" and code in am.REVENUE_CODE_SET:
            vals = [-v for v in vals]
        rows.append(T12Row("line", name, cur_section, side, gl, code,
                           vals, sum(vals)))

    return T12(months=months, rows=rows, sheet_name=ws.title, title=title)


# ===========================================================================
# RENT ROLL
# ===========================================================================
@dataclass
class RRUnit:
    unit: str
    floorplan: str
    sqft: float
    status: str
    resident: str
    market_rent: float
    charges: dict                 # charge_code -> scheduled $
    move_in: object = None
    lease_start: object = None
    lease_end: object = None
    expected_move_out: object = None

    # derived
    contract_rent: float = 0.0
    other_income: float = 0.0
    concessions: float = 0.0
    net_effective: float = 0.0
    lease_type: str = "Market"
    occupancy: str = ""


@dataclass
class RentRoll:
    units: list
    as_of: str = ""
    as_of_date: object = None                              # parsed as-of date (if available)
    charge_summary: dict = field(default_factory=dict)     # code -> (sched$, ycode_type)
    status_summary: dict = field(default_factory=dict)     # status -> (count, pct)
    footer_unit_mix: dict = field(default_factory=dict)    # plan -> (rentable, occ, avgmkt, avgsched)
    totals: dict = field(default_factory=dict)             # sf, market_rent, scheduled, actual


_UNIT_RE = re.compile(r"^[A-Za-z0-9]+-[A-Za-z]*\d")   # e.g. A-101, B-302, J-J3086 (digit after dash)


def _looks_like_unit(s):
    """A rent-roll unit id: dashed (A-101, 01-0111) or plain (1001, 101A, B204).
    Must contain a digit, be short, single-token, and not a 'Total' label. The unit
    loop additionally requires a numeric SQFT cell, which guards against false hits."""
    s = _s(s)
    if not s or len(s) > 12 or " " in s:
        return False
    if not re.search(r"\d", s) or s.endswith("Total:") or s.lower() in ("total", "unit"):
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9.\-]{0,11}", s))


def _occ_from_status(status: str) -> str:
    s = status.lower()
    if "excluded" in s or "model" in s or "down" in s:
        return "Non-Revenue"
    if "vacant" in s or ("notice" in s and "unrented" in s):
        return "Vacant"
    return "Occupied"


def parse_rent_roll(path: str, charge_lookup=None) -> RentRoll:
    charge_lookup = charge_lookup or {}
    wb = _load_workbook(path)
    # choose the sheet that actually holds the unit detail (most unit-id rows in col A)
    ws = max(wb.worksheets,
             key=lambda w: sum(1 for r in range(1, min(w.max_row, 500) + 1)
                               if _looks_like_unit(w.cell(r, 1).value)))

    # as-of date: an 'As Of = ...' label in the first few rows, else fall back
    as_of = ""
    for r in range(1, 9):
        for c in range(1, 5):
            m = re.search(r"as[ _]?of(?:\s+date)?\s*[=:]\s*(.+)", _s(ws.cell(r, c).value), re.I)
            if m:
                as_of = m.group(1).strip(); break
        if as_of:
            break
    if not as_of:
        as_of = _s(ws.cell(4, 1).value)

    # --- locate the detail header row (allowing a two-row header) ---
    def _hdr_score(r):
        labs = [_s(ws.cell(r, c).value).lower() for c in range(1, min(ws.max_column, 24) + 1)]
        has_unit = any(l == "unit" or l.startswith("bldg-unit") or l == "apt"
                       or (l.startswith("unit") and "type" not in l and "sq" not in l)
                       for l in labs[:4])
        keys = sum(1 for l in labs if any(k in l for k in
                   ("market", "charge", "amount", "move", "lease", "resident", "sq ft", "sqft")))
        return has_unit, keys
    hdr_row = None
    for r in range(1, min(ws.max_row, 30) + 1):
        hu, keys = _hdr_score(r)
        if hu and keys >= 2:
            hdr_row = r; break
    if hdr_row is None:
        hdr_row = 7

    # two-row header? the row below holds sub-labels (text, no unit id, no numbers)
    nxt = hdr_row + 1
    nxt_num = any(isinstance(ws.cell(nxt, c).value, (int, float)) for c in range(1, ws.max_column + 1))
    nxt_unit = _looks_like_unit(ws.cell(nxt, 1).value)
    nxt_txt = sum(1 for c in range(1, ws.max_column + 1) if re.search(r"[A-Za-z]", _s(ws.cell(nxt, c).value)))
    two_row = (not nxt_num) and (not nxt_unit) and nxt_txt >= 2

    labels = {}
    for c in range(1, ws.max_column + 1):
        a = _s(ws.cell(hdr_row, c).value)
        b = _s(ws.cell(nxt, c).value) if two_row else ""
        lab = (a + " " + b).strip().lower()
        if lab:
            labels[c] = lab

    def col(*names, exclude=()):
        for n in names:
            for c, lab in labels.items():
                if n in lab and not any(x in lab for x in exclude):
                    return c
        return None

    c_unit     = (col("bldg-unit", "unit number", "unit no")
                  or col("unit", exclude=("type", "sq", "status")) or 1)
    c_unittype = col("unit type", "floor plan", "floorplan", "plan type", "plan")
    c_sqft     = col("sq ft", "sqft", "square")
    c_status   = col("unit status", "lease status", "status")
    c_res      = col("resident name", "tenant name", "name", "resident", "tenant",
                     exclude=("unit", "floor"))
    c_movein   = col("move-in", "move in")
    c_lstart   = col("lease start", "lease from", "lease begin")
    c_lend     = col("lease end", "lease expiration", "lease to", "expiration")
    c_emo      = col("expected move-out", "expected move out", "move-out", "move out")
    c_market   = col("market rent", "market")
    c_ledger   = col("ledger")
    c_charge   = col("charge code", "charge")
    c_sched    = col("scheduled charge", "scheduled", "amount", "charge amt")
    c_actual   = col("actual charge", "actual")

    # --- boundary: detail ends at the first post-detail summary section ---
    detail_end = ws.max_row + 1
    for r in range(hdr_row + 1, ws.max_row + 1):
        a = _s(ws.cell(r, 1).value)
        d = _s(ws.cell(r, 4).value)
        if (a.startswith("Status Summary") or a.startswith("Average Charges")
                or a.startswith("Future Resident") or d.startswith("Charge Code Summ")):
            detail_end = r
            break

    def _resolve_charge(raw):
        """Map a raw charge-code cell to a human name via the optional lookup
        (numeric operator codes -> name); descriptive codes pass through."""
        raw = _s(raw)
        return charge_lookup[raw][0] if raw in charge_lookup else raw

    def _is_total(s):
        s = _s(s).lower()
        return s in ("total", "charge total:") or s.startswith("charge total")

    def _add_charge(unit, r):
        cc = _s(ws.cell(r, c_charge).value) if c_charge else ""
        if cc and not _is_total(cc) and not cc.endswith("Total:"):
            sched = _num(ws.cell(r, c_sched).value) if c_sched else 0.0
            nm = _resolve_charge(cc)
            unit.charges[nm] = unit.charges.get(nm, 0.0) + sched

    data_start = hdr_row + (2 if two_row else 1)

    # ---- FLAT / WIDE format (RealPage OneSite, etc.): ONE row per unit with charges
    #      in COLUMNS rather than charge-code sub-rows. Detected by the absence of a
    #      charge-code column. ----
    if not c_charge:
        core = {c_unit, c_unittype, c_sqft, c_status, c_res, c_movein, c_lstart,
                c_lend, c_emo, c_market, c_ledger}
        stop = re.compile(r"deposit|\bdep\b|on hand|balance|total|market|designation|status|"
                          r"name|move|lease|resh|sqft|sq ?ft|floor|^unit|addl|^id\b|\bid$", re.I)
        charge_cols = [(c, labels[c]) for c in sorted(labels)
                       if c not in core and not stop.search(labels[c])]
        units = []
        for r in range(data_start, ws.max_row + 1):
            a = _s(ws.cell(r, c_unit).value)
            if not _looks_like_unit(a):
                continue
            sqv = ws.cell(r, c_sqft).value if c_sqft else None
            if not (isinstance(sqv, (int, float)) or _s(ws.cell(r, c_status).value)):
                continue
            u = RRUnit(
                unit=a,
                floorplan=_s(ws.cell(r, c_unittype).value) if c_unittype else "",
                sqft=_num(sqv),
                status=_s(ws.cell(r, c_status).value) if c_status else "",
                resident=_s(ws.cell(r, c_res).value) if c_res else "",
                market_rent=_num(ws.cell(r, c_market).value) if c_market else 0.0,
                charges={},
                move_in=ws.cell(r, c_movein).value if c_movein else None,
                lease_start=ws.cell(r, c_lstart).value if c_lstart else None,
                lease_end=ws.cell(r, c_lend).value if c_lend else None,
                expected_move_out=ws.cell(r, c_emo).value if c_emo else None,
            )
            for c, hdr in charge_cols:
                amt = _num(ws.cell(r, c).value)
                if amt:
                    nm = _resolve_charge(hdr)
                    u.charges[nm] = u.charges.get(nm, 0.0) + amt
            units.append(u)
        # OneSite-style rolls list an upcoming lease (Applicant / Pending renewal /
        # Future) as a SECOND row for a unit, alongside the current physical record.
        # Collapse to one row per unit, keeping the physical-status row — it carries
        # market rent, whereas the future-lease row shows market 0. This matches the
        # way RedIQ (and the operator's own unit count) report distinct units.
        if len(units) != len({u.unit for u in units}):
            chosen = {}
            for u in units:
                prev = chosen.get(u.unit)
                if prev is None or (u.market_rent or 0) > (prev.market_rent or 0):
                    chosen[u.unit] = u
            units = list(chosen.values())
    else:
        # ---- BLOCK format (Yardi/CBREI): charge sub-rows beneath each unit row ----
        units, cur_plan = [], ""
        r = data_start
        cur = None
        while r < detail_end:
            a = _s(ws.cell(r, c_unit).value)
            if a.startswith("Unit Type:"):
                cur_plan = a.split(":", 1)[1].strip()
                r += 1
                continue
            sqft_val = ws.cell(r, c_sqft).value if c_sqft else None
            if _looks_like_unit(a) and isinstance(sqft_val, (int, float)):
                if cur:
                    units.append(cur)
                plan = cur_plan
                if c_unittype:
                    v = _s(ws.cell(r, c_unittype).value)
                    if v:
                        plan = v
                cur = RRUnit(
                    unit=a, floorplan=plan, sqft=_num(sqft_val),
                    status=_s(ws.cell(r, c_status).value) if c_status else "",
                    resident=_s(ws.cell(r, c_res).value) if c_res else "",
                    market_rent=_num(ws.cell(r, c_market).value) if c_market else 0.0,
                    charges={},
                    move_in=ws.cell(r, c_movein).value if c_movein else None,
                    lease_start=ws.cell(r, c_lstart).value if c_lstart else None,
                    lease_end=ws.cell(r, c_lend).value if c_lend else None,
                    expected_move_out=ws.cell(r, c_emo).value if c_emo else None,
                )
                _add_charge(cur, r)
                r += 1
                continue
            ledger = _s(ws.cell(r, c_ledger).value) if c_ledger else ""
            if (ledger.startswith("Charge Total")
                    or _s(ws.cell(r, c_market).value).startswith("Charge Total")
                    or (c_charge and _is_total(ws.cell(r, c_charge).value))):
                r += 1
                continue
            if cur is not None and c_charge:
                _add_charge(cur, r)
            r += 1
        if cur:
            units.append(cur)

    # --- derive per-unit economics from charges ---
    for u in units:
        contract = other = conc = 0.0
        for cc, amt in u.charges.items():
            code, is_contract, _rec = am.categorize_charge(cc)
            if is_contract:
                contract += amt
            elif code == "conc" or amt < 0 and re.search(r"concession", cc, re.I):
                conc += amt
            else:
                other += amt
        u.contract_rent = contract
        u.other_income = other
        u.concessions = conc
        u.net_effective = contract + conc   # conc is negative
        if u.status:
            u.occupancy = _occ_from_status(u.status)
        else:
            # no status column on this rent roll -> infer from economics:
            # a unit carrying no contract (base) rent is treated as vacant.
            u.occupancy = "Occupied" if contract > 0 else "Vacant"
        u.lease_type = "Market"

    # --- footer summaries (best-effort; used for validation) ---
    charge_summary, status_summary, footer_mix, totals = _parse_rr_footer(ws)

    # Backfill grand totals from per-unit data when the footer layout didn't yield
    # them (e.g. exports with an extra inline column shift the footer total row).
    # The footer is authoritative when present; this only fills gaps.
    if not totals.get("sqft"):
        totals["sqft"] = sum(u.sqft for u in units)
    if not totals.get("market_rent"):
        totals["market_rent"] = sum(u.market_rent for u in units)
    if not totals.get("scheduled"):
        totals["scheduled"] = sum(sum(u.charges.values()) for u in units)
    if not totals.get("actual"):
        totals["actual"] = totals.get("scheduled", 0.0)

    return RentRoll(units=units, as_of=as_of, as_of_date=_to_date(as_of),
                    charge_summary=charge_summary, status_summary=status_summary,
                    footer_unit_mix=footer_mix, totals=totals)


def _parse_rr_footer(ws) -> tuple:
    charge_summary, status_summary, footer_mix, totals = {}, {}, {}, {}
    mode = None
    last_total = None   # most recent '...Total:' row values (grand total = last before summary)
    for r in range(1, ws.max_row + 1):
        a = _s(ws.cell(r, 1).value)
        d = _s(ws.cell(r, 4).value)
        # everything from the Future-Resident section onward is supplementary -> stop
        if a.startswith("Future Resident"):
            break
        # track running '...Total:' rows while still in the detail block
        if mode is None and a.endswith("Total:") and not a.startswith("Unit Type") \
           and isinstance(ws.cell(r, 2).value, (int, float)):
            last_total = {
                "sqft": _num(ws.cell(r, 2).value),
                "market_rent": _num(ws.cell(r, 9).value),
                "actual": _num(ws.cell(r, 12).value),
                "scheduled": _num(ws.cell(r, 13).value),
            }
        if a.startswith("Status Summary") or d.startswith("Charge Code Summ"):
            if last_total and not totals:
                totals.update(last_total)   # freeze grand total (last before summary)
            mode = "summary"
            continue
        if a.startswith("Average Charges"):
            mode = "mix"
            continue
        if mode == "summary":
            # status summary in cols A/B/C
            if a and not a.startswith("Description") and ws.cell(r, 2).value is not None \
               and not a.startswith("Total"):
                status_summary[a] = (_num(ws.cell(r, 2).value), _num(ws.cell(r, 3).value))
            # charge-code summary in cols D/E/F
            dc = _s(ws.cell(r, 4).value)
            if dc and dc not in ("Charge Code", "Description") and not dc.startswith("Ledger") \
               and not dc.endswith("Total:"):
                charge_summary[dc] = (_num(ws.cell(r, 5).value), _s(ws.cell(r, 6).value))
        if mode == "mix":
            if a and not a.startswith("Unit Type") and not a.startswith("Average") \
               and ws.cell(r, 2).value is not None:
                footer_mix[a] = (_num(ws.cell(r, 2).value), _num(ws.cell(r, 3).value),
                                 _num(ws.cell(r, 4).value), _num(ws.cell(r, 5).value))
    return charge_summary, status_summary, footer_mix, totals


# ===========================================================================
# CHARGE-CODE LOOKUP  (decode rent rolls that bill by numeric operator code)
# ===========================================================================
def parse_charge_codes(path: Optional[str]) -> dict:
    """Optional lookup: operator charge/account code -> (name, type). Some rent rolls
    (e.g. CBREI/Yardi) list each charge by a numeric code rather than a name; this maps
    them so categorization works. Accepts any sheet with a header row exposing an
    account/code column and a name column (+ optional type)."""
    if not path:
        return {}
    try:
        wb = _load_workbook(path)
    except Exception:
        return {}
    out = {}
    for ws in wb.worksheets:
        hdr = None
        for r in range(1, min(ws.max_row, 15) + 1):
            labs = [_s(ws.cell(r, c).value).lower() for c in range(1, min(ws.max_column, 12) + 1)]
            if (any(l in ("name", "description", "charge name") for l in labs)
                    and any(("account" in l or l == "code" or "charge code" in l) for l in labs)):
                hdr = r; break
        if hdr is None:
            continue
        labels = {c: _s(ws.cell(hdr, c).value).lower() for c in range(1, ws.max_column + 1)}

        def find(*names, exact=False):
            for n in names:
                for c, l in labels.items():
                    if (l == n) if exact else (n in l):
                        return c
            return None
        c_code = find("account", "code", exact=True) or find("charge code", "account", "code")
        c_name = find("name", "description", exact=True) or find("name", "description")
        c_type = find("type")
        if not c_code or not c_name:
            continue
        for r in range(hdr + 1, ws.max_row + 1):
            code = _s(ws.cell(r, c_code).value)
            name = _s(ws.cell(r, c_name).value)
            if code and name:
                out[code] = (name, _s(ws.cell(r, c_type).value) if c_type else "")
    return out


# ===========================================================================
# HELLODATA
# ===========================================================================
@dataclass
class HelloData:
    rows: list                      # list[dict] (raw CSV rows, original headers)
    headers: list
    plan_bed_bath: dict             # normalized_plan -> (bed, bath, partial)


def _norm_plan(p: str) -> str:
    """Strip trailing 'NxM' bed/bath token and whitespace; keep base plan id."""
    p = _s(p)
    p = re.sub(r"\s+\d+\s*x\s*\d+\s*$", "", p, flags=re.I)   # 'A1 1x1' -> 'A1'
    return p.strip()


def parse_hellodata(path: str) -> Optional[HelloData]:
    try:
        with open(path, newline="", encoding="utf-8-sig") as f:
            rd = csv.DictReader(f)
            headers = rd.fieldnames or []
            rows = [dict(r) for r in rd]
    except FileNotFoundError:
        return None
    pbb = {}
    for row in rows:
        plan = _norm_plan(row.get("Floorplan", ""))
        bed = row.get("Bedrooms", "")
        bath = row.get("Bathrooms", "")
        partial = row.get("Partial Bathrooms", "")
        if plan and plan not in pbb:
            pbb[plan] = (bed, bath, partial)
    return HelloData(rows=rows, headers=headers, plan_bed_bath=pbb)


# ===========================================================================
# BED / BATH INFERENCE
# ===========================================================================
_LETTER_BED = {"S": 0, "A": 1, "B": 2, "C": 3, "D": 4, "E": 5}


def infer_bed_bath(plan: str, hd: Optional[HelloData]):
    """Return (bed, bath, source) for a floor-plan id."""
    base = _norm_plan(plan)
    # 1) HelloData exact / base match
    if hd:
        for key in (base, re.match(r"^[A-Za-z]+\d+", base).group(0) if re.match(r"^[A-Za-z]+\d+", base) else base):
            if key in hd.plan_bed_bath:
                bed, bath, _p = hd.plan_bed_bath[key]
                try:
                    return int(float(bed)), _bath_num(bath), "HelloData"
                except (ValueError, TypeError):
                    break
    # 2) explicit 'NxM' in the plan name
    m = re.search(r"(\d+)\s*x\s*(\d+)", plan, re.I)
    if m:
        return int(m.group(1)), int(m.group(2)), "plan-name"
    # 3) leading letter convention (A=1bd, B=2bd, C=3bd ...)
    m = re.match(r"^([A-Za-z])", base)
    if m:
        bed = _LETTER_BED.get(m.group(1).upper())
        if bed is not None:
            bath = 1 if bed <= 1 else 2
            return bed, bath, "inferred"
    return None, None, "unknown"


def _bath_num(b):
    try:
        f = float(b)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return b


# ===========================================================================
# UNIT MIX
# ===========================================================================
@dataclass
class UnitMixRow:
    plan: str
    bed: object
    bath: object
    units: int
    occ: int
    vac: int
    nonrev: int
    avg_sqft: float
    total_sqft: float
    avg_market: float
    avg_contract: float
    bedbath_source: str = ""
    hd_names: str = ""            # HelloData (website) floor-plan name(s) for this plan
    # lease-mix / true-market-rent signals
    new_count: int = 0
    renewal_count: int = 0
    new_last5_rents: list = field(default_factory=list)   # most-recent-first contract rents
    avg_new_last5: float = 0.0
    # HelloData trailing executed (off-market) rents, by base plan
    t90_ask: float = 0.0
    t90_eff: float = 0.0
    t90_n: int = 0
    t365_ask: float = 0.0
    t365_eff: float = 0.0
    t365_n: int = 0
    yoy_ask: object = None        # T90 vs year-ago T90 (% change) or None
    yoy_eff: object = None


def mandatory_fee_bundle(rr: RentRoll, occ_threshold: float = 0.8):
    """Detect the mandatory flat monthly fees that a property bills on (nearly) every
    unit — valet trash, pest, utility-billing admin, tech, etc. Property websites fold
    these into the headline "Total Monthly" price, which is what HelloData scrapes, so
    HelloData asking/effective is inflated by this bundle versus base/contract rent.

    Returns (total, [(charge_name, amount), ...]). A fee qualifies if it is NOT contract
    rent, appears on >= occ_threshold of occupied units, is positive, and is essentially
    a flat amount (one dominant value). NOTE: the rent roll only itemizes fees billed
    per-unit; website-only admin fees (e.g. a utility-billing fee embedded in RUBS) may
    not appear here — use --hd-fee-offset to set the exact website bundle when known."""
    from collections import Counter, defaultdict
    occ = [u for u in rr.units if u.contract_rent > 0]
    n = len(occ)
    if not n:
        return 0.0, []
    cnt, amts = Counter(), defaultdict(Counter)
    for u in occ:
        for cc, amt in u.charges.items():
            if am.categorize_charge(cc)[1]:           # skip contract (base/amenity) rent
                continue
            cnt[cc] += 1
            amts[cc][round(amt, 2)] += 1
    comps = []
    for cc, c in cnt.items():
        if c < occ_threshold * n:
            continue
        modal, mc = amts[cc].most_common(1)[0]
        if modal > 0 and mc >= 0.8 * c:               # flat (one dominant amount), positive
            comps.append((cc, modal))
    comps.sort(key=lambda x: -x[1])
    return round(sum(a for _, a in comps), 2), comps


def build_unit_mix(rr: RentRoll, hd: Optional[HelloData],
                   recent_new=None, t90=None, t365=None, t90_prior=None, hd_fee=0.0) -> list:
    ref_date = rr.as_of_date
    if recent_new is None:
        recent_new = recent_new_leases(rr, 5, ref_date)
    ref = _hd_ref_date(hd) if hd else None
    plan_of = _hd_plan_of(rr)                   # join HelloData to RR floor plans by unit #
    if t90 is None:
        t90 = hellodata_window(hd, ref, 90, 0, plan_of)
    if t365 is None:
        t365 = hellodata_window(hd, ref, 365, 0, plan_of)
    if t90_prior is None:                       # the 90-day window one year earlier
        t90_prior = hellodata_window(hd, ref, 455, 365, plan_of)

    # HelloData per-RR-plan metadata via the unit-number join: marketing floor-plan
    # name(s) and bed/bath (used when the rent-roll plan code can't infer them).
    hd_names, hd_bb = {}, {}
    if hd:
        from collections import Counter
        nm, bb = {}, {}
        for row in hd.rows:
            p = plan_of(row)
            nm.setdefault(p, Counter())[_s(row.get("Floorplan"))] += 1
            try:
                bb.setdefault(p, Counter())[(int(float(row.get("Bedrooms"))),
                                             _bath_num(row.get("Bathrooms")))] += 1
            except (ValueError, TypeError):
                pass
        hd_names = {p: ", ".join(n for n, _ in c.most_common(2) if n) for p, c in nm.items()}
        hd_bb = {p: c.most_common(1)[0][0] for p, c in bb.items() if c}

    groups = {}
    for u in rr.units:
        groups.setdefault(u.floorplan, []).append(u)
    mix = []
    for plan, us in groups.items():
        occ = sum(1 for u in us if u.occupancy == "Occupied")
        vac = sum(1 for u in us if u.occupancy == "Vacant")
        nonrev = sum(1 for u in us if u.occupancy == "Non-Revenue")
        sqfts = [u.sqft for u in us if u.sqft]
        markets = [u.market_rent for u in us if u.market_rent]
        contracts = [u.contract_rent for u in us if u.contract_rent > 0]
        bed, bath, src = infer_bed_bath(plan, hd)
        if bed is None and plan in hd_bb:        # fall back to HD bed/bath via unit join
            bed, bath, src = hd_bb[plan][0], hd_bb[plan][1], "HelloData(unit)"
        nc = sum(1 for u in us if classify_lease(u, ref_date) == "new")
        rc = sum(1 for u in us if classify_lease(u, ref_date) == "renewal")
        nl = recent_new.get(plan, [])
        nl_rents = [round(u.contract_rent) for u in nl if u.contract_rent > 0]
        avg_nl = (sum(nl_rents) / len(nl_rents)) if nl_rents else 0.0
        bp = _base_plan(plan)
        t = t90.get(plan) or t90.get(bp, {})     # RR-plan key (unit join) else base name
        t3 = t365.get(plan) or t365.get(bp, {})
        tp = t90_prior.get(plan) or t90_prior.get(bp, {})

        def _net(d, k):                          # HelloData rent net of bundled mandatory fees
            v = d.get(k, 0.0)
            return max(0.0, v - hd_fee) if v else 0.0

        def _yoy(cur, prior):
            return (cur / prior - 1) if (cur and prior) else None
        mix.append(UnitMixRow(
            plan=plan, bed=bed, bath=bath, units=len(us), occ=occ, vac=vac,
            nonrev=nonrev,
            avg_sqft=(sum(sqfts) / len(sqfts)) if sqfts else 0.0,
            total_sqft=sum(u.sqft for u in us),
            avg_market=(sum(markets) / len(markets)) if markets else 0.0,
            avg_contract=(sum(contracts) / len(contracts)) if contracts else 0.0,
            bedbath_source=src, hd_names=hd_names.get(plan, ""),
            new_count=nc, renewal_count=rc,
            new_last5_rents=nl_rents, avg_new_last5=avg_nl,
            t90_ask=_net(t, "ask"), t90_eff=_net(t, "eff"), t90_n=t.get("n", 0),
            t365_ask=_net(t3, "ask"), t365_eff=_net(t3, "eff"), t365_n=t3.get("n", 0),
            yoy_ask=_yoy(_net(t, "ask"), _net(tp, "ask")),
            yoy_eff=_yoy(_net(t, "eff"), _net(tp, "eff")),
        ))
    mix.sort(key=lambda m: ((m.bed if isinstance(m.bed, int) else 99), m.plan))
    return mix


# ===========================================================================
# LEASE CLASSIFICATION + NEW-LEASE / HELLODATA MARKET-RENT SIGNALS
# ===========================================================================
def _to_date(v):
    if v is None or v == "":
        return None
    if hasattr(v, "date"):
        try:
            return v.date()
        except Exception:
            return None
    if isinstance(v, _dt.date):
        return v
    s = _s(v)
    for f in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%m/%d/%y"):
        try:
            return _dt.datetime.strptime(s[:19], f).date()
        except Exception:
            pass
    return None


_NEW_LEASE_MAXAGE = 400          # ~13 months: a move-in newer than this (when lease
#                                  start is absent) is treated as still on its original
#                                  (new) lease; older implies a likely renewal.


def classify_lease(u: RRUnit, ref_date=None) -> str:
    """Classify a unit's CURRENT lease:
      * both dates present  -> new if lease start <= move-in (first lease),
                               renewal if move-in is older than lease start.
      * only move-in (no lease start) -> new if the move-in is within ~12 months of
                               `ref_date` (rent-roll as-of), else renewal (likely
                               renewed at least once). This is the best we can do
                               without a lease-start date.
      * otherwise -> unknown."""
    ls, mi = _to_date(u.lease_start), _to_date(u.move_in)
    if ls and mi:
        return "renewal" if ls > mi else "new"
    if mi and not ls and ref_date:
        return "new" if (ref_date - mi).days <= _NEW_LEASE_MAXAGE else "renewal"
    return "unknown"


def _lease_date(u: RRUnit):
    """The date that best dates the current lease for trend/recency: lease start if
    present, else move-in (used when the rent roll carries only move-in)."""
    return _to_date(u.lease_start) or _to_date(u.move_in)


def _base_plan(p: str) -> str:
    """Reduce a rent-roll plan id to the HelloData base plan: 'B1.1'->'B1', 'B2-A'->'B2',
    'A1 1x1'->'A1'."""
    p = _s(p)
    m = re.match(r"([A-Za-z]+\s*\d+)", p)
    return re.sub(r"\s+", "", m.group(1)).upper() if m else p.upper()


def _unit_plan_map(rr: RentRoll) -> dict:
    """{unit-id -> floorplan} from the rent roll, for joining HelloData by unit number."""
    return {_s(u.unit): u.floorplan for u in rr.units if _s(u.unit)}


def _hd_plan_of(rr: Optional[RentRoll]):
    """Return a function HD-row -> plan key. Prefer joining HelloData to the rent roll by
    UNIT NUMBER (so HD's marketing floor-plan names, which need not match the rent roll's
    internal codes, resolve to the rent roll's floor plan); fall back to the HD floor-plan
    base name when no unit match (e.g. Canyon Ridge, where the names already align)."""
    umap = _unit_plan_map(rr) if rr else {}

    def plan_of(row):
        p = umap.get(_s(row.get("Unit")))
        return p if p else _base_plan(row.get("Floorplan", ""))
    return plan_of


def recent_new_leases(rr: RentRoll, n: int = 5, ref_date=None) -> dict:
    """{floorplan -> [RRUnit, ...]} the n most-recent NEW leases per plan."""
    ref_date = ref_date or rr.as_of_date
    by_plan = {}
    for u in rr.units:
        if classify_lease(u, ref_date) != "new" or not _lease_date(u):
            continue
        by_plan.setdefault(u.floorplan, []).append(u)
    out = {}
    for plan, us in by_plan.items():
        us.sort(key=lambda u: _lease_date(u), reverse=True)
        out[plan] = us[:n]
    return out


def new_lease_t90(rr: RentRoll, days: int = 90, ref_date=None) -> float:
    """Avg contract rent of NEW leases dated in the trailing `days`, anchored on the most
    recent new lease. The preferred 'true market rent' read alongside HelloData."""
    ref_date = ref_date or rr.as_of_date
    pts = []
    for u in rr.units:
        if classify_lease(u, ref_date) == "new":
            d = _lease_date(u)
            if d and u.contract_rent > 0:
                pts.append((d, u.contract_rent))
    if not pts:
        return 0.0
    anchor = max(d for d, _ in pts)
    sel = [c for d, c in pts if (anchor - d).days <= days]
    return sum(sel) / len(sel) if sel else 0.0


def _hd_dates(hd: HelloData):
    return [d for d in (_to_date(r.get("Off Market Date")) for r in hd.rows) if d]


def _hd_ref_date(hd):
    ds = _hd_dates(hd)
    return max(ds) if ds else _dt.date.today()


def hellodata_window(hd, ref_date, older_days, newer_days=0, plan_of=None) -> dict:
    """{plan -> {'ask','eff','n'}} averaged over leases that went OFF market (executed)
    in the window [ref-older_days, ref-newer_days]. T90 = window(90,0); T365 =
    window(365,0); the year-ago T90 = window(455,365). `plan_of` maps an HD row to a
    plan key (defaults to the HD floor-plan base name)."""
    if hd is None:
        return {}
    if ref_date is None:
        ref_date = _hd_ref_date(hd)
    if plan_of is None:
        plan_of = lambda row: _base_plan(row.get("Floorplan", ""))
    acc = {}
    for r in hd.rows:
        od = _to_date(r.get("Off Market Date"))
        if not od:
            continue
        age = (ref_date - od).days
        if not (newer_days <= age <= older_days):
            continue
        bp = plan_of(r)
        d = acc.setdefault(bp, {"ask": [], "eff": []})
        for fld, key in (("Last Asking Rent", "ask"), ("Last Effective Rent", "eff")):
            try:
                v = float(r.get(fld) or 0)
                if v > 0:
                    d[key].append(v)
            except (ValueError, TypeError):
                pass
    return {bp: {"ask": (sum(v["ask"]) / len(v["ask"])) if v["ask"] else 0.0,
                 "eff": (sum(v["eff"]) / len(v["eff"])) if v["eff"] else 0.0,
                 "n": len(v["eff"])}
            for bp, v in acc.items()}


def hellodata_t90(hd, ref_date=None, window: int = 90) -> dict:
    return hellodata_window(hd, ref_date, window, 0)


# ===========================================================================
# RECONCILIATION  (rent roll  <->  T12, latest month)
# ===========================================================================
@dataclass
class ReconLine:
    label: str
    rr_value: float
    t12_value: float
    note: str = ""

    @property
    def variance(self):
        return self.rr_value - self.t12_value

    @property
    def pct(self):
        return (self.variance / self.t12_value) if self.t12_value else None


@dataclass
class Reconciliation:
    latest_month: str
    lines: list                 # list[ReconLine]
    flags: list                 # list[str]
    charge_map: list            # list[(charge_code, code, ytype, sched$, in_contract)]
    correlations: list = field(default_factory=list)   # list[(label, detail_str)]
    noi_tie: list = field(default_factory=list)        # list[dict] per statement (see reconcile_noi)
    hd_fee_netting: dict = field(default_factory=dict)  # HD asking gross→net disclosure (see build())
    charge_t12: list = field(default_factory=list)      # empirical charge→T12 placement (see match_charges_to_t12)


def reconcile(t12: T12, rr: RentRoll) -> Reconciliation:
    midx = t12.n_months - 1
    latest = t12.month_labels[midx] if t12.months else ""
    # trailing-12 (or all) monthly average per code — steadier than a single month
    n = t12.n_months
    last12 = list(range(max(0, n - 12), n))
    _k = len(last12) or 1

    def t12_avg(*codes):
        return sum(t12.code_total(c, i) for c in codes for i in last12) / _k

    # rent-roll scheduled charge totals by our code
    rr_market = sum(u.market_rent for u in rr.units)
    rr_contract = sum(u.contract_rent for u in rr.units)
    rr_other = sum(u.other_income for u in rr.units)
    amenity = sum(v for u in rr.units for cc, v in u.charges.items() if "amenity" in cc.lower())

    rentinc = t12.code_total("Rentinc", midx)
    ltl = t12.code_total("ltl", midx)
    vac = t12.code_total("vac", midx)
    nr = t12.code_total("nr", midx)
    agpr_t1 = rentinc + ltl                       # full-occupancy contract rent (AGPR)
    econ_contract = rentinc + ltl + vac + nr      # occupied contract rent

    # charge-code map (what's in contract rent vs other income)
    charge_map = []
    cc_totals = {}
    for u in rr.units:
        for cc, amt in u.charges.items():
            cc_totals[cc] = cc_totals.get(cc, 0.0) + amt
    for cc, amt in sorted(cc_totals.items(), key=lambda kv: -abs(kv[1])):
        code, is_contract, _rec = am.categorize_charge(cc)
        ytype = rr.charge_summary.get(cc, (None, ""))[1]
        charge_map.append((cc, code, ytype, amt, is_contract))

    mkt_note = ("RR Market column \u2194 T12 GPR \u2014 both are seller-set ASKING rents and are "
                "NOT a market-rent signal (either can be set at will). True market rent = latest "
                "new-lease contract rents + HelloData executed asking; see the unit mix and Lease "
                "Trend tabs. Shown here only as a data-integrity check that the two files describe "
                "the same property/period.")
    # Label adapts to the data: only call out "incl. amenity rent" when the rent roll
    # actually carries amenity rent that rolls into contract rent (it doesn't, e.g., at Aura,
    # where the only "amenity" item is a $10 amenity FEE booked as Other Income \u2014 not rent).
    has_amen = amenity > 0
    clabel = ("Contract Rent incl. amenity rent (occupied), monthly" if has_amen
              else "Contract Rent (occupied), monthly")
    cnote = ("RR contract (Rent + amenity rent) \u2194 T12 contract net of vacancy "
             "(Rentinc+LtL+Vac+NonRev)" if has_amen else
             "RR contract rent \u2194 T12 contract net of vacancy (Rentinc+LtL+Vac+NonRev)")
    agpr_note = "RR contract \u00d712 \u2194 T12 latest-month AGPR \u00d712 (Rentinc+LtL)."
    if has_amen:
        agpr_note += (" Amenity rent is IN contract rent \u2014 it is not a separate T12 line, "
                      "and including it ties RR contract closer to AGPR.")
    lines = [
        ReconLine("Gross Market Rent (asking), monthly", rr_market, rentinc, mkt_note),
        ReconLine(clabel, rr_contract, econ_contract, cnote),
        ReconLine("T1 AGPR, annualized", rr_contract * 12, agpr_t1 * 12, agpr_note),
        ReconLine("Other Income (recurring), monthly", rr_other,
                  t12.code_total("OI", midx) + t12.code_total("park", midx)
                  + t12.code_total("RF", midx) + t12.code_total("RT", midx)
                  + t12.code_total("RWS", midx) + t12.code_total("cable", midx),
                  "RR ancillary charges \u2194 T12 other-income codes"),
    ]

    flags = []
    # amenity-rent verification (the AGPR question)
    if amenity > 0 and agpr_t1:
        gap_with = rr_contract - agpr_t1
        gap_wo = rr_contract - amenity - agpr_t1
        flags.append(
            f"Amenity rent ${amenity:,.0f}/mo is mapped to Rental Income / contract rent "
            f"(it appears nowhere else on the T12). Including it narrows the gap to T1 AGPR "
            f"from ${abs(gap_wo):,.0f} to ${abs(gap_with):,.0f}/mo \u2014 confirming amenity "
            f"belongs in contract rent / AGPR.")
    # contract-rent gap (occupied basis)
    cl = lines[1]
    if cl.t12_value and abs(cl.variance) / abs(cl.t12_value) > 0.03:
        flags.append(
            f"Contract rent (RR ${cl.rr_value:,.0f}) vs T12 occupied contract "
            f"(${cl.t12_value:,.0f}) differ by ${cl.variance:,.0f} "
            f"({cl.pct*100:+.1f}%) -- check vacant-unit treatment / amenity timing.")
    # NOTE: the RR-vs-T12 market-rent (asking) difference is intentionally NOT raised as a
    # flag. Both sides are seller-set asking rents, not a market-rent signal — flagging the
    # gap implies it is something to reconcile, which misdirects. The true market-rent reads
    # live in the unit mix (new-lease + HelloData executed) and Lease Trend tabs; the
    # informational note on the "Gross Market Rent (asking)" line above explains this.
    # RUBS / trash recovery charges with no clean T12 line
    for cc, code, ytype, amt, isc in charge_map:
        if re.search(r"valet|trash", cc, re.I) and code in ("RT", "cont"):
            flags.append(
                f"'{cc}' (${amt:,.0f}/mo, RUBS-style recovery) -- confirm whether it "
                f"nets against the trash expense or books as Other Income (RT).")
            break

    # ---- correlated cross-checks: rent-roll counts/charges vs T12 income lines ----
    correlations = []

    # Parking — units billed for parking on the RR vs T12 parking income (T12 avg/mo)
    park_units, rr_park = 0, 0.0
    for u in rr.units:
        upark = sum(amt for cc, amt in u.charges.items()
                    if am.categorize_charge(cc)[0] == "park")
        if upark > 0:
            park_units += 1
            rr_park += upark
    t12_park = t12_avg("park")
    if park_units or t12_park:
        per = (rr_park / park_units) if park_units else 0
        cap = (t12_park / rr_park * 100) if rr_park else None
        correlations.append((
            "Parking",
            f"RR: {park_units} units billed parking, ${rr_park:,.0f}/mo "
            f"(avg ${per:,.0f}/space)  ↔  T12 parking income ${t12_park:,.0f}/mo"
            + (f"  —  {cap:.0f}% of RR-scheduled parking appears on the T12."
               if cap is not None else "")))

    # Utility recapture (RUBS / billbacks) — what % of utility expense is recovered
    util_exp = t12_avg("UWS", "UC", "UF", "UT")
    rubs_inc = t12_avg("RWS", "RT", "RF")
    ws_exp, ws_rec = t12_avg("UWS"), t12_avg("RWS")
    if util_exp:
        detail = (f"T12 utility expense ${util_exp:,.0f}/mo (UWS+UC+UF+UT)  ↔  RUBS/billback "
                  f"income ${rubs_inc:,.0f}/mo (RWS+RT+RF)  =  {rubs_inc/util_exp*100:.0f}% recaptured")
        if ws_rec > 0 and ws_exp:
            detail += f"  ·  water/sewer alone {ws_rec/ws_exp*100:.0f}% (${ws_rec:,.0f}/${ws_exp:,.0f})"
        elif rubs_inc > 0:
            detail += "  ·  operator books reimbursement in a single line (RF), not split by utility"
        correlations.append(("Utility recapture (RUBS)", detail))

    return Reconciliation(latest_month=latest, lines=lines, flags=flags,
                          charge_map=charge_map, correlations=correlations)


def reconcile_noi(st):
    """Tie the standardized NOI back to each source statement's OWN reported
    'Net Operating Income' subtotal. The RUBS gross-up (reimbursements moved from a
    contra-expense to revenue) nets to zero at NOI, so a tie confirms the
    categorization kept every operating line on the right side of the NOI boundary;
    a gap flags a real difference (e.g. items pushed below NOI, or no clean NOI line
    in the source). Returns one dict per stitched statement."""
    rev_codes = am.REVENUE_CODE_SET
    exp_codes = {c for c, _ in am.EXPENSE_CODES}
    out = []
    for f in st.files:
        rep = {"noi": None, "rev": None, "opex": None}
        for row in f.t12.rows:
            if row.rtype != "subtotal" or not row.name:
                continue
            nm = row.name.upper()
            val = row.total if row.total is not None else sum(row.values)
            if "NET OPERATING INCOME" in nm:                        rep["noi"] = val
            elif "TOTAL REVENUE" in nm or "EFFECTIVE GROSS" in nm:   rep["rev"] = val
            elif "TOTAL OPERATING EXPENSE" in nm:                    rep["opex"] = val
        comp_rev = sum(sum(r.values) for r in f.t12.rows
                       if r.rtype == "line" and r.code in rev_codes)
        comp_opex = sum(sum(r.values) for r in f.t12.rows
                        if r.rtype == "line" and r.code in exp_codes)
        out.append({"label": f.label, "rep_noi": rep["noi"], "comp_noi": comp_rev - comp_opex,
                    "rep_rev": rep["rev"], "comp_rev": comp_rev,
                    "rep_opex": rep["opex"], "comp_opex": comp_opex})
    return out


def match_charges_to_t12(st, rr, tol=0.25):
    """Empirically determine where each rent-roll charge LANDS on the T12 — by matching the
    charge's monthly $ (and, where the code is mnemonic, its name) to an actual T12 line —
    and classify it from WHERE IT TIES, not from the charge name alone.

    Buckets: 'contract' (ties to a Rental Income / Potential Rent line) | 'other_income'
    (ties to an Other Income line) | 'rubs_recovery' (ties to a utility/RUBS rebill, which
    operators book on the EXPENSE side as a contra-expense — so we search both sides) |
    'unmatched' (no T12 line within tolerance — flag it). The base contract-rent charge is
    contract by definition (it ties to the whole Rental Income section, not one line).

    Each result carries the matched T12 line, the empirical bucket, the name-heuristic
    verdict (categorize_charge), and whether the two AGREE — disagreements are surfaced on
    the Reconciliation tab so the contract-rent determination can be audited."""
    if not st.files:
        return []
    latest = max(st.files, key=lambda f: (f.monthkeys[-1] if f.monthkeys else (0, 0)))
    t12 = latest.t12
    n = t12.n_months
    last12 = list(range(max(0, n - 12), n)) or [0]
    # Candidates = where a RESIDENT charge could legitimately land: any revenue line, or a
    # genuine rebill/recovery on the expense side (RUBS — booked as a contra-expense). Plain
    # expense lines (Sewer, Alarms, Telephone, …) are costs, not resident charges, so they
    # are excluded — otherwise small charges find coincidental same-magnitude expense matches.
    _REC_CODES = {"RF", "RT", "RWS", "RUBS"}
    _REC_RE = re.compile(r"rebill|recover|recaptur|reimburs|rubs", re.I)
    cand = []
    for row in t12.rows:
        if row.rtype != "line":
            continue
        vals = [row.values[i] for i in last12 if i < len(row.values)]
        avg = sum(vals) / len(vals) if vals else 0.0
        if abs(avg) < 1:
            continue
        is_recovery = (row.code in _REC_CODES) or bool(_REC_RE.search(row.name))
        if row.side == "rev":
            kind = "rev"
        elif row.side == "exp" and is_recovery:
            kind = "rec"
        else:
            continue
        cand.append({"name": row.name, "section": row.section or "", "side": row.side,
                     "kind": kind, "code": row.code, "avg": avg})

    tot = {}
    for u in rr.units:
        for cc, amt in u.charges.items():
            tot[cc] = tot.get(cc, 0.0) + amt
    base_cc, base_amt = None, 0.0
    for cc, amt in tot.items():
        _code, isc, _r = am.categorize_charge(cc)
        if isc and amt > base_amt:
            base_cc, base_amt = cc, amt

    def name_sim(cc, name):
        c = re.sub(r"[^a-z]", "", cc.lower())
        best = 0
        for w in re.findall(r"[a-z]+", name.lower()):
            k = 0
            for a, b in zip(c, w):
                if a == b:
                    k += 1
                else:
                    break
            if k >= 3:
                best = max(best, k)
        return best

    _REC_CODE_SET = {"RF", "RT", "RWS", "RUBS"}

    def heuristic_bucket(code, isc):
        if isc:
            return "contract"
        return "rubs_recovery" if code in _REC_CODE_SET else "other_income"

    rows = []
    for cc, amt in sorted(tot.items(), key=lambda kv: -abs(kv[1])):
        if abs(amt) < 5:          # ignore charge-code list entries with no current $
            continue
        code, isc, _r = am.categorize_charge(cc)
        if cc == base_cc:
            rows.append({"cc": cc, "amt": amt, "code": code, "match": None,
                         "bucket": "contract", "conf": "high", "name_contract": True,
                         "agrees": True,
                         "note": "Base contract rent — ties to the T12 Rental Income section."})
            continue
        scored = []
        for c in cand:
            denom = max(abs(amt), abs(c["avg"]), 1.0)
            # Revenue lines must match by SIGN (a positive resident charge is not a negative
            # contra-revenue adjustment like Vacancy/Models). Recoveries may be booked +/-,
            # so they match on magnitude.
            rel = (abs(abs(amt) - abs(c["avg"])) if c["kind"] == "rec"
                   else abs(amt - c["avg"])) / denom
            sim = name_sim(cc, c["name"])
            scored.append((rel - 0.10 * sim, rel, sim, c))
        scored.sort(key=lambda x: x[0])
        best = scored[0] if scored else None
        # A match only OVERRIDES the name categorization when the evidence is STRONG: a very
        # tight amount tie, or a decent amount tie corroborated by the line name. A merely
        # coincidental same-magnitude match (no name support) is NOT enough to flip a charge —
        # a charge folded into Rental Income (e.g. amenity rent) has no distinct T12 line and
        # would otherwise grab an unrelated OI line of similar size. In the weak/no-match case
        # we DEFER to the categorization rather than contradict it.
        strong = bool(best) and best[1] <= tol and (best[1] <= 0.05 or best[2] >= 3)
        if strong:
            _s, rel, sim, c = best
            sect = c["section"].upper()
            _CONTRA = {"conc", "ltl", "vac", "nr", "cl"}   # rental-income contras, NOT contract rent
            if c["kind"] == "rec":
                bucket = "rubs_recovery"
            elif c["code"] in _CONTRA or c["avg"] < 0:
                bucket = "rental_contra"   # vacancy / concessions / loss-to-lease / write-offs
            elif "OTHER INCOME" in sect or c["code"] in ("OI", "park", "cable"):
                bucket = "other_income"
            else:
                bucket = "contract"
            match, conf = c, ("high" if (rel <= 0.05 and sim >= 3) or rel <= 0.03 else "med")
        else:
            bucket = heuristic_bucket(code, isc)
            match, conf, rel = None, "low", None

        emp_contract = (bucket == "contract")
        agrees = (emp_contract == isc) if strong else None
        if strong and bucket == "rubs_recovery":
            where = f"T12 '{match['name']}'" if best[2] >= 3 else "the T12's RUBS/utility-rebill block"
            note = (f"Ties to {where} (~{match['avg']:,.0f}/mo, booked as a contra-expense) — "
                    f"a resident recovery, not contract rent.")
        elif strong and bucket == "rental_contra":
            what = match['name'] if best[2] >= 3 else "concession / credit / vacancy"
            note = (f"Ties to a rental-income contra (~{match['avg']:,.0f}/mo, {what}) — a credit "
                    f"that REDUCES rental income; it is NOT a component of contract rent.")
        elif strong:
            note = (f"Ties to T12 '{match['name']}' ({match['avg']:,.0f}/mo, "
                    f"{match['section'].title() or 'revenue'}); {rel*100:.0f}% from "
                    f"rent-roll {amt:,.0f}/mo.")
        elif isc:
            note = ("No distinct T12 line within tolerance — consistent with being folded into "
                    "Rental Income / contract rent (per the charge categorization).")
        else:
            note = (f"No T12 income/rebill line ties to {amt:,.0f}/mo — treated as "
                    f"{'a utility recovery' if bucket=='rubs_recovery' else 'other income / pass-through'} "
                    f"(per the charge categorization).")
        if agrees is False:
            note += (f"  ⚠ Categorization calls this {'contract rent' if isc else 'non-contract'}, "
                     f"but the T12 placement says {'contract rent' if emp_contract else 'NOT contract rent'} "
                     f"— review the Code column.")
        rows.append({"cc": cc, "amt": amt, "code": code, "match": match, "bucket": bucket,
                     "conf": conf, "name_contract": isc, "agrees": agrees, "note": note})
    return rows


# ===========================================================================
# TRENDS
# ===========================================================================
@dataclass
class Trends:
    months: list
    egr_by_month: list
    opex_by_month: list
    noi_by_month: list
    rentinc_by_month: list
    vac_by_month: list
    ltl_by_month: list
    # annualizations
    periods: dict               # 'T12'/'T6'/'T3' -> dict(code -> annualized $)
    notes: list


def _egr_codes():
    return [c for c, _ in am.REVENUE_CODES]


def _opex_codes():
    return [c for c, _ in am.EXPENSE_CODES]


def build_trends(t12, rr: RentRoll) -> Trends:
    n = t12.n_months
    labels = t12.month_labels
    ymkeys = [m[1] for m in t12.months]

    def month_sum(codes, i):
        return sum(t12.code_total(c, i) for c in codes)

    egr = [month_sum(_egr_codes(), i) for i in range(n)]
    opex = [month_sum(_opex_codes(), i) for i in range(n)]
    noi = [egr[i] - opex[i] for i in range(n)]
    rentinc = [t12.code_total("Rentinc", i) for i in range(n)]
    vac = [t12.code_total("vac", i) for i in range(n)]
    ltl = [t12.code_total("ltl", i) for i in range(n)]

    # trailing annualizations: T12 = sum, T6 = last6*2, T3 = last3*4
    def annualize(window):
        mult = 12 / window
        out = {}
        for c in am.ALL_CODES:
            s = sum(t12.code_total(c, i) for i in range(n - window, n))
            out[c] = s * mult
        out["_EGR"] = sum(egr[n - window:]) * mult
        out["_OPEX"] = sum(opex[n - window:]) * mult
        out["_NOI"] = sum(noi[n - window:]) * mult
        return out

    periods = {}
    if n >= 12:
        periods["T12"] = annualize(12)
    if n >= 6:
        periods["T6"] = annualize(6)
    if n >= 3:
        periods["T3"] = annualize(3)

    notes = []
    if ltl and ltl[0] != 0:
        chg = ltl[-1] - ltl[0]
        if abs(chg) > abs(ltl[0]) * 0.25:
            notes.append(
                f"Loss-to-Lease moved from ${ltl[0]:,.0f} ({labels[0]}) to "
                f"${ltl[-1]:,.0f} ({labels[-1]}) -- a ${chg:,.0f} swing; rents are "
                f"{'compressing vs market' if chg < 0 else 'catching up to market'}.")
    if "T3" in periods and "T12" in periods:
        e3, e12 = periods["T3"]["_EGR"], periods["T12"]["_EGR"]
        if e12:
            d = (e3 - e12) / e12 * 100
            notes.append(f"Annualized EGR: T3 ${e3:,.0f} vs T12 ${e12:,.0f} ({d:+.1f}%) "
                         f"-- recent run-rate is {'ahead of' if d>0 else 'below'} the trailing year.")
        n3, n12 = periods["T3"]["_NOI"], periods["T12"]["_NOI"]
        if n12:
            d = (n3 - n12) / n12 * 100
            notes.append(f"Annualized NOI: T3 ${n3:,.0f} vs T12 ${n12:,.0f} ({d:+.1f}%).")

    # year-over-year on any overlapping calendar months (only possible when the
    # stitched history spans >12 months)
    idx = {ym: i for i, ym in enumerate(ymkeys)}
    yoy = []
    for i, (y, m) in enumerate(ymkeys):
        j = idx.get((y - 1, m))
        if j is not None and egr[j]:
            yoy.append((labels[i], (egr[i] - egr[j]) / egr[j] * 100))
    if yoy:
        avg = sum(p for _, p in yoy) / len(yoy)
        pairs = ", ".join(f"{lab} {p:+.0f}%" for lab, p in yoy)
        notes.append(f"Year-over-year EGR (same calendar month): {pairs} (avg {avg:+.1f}%). "
                     f"On a lease-up this reflects occupancy ramp more than rent growth — "
                     f"see the Lease Trend tab for per-unit market-rent movement.")
    return Trends(months=labels, egr_by_month=egr, opex_by_month=opex,
                  noi_by_month=noi, rentinc_by_month=rentinc, vac_by_month=vac,
                  ltl_by_month=ltl, periods=periods, notes=notes)


# ===========================================================================
# MULTI-PERIOD STITCHING  (stitch several T12 / monthly statements together)
# ===========================================================================
import os as _os


def _stmt_label(path: str) -> str:
    b = _os.path.splitext(_os.path.basename(path))[0]
    m = re.search(r"(20\d{2})[._\-]?(\d{2})(?!\d)", b)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return b.replace("_", " ").strip()[:24]


@dataclass
class StmtFile:
    label: str
    path: str
    t12: T12
    monthkeys: list       # [(y,m), ...] in file order
    end: tuple            # (y,m) of last month
    detail: int           # distinct nonzero codes -> granularity score


class Stitched:
    """A virtual T12 spanning the UNION of months across several statement files.
    Quacks like T12 (month_labels / n_months / code_total) so trends & reconcile
    reuse it unchanged. Overlapping months are owned by the freshest extract
    (latest window end, then most detailed); each (code, month) is counted once."""

    def __init__(self, files):
        self.files = files
        self.title = files[0].t12.title if files else ""
        seen = {}
        for f in files:
            for (lab, ym) in f.t12.months:
                seen[ym] = lab
        self.monthkeys = sorted(seen.keys())
        self.months = [(seen[k], k) for k in self.monthkeys]
        maxdetail = max((f.detail for f in files), default=0)
        self.owner = []
        for k in self.monthkeys:
            cand = [i for i, f in enumerate(files) if k in f.monthkeys]
            self.owner.append(max(cand, key=lambda i: (files[i].end, files[i].detail)))
        self._cm = {}
        for mi, k in enumerate(self.monthkeys):
            f = files[self.owner[mi]]
            col = f.monthkeys.index(k)
            for r in f.t12.rows:
                if r.rtype == "line" and r.code:
                    val = r.values[col] if col < len(r.values) else 0.0
                    self._cm[(r.code, k)] = self._cm.get((r.code, k), 0.0) + val
        self.overlap_flags = self._overlap_flags()
        self.granularity_flags = self._granularity_flags(maxdetail)

    @property
    def month_labels(self):
        return [m[0] for m in self.months]

    @property
    def n_months(self):
        return len(self.monthkeys)

    def code_total(self, code, month_idx=None):
        if month_idx is None:
            return sum(v for (c, k), v in self._cm.items() if c == code)
        return self._cm.get((code, self.monthkeys[month_idx]), 0.0)

    def render_plan(self):
        """For the workbook: list of (file, [(union_col_idx, file_col_idx), ...]) covering
        ONLY the months each file owns. Writing blanks elsewhere prevents double counts
        so a SUMIFS over the whole tab reproduces the resolved series exactly."""
        plan = []
        for i, f in enumerate(self.files):
            owned = [(mi, f.monthkeys.index(k)) for mi, k in enumerate(self.monthkeys)
                     if self.owner[mi] == i and k in f.monthkeys]
            plan.append((f, owned))
        return plan

    def _egr_of(self, f, col):
        rev = {c for c, _ in REVENUE_CODES_M}
        return sum((r.values[col] if col < len(r.values) else 0)
                   for r in f.t12.rows if r.rtype == "line" and r.code in rev)

    def _overlap_flags(self):
        disagree = []
        for mi, k in enumerate(self.monthkeys):
            covering = [f for f in self.files if k in f.monthkeys]
            if len(covering) < 2:
                continue
            egrs = [self._egr_of(f, f.monthkeys.index(k)) for f in covering]
            if max(egrs) - min(egrs) > max(1.0, abs(max(egrs)) * 0.005):
                disagree.append(self.month_labels[mi])
        if disagree:
            return [f"Overlapping months disagree between uploaded statements "
                    f"({', '.join(disagree)}); the most recent extract was used. "
                    f"Check for restated financials."]
        return []

    def _granularity_flags(self, maxdetail):
        lo = [mi for mi in range(self.n_months)
              if self.files[self.owner[mi]].detail < maxdetail]
        if not lo:
            return []
        labs = [self.month_labels[mi] for mi in lo]
        span = f"{labs[0]}–{labs[-1]}" if len(labs) > 1 else labs[0]
        return [f"{span} sourced from a summary-level statement; sub-line detail "
                f"(concessions, loss-to-lease, RUBS, etc.) is not itemized for those "
                f"months and reads as 0 in those codes."]


# small alias so the helper above doesn't shadow the module constant import order
REVENUE_CODES_M = am.REVENUE_CODES


def stitch_statements(paths) -> Stitched:
    files = []
    for p in paths:
        t = parse_t12(p)
        nz = len({r.code for r in t.rows
                  if r.rtype == "line" and r.code and any(r.values)})
        files.append(StmtFile(_stmt_label(p), p, t,
                              [ym for _, ym in t.months], t.months[-1][1], nz))
    files.sort(key=lambda f: f.end)         # oldest extract first (stable display)
    return Stitched(files)


def _line_key(name):
    return re.sub(r"\s+", " ", _s(name)).lower()


def _side_of_code(code, fallback="unknown"):
    if any(code == c for c, _ in am.REVENUE_CODES):
        return "revenue"
    if any(code == c for c, _ in am.EXPENSE_CODES):
        return "expense"
    if any(code == c for c, _ in am.NONOP_REV_CODES):
        return "nonop_rev"
    if any(code == c for c, _ in am.NONOP_EXP_CODES):
        return "nonop_exp"
    return fallback


_SIDE_RANK = {"revenue": 0, "expense": 1, "nonop_rev": 2, "nonop_exp": 3, "unknown": 4}
_SIDE_LABEL = {"revenue": "INCOME", "expense": "OPERATING EXPENSES",
               "nonop_rev": "NON-OPERATING REVENUE", "nonop_exp": "NON-OPERATING EXPENSES",
               "unknown": "OTHER / UNCATEGORIZED"}


def unified_lines(st: Stitched):
    """Merge every statement's GL lines into ONE ordered set: each distinct (code, line)
    appears once, with values across the full union of months sourced from the owning
    (freshest) statement for each month. Same-named lines within a statement are summed,
    so a SUMIFS by code over the result reproduces the resolved per-code series exactly."""
    # first-appearance order (richest statement first). Lines are merged across
    # statements by GL account number when present (robust to differing descriptions
    # or name/number column layouts), else by normalized name.
    def _merge_key(row):
        return ("gl:" + row.acct) if getattr(row, "acct", "") else ("nm:" + _line_key(row.name))

    order, info = [], {}
    for f in sorted(st.files, key=lambda f: -f.detail):
        for row in f.t12.rows:
            if row.rtype != "line":
                continue
            key = _merge_key(row)
            if key not in info:
                side = _side_of_code(row.code,
                                     row.side if getattr(row, "side", "") in ("revenue", "expense") else "unknown")
                info[key] = {"name": _s(row.name), "code": row.code, "side": side,
                             "appearance": len(order)}
                order.append(key)
    # per-statement summed value vectors by key (handles duplicate-named lines)
    flines = []
    for f in st.files:
        nm = f.t12.n_months
        d = {}
        for row in f.t12.rows:
            if row.rtype != "line":
                continue
            key = _merge_key(row)
            arr = d.setdefault(key, [0.0] * nm)
            for j in range(nm):
                arr[j] += row.values[j] if j < len(row.values) else 0.0
        flines.append(d)
    # each unified line's monthly values (owner-sourced; blank where that month's owner lacks it)
    lines = []
    for key in order:
        meta = info[key]
        vals = []
        for mi, ym in enumerate(st.monthkeys):
            oi = st.owner[mi]
            arr = flines[oi].get(key)
            if arr is None:
                vals.append(None)
            else:
                col = st.files[oi].monthkeys.index(ym)
                vals.append(arr[col] if col < len(arr) else 0.0)
        lines.append({"type": "line", "code": meta["code"], "name": meta["name"],
                      "side": meta["side"], "values": vals, "appearance": meta["appearance"]})
    lines.sort(key=lambda L: (_SIDE_RANK[L["side"]], L["appearance"]))
    out, last = [], None
    for L in lines:
        if L["side"] != last:
            out.append({"type": "header", "side": L["side"], "name": _SIDE_LABEL[L["side"]]})
            last = L["side"]
        out.append(L)
    return out


# ===========================================================================
# LEASE TREND  /  SEASONALITY  (new-lease rents + HelloData executed, mix-weighted)
# ===========================================================================
@dataclass
class LeaseTrend:
    months: list                    # [(y,m), ...] sorted — full HelloData + new-lease history
    hd_ask: dict                    # (y,m) -> mix-weighted executed asking rent / unit
    hd_eff: dict                    # (y,m) -> mix-weighted executed effective rent / unit
    hd_conc: dict                   # (y,m) -> HelloData concession % (1 - eff/ask)
    hd_n: dict                      # (y,m) -> executed (off-market) lease count
    new_rent: dict                  # (y,m) -> avg NEW-lease contract rent (rent roll, by start month)
    new_n: dict                     # (y,m) -> new-lease count
    recent_new: dict                # plan -> [RRUnit] (last 5 new leases)
    plan_t90: dict                  # base_plan -> {'ask','eff','n'}  (trailing 90d)
    notes: list


def _quarter(d):
    return (d.year, (d.month - 1) // 3 + 1)


def build_lease_trend(rr: RentRoll, hd: Optional[HelloData], hd_fee: float = 0.0,
                      extra_rolls=None) -> LeaseTrend:
    ref_date = rr.as_of_date
    recent = recent_new_leases(rr, 5, ref_date)
    plan_of = _hd_plan_of(rr)        # HelloData -> RR floor plan via unit-number join
    umap = _unit_plan_map(rr)
    join_active = bool(hd) and sum(1 for r in hd.rows if umap.get(_s(r.get("Unit")))) \
        >= 0.3 * max(1, len(hd.rows))
    # weight key must match plan_of's output: RR floor plan when the unit join is live,
    # else the HelloData base-plan name (Canyon Ridge, where they already align)
    def rr_key(u):
        return u.floorplan if join_active else _base_plan(u.floorplan)
    weights = {}
    for u in rr.units:
        k = rr_key(u); weights[k] = weights.get(k, 0) + 1
    t90 = hellodata_window(hd, _hd_ref_date(hd) if hd else None, 90, 0, plan_of)

    months = set()
    hd_pm = {}                       # (y,m,plan) -> {'ask':[],'eff':[]}
    if hd:
        for r in hd.rows:
            od = _to_date(r.get("Off Market Date"))
            if not od:
                continue
            ym = (od.year, od.month); months.add(ym)
            bp = plan_of(r)
            cell = hd_pm.setdefault((ym, bp), {"ask": [], "eff": []})
            for fld, key in (("Last Asking Rent", "ask"), ("Last Effective Rent", "eff")):
                try:
                    v = float(r.get(fld) or 0)
                    if v > 0:
                        cell[key].append(v)
                except (ValueError, TypeError):
                    pass
    hd_ask, hd_eff, hd_conc, hd_n = {}, {}, {}, {}
    for ym in months:
        na = da = ne = de = 0.0; cnt = 0
        for bp, w in weights.items():
            cell = hd_pm.get((ym, bp))
            if not cell:
                continue
            if cell["ask"]:
                na += (sum(cell["ask"]) / len(cell["ask"])) * w; da += w
            if cell["eff"]:
                ne += (sum(cell["eff"]) / len(cell["eff"])) * w; de += w; cnt += len(cell["eff"])
        a = na / da if da else 0.0
        e = ne / de if de else 0.0
        hd_conc[ym] = (1 - e / a) if (a > 0 and e > 0) else 0.0   # ratio from GROSS rents
        hd_ask[ym] = max(0.0, a - hd_fee) if a else 0.0           # net of bundled fees
        hd_eff[ym] = max(0.0, e - hd_fee) if e else 0.0
        hd_n[ym] = cnt

    # New-lease contract-rent history. The current rent roll only carries leases still in
    # place, so its new-lease signings reach back ~12-18 months. Older rent rolls (passed as
    # extra_rolls) capture signings that have since turned over, extending this series back in
    # time — each roll is classified against ITS OWN as-of date, and a signing seen in more
    # than one roll (same unit + lease month) is counted once.
    new_pm, new_cnt = {}, {}          # (ym,plan) -> [contract rents] ; ym -> count
    seen_new = set()
    for roll in [rr] + list(extra_rolls or []):
        rdate = roll.as_of_date
        for u in roll.units:
            if classify_lease(u, rdate) != "new":
                continue
            d = _lease_date(u)
            if not d or u.contract_rent <= 0:
                continue
            ym = (d.year, d.month)
            if (u.unit, ym) in seen_new:
                continue
            seen_new.add((u.unit, ym))
            months.add(ym)
            new_pm.setdefault((ym, rr_key(u)), []).append(u.contract_rent)
            new_cnt[ym] = new_cnt.get(ym, 0) + 1
    new_rent, new_n = {}, {}
    for ym in {k[0] for k in new_pm}:
        num = den = 0.0                # mix-weight by floor-plan unit count (like the HD rows)
        for bp, w in weights.items():
            lst = new_pm.get((ym, bp))
            if lst:
                num += (sum(lst) / len(lst)) * w; den += w
        new_rent[ym] = num / den if den else 0.0
        new_n[ym] = new_cnt.get(ym, 0)

    months = sorted(months)
    notes = _trend_notes(months, hd_ask, hd_n)
    return LeaseTrend(months, hd_ask, hd_eff, hd_conc, hd_n, new_rent, new_n,
                      recent, t90, notes)


def _trend_notes(months, hd_ask, hd_n):
    import calendar
    notes = []
    bym = {}
    for ym in months:
        if hd_ask.get(ym, 0) > 0 and hd_n.get(ym, 0) > 0:
            bym.setdefault(ym[1], []).append(hd_ask[ym])
    if len(bym) >= 6:
        avg = {m: sum(v) / len(v) for m, v in bym.items()}
        hi, lo = max(avg, key=avg.get), min(avg, key=avg.get)
        spread = (avg[hi] / avg[lo] - 1) * 100 if avg[lo] else 0
        if spread > 2:
            notes.append(
                f"Seasonality: executed asking rents peak in {calendar.month_abbr[hi]} and "
                f"trough in {calendar.month_abbr[lo]} (~{spread:.0f}% spread across the HelloData history).")
    pairs = [(hd_ask[ym] / hd_ask[(ym[0] - 1, ym[1])] - 1) * 100
             for ym in months
             if hd_ask.get(ym, 0) > 0 and hd_ask.get((ym[0] - 1, ym[1]), 0) > 0]
    if pairs:
        notes.append(f"HelloData asking-rent YoY (same calendar month): {sum(pairs) / len(pairs):+.1f}% "
                     f"average across {len(pairs)} month-pair(s).")
    return notes
